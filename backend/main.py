"""
DrillOps — FastAPI Backend v3
Database: Supabase (PostgreSQL)
Multi-contractor: every query is filtered by contractor
"""

import re
import os
import base64
import json
import csv
from math import ceil
from contextlib import contextmanager
from functools import lru_cache
from io import BytesIO
from io import StringIO
from threading import BoundedSemaphore
from typing import Optional

import pdfplumber
import pandas as pd
import psycopg2
import psycopg2.extras
from psycopg2.pool import ThreadedConnectionPool
import httpx
from fastapi import FastAPI, UploadFile, File, HTTPException, Query, Form, Request
from fastapi.middleware.cors import CORSMiddleware

app = FastAPI(title="DrillOps API", version="3.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

DATABASE_URL = os.environ.get("DATABASE_URL")
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL environment variable is not set.")

# Keep the production model configurable so future model migrations do not
# require touching every Gemini-powered workflow.
GEMINI_MODEL = os.environ.get("GEMINI_MODEL", "gemini-3.5-flash").strip() or "gemini-3.5-flash"


def gemini_generate_content_url(api_key: str) -> str:
    return f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent?key={api_key}"


def gemini_response_text(result: dict) -> str:
    """Join every text part returned by Gemini into one response body."""
    parts = result["candidates"][0]["content"]["parts"]
    return "".join(
        part.get("text", "")
        for part in parts
        if isinstance(part, dict) and not part.get("thought", False)
    ).strip()

CONTRACTORS = [
    ("Allianz Drilling",   "ALZ"),
    ("Mitchells Drilling", "MIT"),
    ("MCC Group",          "MCC"),
    ("CHMS",               "CHMS"),
    ("King Konstruct",     "KK"),
    ("Weatherfords",       "WFD"),
    ("Epiroc",             "EPI"),
    ("Fortem",             "FOR"),
    ("GEOGAS",             "GEO"),
    ("Geochempet",         "GCP"),
    ("STANTEC",            "STN"),
    ("SIGRA",              "SIG"),
    ("Earth Works",        "EW"),
]

CONTRACTOR_CATEGORIES = ["Drilling", "Earthworks", "Labour", "Geological Support", "Lab Testing", "Misc"]
DEFAULT_CONTRACTOR_CATEGORIES = {
    "Allianz Drilling": "Drilling",
    "Mitchells Drilling": "Drilling",
    "MCC Group": "Labour",
    "CHMS": "Earthworks",
    "King Konstruct": "Earthworks",
    "Weatherfords": "Geological Support",
    "Epiroc": "Geological Support",
    "Fortem": "Geological Support",
    "GEOGAS": "Lab Testing",
    "Geochempet": "Lab Testing",
    "STANTEC": "Lab Testing",
    "SIGRA": "Lab Testing",
    "Earth Works": "Earthworks",
}

CONTRACTOR_EXPENSE_GL_LABELS = {
    "4200": "Consulting Services",
    "4250": "Contractors",
    "4350": "Drilling Services",
}


def default_contractor_expense_gl(category: str) -> str:
    """Return the legacy category-derived GL used to initialise contractors."""
    if category == "Drilling":
        return "4350"
    if category in ("Earthworks", "Labour"):
        return "4250"
    return "4200"


def contractor_gl_category(expense_gl: str) -> str:
    code = str(expense_gl or "").strip()
    label = CONTRACTOR_EXPENSE_GL_LABELS.get(code, "Assigned contractor GL")
    return f"{code} - {label}" if code else ""

CONTRACTOR_REFERENCE_TABLES = [
    "activities",
    "consumables",
    "crew",
    "imported_files",
    "drilling_rates",
    "hourly_rates",
    "consumable_rates",
    "boreholes",
    "purchase_orders",
    "source_files",
    "report_approvals",
    "activity_sheet_locks",
    "minimum_shift_topup_preferences",
    "projects",
    "project_budgets",
    "invoices",
    "invoice_lines",
    "invoice_imports",
]


DB_POOL_MIN = max(1, int(os.environ.get("DB_POOL_MIN", "1")))
DB_POOL_MAX = max(DB_POOL_MIN, int(os.environ.get("DB_POOL_MAX", "5")))
DB_POOL_ACQUIRE_TIMEOUT = max(1, int(os.environ.get("DB_POOL_ACQUIRE_TIMEOUT", "30")))

_db_pool = ThreadedConnectionPool(
    DB_POOL_MIN,
    DB_POOL_MAX,
    DATABASE_URL,
    cursor_factory=psycopg2.extras.RealDictCursor,
    connect_timeout=10,
    application_name="drillops-api",
    keepalives=1,
    keepalives_idle=30,
    keepalives_interval=10,
    keepalives_count=3,
)
_db_pool_slots = BoundedSemaphore(DB_POOL_MAX)


def _connection_failed(exc, conn):
    """Return True when an exception means the connection cannot be reused."""
    return bool(
        conn.closed
        or isinstance(exc, (psycopg2.InterfaceError, psycopg2.OperationalError))
        or (isinstance(exc, psycopg2.DatabaseError) and exc.pgcode is None)
    )


def _get_live_connection():
    """Check pooled connections before use and replace stale ones once."""
    last_error = None
    for _ in range(2):
        conn = _db_pool.getconn()
        try:
            if conn.closed:
                raise psycopg2.InterfaceError("pooled database connection is closed")
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
            conn.rollback()
            return conn
        except psycopg2.Error as exc:
            last_error = exc
            _db_pool.putconn(conn, close=True)

    raise last_error


@contextmanager
def get_conn():
    """Borrow a PostgreSQL connection and return it to the shared pool."""
    if not _db_pool_slots.acquire(timeout=DB_POOL_ACQUIRE_TIMEOUT):
        raise RuntimeError("Database connection pool is busy; please retry")

    conn = None
    discard = False
    try:
        conn = _get_live_connection()
        yield conn
        conn.commit()
    except Exception as exc:
        if conn is not None:
            discard = _connection_failed(exc, conn)
            if not discard:
                try:
                    conn.rollback()
                except psycopg2.Error:
                    discard = True
        raise
    finally:
        if conn is not None:
            discard = discard or bool(conn.closed)
            _db_pool.putconn(conn, close=discard)
        _db_pool_slots.release()


@app.on_event("shutdown")
def close_db_pool():
    _db_pool.closeall()


PCD_SMALL_LABEL = "PCD or Blade 99-125mm"
PCD_MEDIUM_LABEL = "PCD or Blade 125-175mm"
PCD_LARGE_LABEL = "PCD or Blade 175-305mm"
PQ_PQ3_LABEL = "PQ_PQ3"

LEGACY_DRILLING_BIT_LABELS = {
    "PCD_S": PCD_SMALL_LABEL,
    "PCD_M": PCD_MEDIUM_LABEL,
    "PCD_L": PCD_LARGE_LABEL,
}

ALLIANZ_CONTRACT_DRILLING_RATES = [
    (PCD_SMALL_LABEL, 0, 100, 46.00), (PCD_SMALL_LABEL, 100, 200, 51.00),
    (PCD_SMALL_LABEL, 200, 300, 59.00), (PCD_SMALL_LABEL, 300, 400, 66.00),
    (PCD_SMALL_LABEL, 400, 500, 70.00),
    (PCD_MEDIUM_LABEL, 0, 100, 51.00), (PCD_MEDIUM_LABEL, 100, 200, 62.00),
    (PCD_MEDIUM_LABEL, 200, 300, 76.00), (PCD_MEDIUM_LABEL, 300, 400, 81.00),
    (PCD_MEDIUM_LABEL, 400, 500, 95.00),
    (PCD_LARGE_LABEL, 0, 100, 84.58), (PCD_LARGE_LABEL, 100, 200, 94.58),
    (PCD_LARGE_LABEL, 200, 300, 100.50), (PCD_LARGE_LABEL, 300, 400, 111.00),
    (PCD_LARGE_LABEL, 400, 500, 126.00),
    ("HAMMER_S", 0, 200, 51.00), ("HAMMER_S", 200, 300, 61.00),
    ("HAMMER_S", 400, 500, 90.00),
    ("HAMMER_M", 0, 200, 60.00), ("HAMMER_M", 200, 300, 73.00),
    ("HAMMER_M", 400, 500, 103.00),
    ("HAMMER_L", 0, 200, 76.00), ("HAMMER_L", 200, 300, 86.00),
    ("HQ_HQ3", 0, 100, 196.00), ("HQ_HQ3", 100, 200, 239.00),
    ("HQ_HQ3", 200, 300, 278.00), ("HQ_HQ3", 300, 400, 324.00),
    ("HQ_HQ3", 400, 500, 367.00), ("HQ_HQ3", 500, 600, 422.00),
    ("4C", 0, 100, 227.00), ("4C", 100, 200, 273.00),
    ("4C", 200, 300, 327.00), ("4C", 300, 400, 374.00),
]


ALLIANZ_CONTRACT_HOURLY_RATES = [
    ("MOB", "Mobilisation: Initial mobilisation of standard drilling rig and ancillary equipment", 38760.00, "event"),
    ("DEMOB", "Demobilisation: Demobilisation of standard drilling rig and ancillary equipment from site", 38760.00, "event"),
    ("H_Active", "Active rate: drilling operations and essential related tasks", 745.00, "hour"),
    ("H_Inactive", "Inactive rate: on-site, ready-for-work non-operational delays", 675.00, "hour"),
    ("H_Min_Shift", "Minimum shift rate: 12 hours at active rate", 8940.00, "shift"),
    ("H_Standby_Without_Crew", "Standby without crew", 4470.00, "day"),
    ("H_Standby_NoCrew", "Standby without crew", 4470.00, "day"),
    ("H_Inspections", "Company initiated inspections of contractor plant and machinery", 675.00, "hour"),
    ("H_Training", "Training on site including personnel inductions and authorisations", 675.00, "hour"),
    ("H_Con_Collect_Plan", "Planned time spent picking up consumables from local stockpile before day starts", 675.00, "hour"),
    ("H_Con_Collect_Unplan", "Unplanned time spent picking up consumables from local stockpile", 0.00, "hour"),
    ("H_Water_Collect", "Time spent collecting water for operations", 675.00, "hour"),
    ("H_Truck_Vacuum", "Vacuum truck for desilting and dewatering drilling pits", 675.00, "hour"),
    ("H_Crew_Travel_On", "Crew travel time whilst on Company lease from site gate to drilling location", 675.00, "hour"),
    ("H_Crew_Travel_Off", "Crew travel time whilst off Company lease", 0.00, "hour"),
    ("H_Safety_Contractor", "Contractor initiated safety work charged in 15 minute increments", 675.00, "hour"),
    ("H_Safety_Company", "Company toolbox talks charged in 15 minute increments", 675.00, "hour"),
    ("H_Safety_Permits", "Company risk assessments, permits, SLAMs, JSAs and safety audits", 675.00, "hour"),
    ("H_Safety_Prestart", "Pre-start inspections of plant and equipment", 675.00, "hour"),
    ("H_Safety_Gas", "Standby due to gas detected from hole", 675.00, "hour"),
    ("H_Drilling", "Time spent drilling; covered under metreage charges and non chargeable", 0.00, "hour"),
    ("H_Water_Boring", "Rig operating rate for water bores", 745.00, "hour"),
    ("H_Tripping_Rods", "Tripping drilling rods in and out of a borehole", 745.00, "hour"),
    ("H_Surface_Setup", "Surface setup, T-piece, blooey line and blowout prevention setup", 745.00, "hour"),
    ("H_Casing_Install", "Installing casing", 745.00, "hour"),
    ("H_Casing_Retrieval", "Extracting casing from borehole", 745.00, "hour"),
    ("H_Change_Drill_Mthd", "Changing drilling method or bit types due to geological conditions", 745.00, "hour"),
    ("H_Reaming", "Reaming", 745.00, "hour"),
    ("H_Cleanouts", "Cleaning out boreholes for logging or blocked holes", 745.00, "hour"),
    ("H_Circulation_Flush", "Flushing borehole and attaining circulation to commence drilling", 745.00, "hour"),
    ("H_Circulation_Lost", "Regaining fluid circulation where circulation has been lost", 745.00, "hour"),
    ("H_Water_Flow_Measure", "Measuring and sampling water", 745.00, "hour"),
    ("H_Mud_Mixing", "Mixing drilling fluids", 745.00, "hour"),
    ("H_Fishing_Equipment", "Fishing or retrieving equipment out of a borehole", 745.00, "hour"),
    ("H_Fishing_Equipment_NC", "Fishing or retrieving equipment, non chargeable", 0.00, "hour"),
    ("H_Rig_Cementing", "Cementing boreholes using a drilling rig", 745.00, "hour"),
    ("H_Setup_Packup_Site", "Setting up or packing up on a drill site", 745.00, "hour"),
    ("H_Rig_Move", "Moving between drill sites including areas on a specific hub", 745.00, "hour"),
    ("H_Standby_AAC", "Standby due to Company instruction to stop work", 675.00, "hour"),
    ("H_Standby_Contractor", "Standby due to Contractor instruction to stop work", 675.00, "hour"),
    ("H_Standby_Cement_Set", "Standby waiting for grout to set", 675.00, "hour"),
    ("H_Standby_NoGasMon", "Standby due to gas monitor failure or unavailability", 0.00, "hour"),
    ("H_Standby_Grout", "Standby whilst grouting unit operating", 675.00, "hour"),
    ("H_Standby_Grouter", "Standby waiting for grouting unit", 675.00, "hour"),
    ("H_Standby_Water", "Standby waiting on water delivery", 675.00, "hour"),
    ("H_Standby_Blasting", "Standby for mine production blasting", 675.00, "hour"),
    ("H_Standby_Fatigue", "Standby for fatigue management including heat breaks and lunch breaks", 675.00, "hour"),
    ("H_Standby_Incidents", "Standby for incidents, investigations and ICAMS", 675.00, "hour"),
    ("H_Standby_Logger", "Standby waiting for geophysical logger to arrive", 675.00, "hour"),
    ("H_Standby_Logging", "Standby while geophysical logging is occurring", 675.00, "hour"),
    ("H_Standby_Mine_Shut", "Standby for mine shutdowns", 675.00, "hour"),
    ("H_Standby_Site_Insp", "Contractor work area inspection standby, non chargeable", 0.00, "hour"),
    ("H_Standby_Wet_Weath", "Standby for adverse weather with crew", 675.00, "hour"),
    ("H_Standby_Vac", "Standby waiting on vacuum truck to remove waste from sumps", 675.00, "hour"),
    ("H_Standby_Sumps", "Standby waiting on sumps to be delivered or dug deeper", 675.00, "hour"),
    ("H_Travel_Pitless", "Travel time for pitless system to and from site", 675.00, "hour"),
    ("H_Repairs", "Unplanned repairs of plant and equipment, non chargeable", 0.00, "hour"),
    ("H_Maintenance", "Scheduled maintenance of plant and equipment, non chargeable", 0.00, "hour"),
    ("H_Cement_Cart_Wait", "Cementing subcontractor waiting for drill", 675.00, "hour"),
    ("H_Cementing_Top_Up", "Grouting unit top ups", 745.00, "hour"),
    ("H_Rig_Testing", "Downhole testing for permeability while rig running", 745.00, "hour"),
    ("H_Standby_Testing", "Shutdown rig standby during permeability testing or piezometer installs", 675.00, "hour"),
    ("D_Accommodation", "Accommodation charge per person supplied by Company", 0.00, "day"),
    ("D_Backhoe", "Daily backhoe rate with operator", 1850.00, "day"),
    ("D_Water_Cart", "Daily water cart rate, 20,000L capacity", 1650.00, "day"),
    ("D_Backhoe_Standby", "Daily backhoe standby rate, 4 hour minimum", 620.00, "day"),
    ("D_Water_Cart_Standby", "Daily water cart standby rate, 4 hour minimum", 550.00, "day"),
]


MITCHELLS_CONTRACT_DRILLING_RATES = [
    (PCD_SMALL_LABEL, 0, 100, 48.00), (PCD_SMALL_LABEL, 100, 200, 52.00),
    (PCD_SMALL_LABEL, 200, 300, 62.00), (PCD_SMALL_LABEL, 300, 400, 67.00),
    (PCD_SMALL_LABEL, 400, 500, 71.00),
    (PCD_MEDIUM_LABEL, 0, 100, 52.00), (PCD_MEDIUM_LABEL, 100, 200, 62.00),
    (PCD_MEDIUM_LABEL, 200, 300, 71.00), (PCD_MEDIUM_LABEL, 300, 400, 81.00),
    (PCD_MEDIUM_LABEL, 400, 500, 90.00), (PCD_MEDIUM_LABEL, 500, 600, 100.00),
    (PCD_MEDIUM_LABEL, 600, 700, 109.00),
    (PCD_LARGE_LABEL, 0, 100, 62.00), (PCD_LARGE_LABEL, 100, 200, 71.00),
    (PCD_LARGE_LABEL, 200, 300, 81.00), (PCD_LARGE_LABEL, 300, 400, 90.00),
    (PCD_LARGE_LABEL, 400, 500, 100.00),
    ("HQ_HQ3", 0, 100, 190.00), ("HQ_HQ3", 100, 200, 238.00),
    ("HQ_HQ3", 200, 300, 285.00), ("HQ_HQ3", 300, 400, 333.00),
    ("HQ_HQ3", 400, 500, 380.00),
    (PQ_PQ3_LABEL, 0, 100, 238.00), (PQ_PQ3_LABEL, 100, 200, 285.00),
    (PQ_PQ3_LABEL, 200, 300, 333.00), (PQ_PQ3_LABEL, 300, 400, 380.00),
    (PQ_PQ3_LABEL, 400, 500, 428.00),
]

MITCHELLS_CONTRACT_HOURLY_RATE_OVERRIDES = {
    "MOB": (40000.00, "event"),
    "DEMOB": (40000.00, "event"),
    "H_Active": (725.00, "hour"),
    "H_Inactive": (650.00, "hour"),
    "H_Min_Shift": (7800.00, "shift"),
    "H_Standby_Without_Crew": (0.00, "day"),
    "H_Standby_NoCrew": (0.00, "day"),
    "D_Accommodation": (0.00, "item"),
    "D_Backhoe": (0.00, "item"),
    "D_Water_Cart": (0.00, "item"),
    "D_Backhoe_Standby": (0.00, "item"),
    "D_Water_Cart_Standby": (0.00, "item"),
}


def contract_drilling_rows(contractor: str, year: str):
    rows = []
    source = MITCHELLS_CONTRACT_DRILLING_RATES if contractor == "Mitchells Drilling" else ALLIANZ_CONTRACT_DRILLING_RATES
    for bit, depth_from, depth_to, rate in source:
        rows.append((contractor, year, bit, depth_from, depth_to, rate))
    return rows


def contract_hourly_rows(contractor: str, year: str):
    rows = []
    for code, desc, rate, unit in ALLIANZ_CONTRACT_HOURLY_RATES:
        if contractor == "Mitchells Drilling":
            if code in MITCHELLS_CONTRACT_HOURLY_RATE_OVERRIDES:
                rate, unit = MITCHELLS_CONTRACT_HOURLY_RATE_OVERRIDES[code]
            elif rate == 745.00:
                rate = 725.00
            elif rate == 675.00:
                rate = 650.00
            elif unit in {"day", "item"}:
                rate = 0.00
                unit = "item"
            if code == "H_Min_Shift":
                desc = "Minimum shift rate: 12 hours at active rate"
            elif code.startswith("D_") or code == "D_Accommodation":
                desc = f"{desc}; Mitchells schedule recharges are cost plus 10%"
        rows.append((contractor, year, code, desc, rate, unit))
    return rows


def migrate_legacy_drilling_bit_labels():
    """Rename old internal PCD bucket labels to the contract schedule labels."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            for old_label, new_label in LEGACY_DRILLING_BIT_LABELS.items():
                cur.execute(
                    "UPDATE drilling_rates SET bit_type=%s WHERE bit_type=%s",
                    (new_label, old_label),
                )
        conn.commit()


# ── Schema ────────────────────────────────────────────────────────────────────
def apply_mitchells_contract_exceptions():
    """Keep Mitchells contract details aligned for existing imported data."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE activities
                SET contractor='Mitchells Drilling'
                WHERE contractor <> 'Mitchells Drilling'
                  AND (
                    UPPER(COALESCE(drill_rig, '')) = 'IB652C'
                    OR source_file IN (
                        SELECT filename FROM source_files
                        WHERE file_type='coreplan_csv'
                    )
                  )
                """
            )
            cur.execute(
                """
                UPDATE consumables
                SET contractor='Mitchells Drilling'
                WHERE contractor <> 'Mitchells Drilling'
                  AND source_file IN (
                    SELECT filename FROM source_files
                    WHERE file_type='coreplan_csv'
                  )
                """
            )
            cur.execute(
                """
                UPDATE crew
                SET contractor='Mitchells Drilling'
                WHERE contractor <> 'Mitchells Drilling'
                  AND source_file IN (
                    SELECT filename FROM source_files
                    WHERE file_type='coreplan_csv'
                  )
                """
            )
            for bit, depth_from, depth_to, rate in MITCHELLS_CONTRACT_DRILLING_RATES:
                cur.execute(
                    """
                    UPDATE drilling_rates
                    SET rate=%s
                    WHERE contractor='Mitchells Drilling'
                      AND bit_type=%s
                      AND depth_from=%s
                      AND depth_to=%s
                    """,
                    (rate, bit, depth_from, depth_to),
                )
            cur.execute(
                """
                UPDATE activities
                SET code='H_Standby_Fatigue',
                    unit_rate=0,
                    line_cost=0,
                    rate_basis='not chargeable - Mitchells fatigue management'
                WHERE contractor='Mitchells Drilling'
                  AND (code='H_Standby_Fatigue' OR code='H_Standby_AAC')
                  AND COALESCE(notes, '') ILIKE '%fatigue%'
                """
            )
            cur.execute(
                """
                UPDATE source_files sf
                SET contractor='Mitchells Drilling'
                WHERE sf.contractor <> 'Mitchells Drilling'
                  AND sf.file_type='coreplan_csv'
                  AND NOT EXISTS (
                    SELECT 1 FROM source_files existing
                    WHERE existing.filename=sf.filename
                      AND existing.contractor='Mitchells Drilling'
                  )
                """
            )
            cur.execute(
                """
                UPDATE imported_files im
                SET contractor='Mitchells Drilling'
                WHERE im.contractor <> 'Mitchells Drilling'
                  AND im.filename IN (
                    SELECT filename FROM source_files
                    WHERE file_type='coreplan_csv'
                  )
                  AND NOT EXISTS (
                    SELECT 1 FROM imported_files existing
                    WHERE existing.filename=im.filename
                      AND existing.contractor='Mitchells Drilling'
                  )
                """
            )
        conn.commit()


def init_db():
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS activities (
                    id              SERIAL PRIMARY KEY,
                    source_file     TEXT,
                    contractor      TEXT DEFAULT 'Allianz Drilling',
                    date            TEXT,
                    hole_num        TEXT,
                    site_name       TEXT,
                    program         TEXT,
                    project         TEXT,
                    location        TEXT,
                    drill_rig       TEXT,
                    client          TEXT,
                    contract        TEXT,
                    shift           TEXT,
                    time_from       TEXT,
                    time_to         TEXT,
                    total_time      TEXT,
                    bit_type        TEXT,
                    diameter        TEXT,
                    metres_from     FLOAT,
                    metres_to       FLOAT,
                    total_metres    FLOAT,
                    code            TEXT,
                    notes           TEXT,
                    rate_year       TEXT,
                    unit_rate       FLOAT,
                    quantity        FLOAT,
                    line_cost       FLOAT,
                    rate_basis      TEXT,
                    po_id           INTEGER
                )
            """)
            # Safe migrations for existing tables
            for col, typedef in [
                ("contractor", "TEXT DEFAULT 'Allianz Drilling'"),
                ("program",    "TEXT"),
                ("project",    "TEXT"),
                ("po_id",      "INTEGER"),
                ("rate_year",  "TEXT"),
                ("unit_rate",  "FLOAT"),
                ("quantity",   "FLOAT"),
                ("line_cost",  "FLOAT"),
                ("rate_basis", "TEXT"),
            ]:
                try:
                    cur.execute(f"ALTER TABLE activities ADD COLUMN IF NOT EXISTS {col} {typedef}")
                except Exception:
                    conn.rollback()

            cur.execute("""
                CREATE TABLE IF NOT EXISTS consumables (
                    id          SERIAL PRIMARY KEY,
                    source_file TEXT,
                    contractor  TEXT DEFAULT 'Allianz Drilling',
                    date        TEXT,
                    hole_num    TEXT,
                    site_name   TEXT,
                    consumable  TEXT,
                    type        TEXT,
                    quantity    TEXT,
                    unit        TEXT,
                    unit_price  FLOAT,
                    line_cost   FLOAT
                )
            """)
            for col, typedef in [
                ("contractor", "TEXT DEFAULT 'Allianz Drilling'"),
                ("unit_price", "FLOAT"),
                ("line_cost", "FLOAT"),
            ]:
                try:
                    cur.execute(f"ALTER TABLE consumables ADD COLUMN IF NOT EXISTS {col} {typedef}")
                except Exception:
                    conn.rollback()

            cur.execute("""
                CREATE TABLE IF NOT EXISTS crew (
                    id          SERIAL PRIMARY KEY,
                    source_file TEXT,
                    contractor  TEXT DEFAULT 'Allianz Drilling',
                    date        TEXT,
                    hole_num    TEXT,
                    site_name   TEXT,
                    role        TEXT,
                    name        TEXT,
                    hours       TEXT
                )
            """)
            try:
                cur.execute("ALTER TABLE crew ADD COLUMN IF NOT EXISTS contractor TEXT DEFAULT 'Allianz Drilling'")
            except Exception:
                conn.rollback()

            cur.execute("""
                CREATE TABLE IF NOT EXISTS imported_files (
                    filename   TEXT,
                    contractor TEXT DEFAULT 'Allianz Drilling',
                    PRIMARY KEY (filename, contractor)
                )
            """)
            try:
                cur.execute("ALTER TABLE imported_files ADD COLUMN IF NOT EXISTS contractor TEXT DEFAULT 'Allianz Drilling'")
            except Exception:
                conn.rollback()

            cur.execute("""
                CREATE TABLE IF NOT EXISTS drilling_rates (
                    id         SERIAL PRIMARY KEY,
                    contractor TEXT NOT NULL DEFAULT 'Allianz Drilling',
                    year       TEXT NOT NULL,
                    bit_type   TEXT NOT NULL,
                    depth_from FLOAT NOT NULL,
                    depth_to   FLOAT NOT NULL,
                    rate       FLOAT NOT NULL
                )
            """)
            try:
                cur.execute("ALTER TABLE drilling_rates ADD COLUMN IF NOT EXISTS contractor TEXT NOT NULL DEFAULT 'Allianz Drilling'")
            except Exception:
                conn.rollback()

            cur.execute("""
                CREATE TABLE IF NOT EXISTS hourly_rates (
                    id          SERIAL PRIMARY KEY,
                    contractor  TEXT NOT NULL DEFAULT 'Allianz Drilling',
                    year        TEXT NOT NULL,
                    code        TEXT NOT NULL,
                    description TEXT,
                    rate        FLOAT NOT NULL,
                    unit        TEXT NOT NULL
                )
            """)
            try:
                cur.execute("ALTER TABLE hourly_rates ADD COLUMN IF NOT EXISTS contractor TEXT NOT NULL DEFAULT 'Allianz Drilling'")
            except Exception:
                conn.rollback()

            cur.execute("""
                CREATE TABLE IF NOT EXISTS boreholes (
                    id              SERIAL PRIMARY KEY,
                    contractor      TEXT NOT NULL DEFAULT 'Allianz Drilling',
                    project         TEXT,
                    planned_year    TEXT,
                    site_id         TEXT,
                    hole_id         TEXT NOT NULL,
                    drill_order     INTEGER,
                    days_budgeted   FLOAT,
                    bh_type         TEXT,
                    bit_type        TEXT,
                    purpose         TEXT,
                    easting         FLOAT,
                    northing        FLOAT,
                    rl              FLOAT,
                    chip_depth      FLOAT,
                    eoh_depth       FLOAT,
                    total_core      FLOAT,
                    seam_tk         FLOAT,
                    lat             FLOAT,
                    lng             FLOAT,
                    status          TEXT DEFAULT 'Planned',
                    notes           TEXT,
                    drilling_budget_total FLOAT,
                    earthworks_budget_total FLOAT,
                    geophysical_budget_total FLOAT,
                    geological_support_budget_total FLOAT,
                    misc_budget_total FLOAT,
                    budget_total    FLOAT,
                    actual_total    FLOAT,
                    UNIQUE(contractor, hole_id)
                )
            """)
            for col, typedef in [
                ("project", "TEXT"),
                ("planned_year", "TEXT"),
                ("site_id", "TEXT"),
                ("eoh_depth", "FLOAT"),
                ("total_core", "FLOAT"),
                ("assigned_rig", "TEXT"),
                ("scheduled_start", "TEXT"),
                ("scheduled_end", "TEXT"),
                ("drilling_budget_total", "FLOAT"),
                ("earthworks_budget_total", "FLOAT"),
                ("geophysical_budget_total", "FLOAT"),
                ("geological_support_budget_total", "FLOAT"),
                ("misc_budget_total", "FLOAT"),
            ]:
                try:
                    cur.execute(f"ALTER TABLE boreholes ADD COLUMN IF NOT EXISTS {col} {typedef}")
                except Exception:
                    conn.rollback()

            cur.execute("""
                CREATE TABLE IF NOT EXISTS purchase_orders (
                    id              SERIAL PRIMARY KEY,
                    po_number       TEXT NOT NULL,
                    contractor      TEXT NOT NULL DEFAULT 'Allianz Drilling',
                    project         TEXT,
                    description     TEXT,
                    issue_date      TEXT,
                    expiry_date     TEXT,
                    po_value        FLOAT DEFAULT 0,
                    status          TEXT DEFAULT 'Active',
                    notes           TEXT
                )
            """)
            for col, typedef in [("project", "TEXT")]:
                try:
                    cur.execute(f"ALTER TABLE purchase_orders ADD COLUMN IF NOT EXISTS {col} {typedef}")
                except Exception:
                    conn.rollback()

            cur.execute("""
                CREATE TABLE IF NOT EXISTS source_files (
                    id          SERIAL PRIMARY KEY,
                    filename    TEXT NOT NULL,
                    contractor  TEXT,
                    file_type   TEXT,
                    pdf_data    BYTEA,
                    uploaded_at TIMESTAMP DEFAULT NOW(),
                    UNIQUE(filename, contractor)
                )
            """)

            cur.execute("""
                CREATE TABLE IF NOT EXISTS report_approvals (
                    id          SERIAL PRIMARY KEY,
                    contractor  TEXT NOT NULL DEFAULT 'Allianz Drilling',
                    report_date TEXT,
                    hole_num    TEXT,
                    source_file TEXT,
                    status      TEXT,
                    reason      TEXT,
                    log         JSONB DEFAULT '[]'::jsonb,
                    updated_at  TIMESTAMP DEFAULT NOW(),
                    UNIQUE(contractor, report_date, hole_num, source_file)
                )
            """)

            cur.execute("""
                CREATE TABLE IF NOT EXISTS activity_sheet_locks (
                    id          SERIAL PRIMARY KEY,
                    contractor  TEXT NOT NULL DEFAULT 'Allianz Drilling',
                    report_date TEXT,
                    hole_num    TEXT,
                    source_file TEXT,
                    locked      BOOLEAN DEFAULT TRUE,
                    reason      TEXT,
                    updated_at  TIMESTAMP DEFAULT NOW(),
                    UNIQUE(contractor, report_date, hole_num, source_file)
                )
            """)

            cur.execute("""
                CREATE TABLE IF NOT EXISTS minimum_shift_topup_preferences (
                    id            SERIAL PRIMARY KEY,
                    contractor    TEXT NOT NULL DEFAULT 'Allianz Drilling',
                    report_date   TEXT,
                    hole_num      TEXT,
                    source_file   TEXT,
                    include_topup BOOLEAN DEFAULT TRUE,
                    reason        TEXT,
                    updated_at    TIMESTAMP DEFAULT NOW(),
                    UNIQUE(contractor, report_date, hole_num, source_file)
                )
            """)

            cur.execute("""
                CREATE TABLE IF NOT EXISTS consumable_rates (
                    id          SERIAL PRIMARY KEY,
                    contractor  TEXT NOT NULL DEFAULT 'Allianz Drilling',
                    year        TEXT NOT NULL DEFAULT '2025',
                    product     TEXT NOT NULL,
                    description TEXT,
                    unit_price  FLOAT DEFAULT 0,
                    unit        TEXT DEFAULT 'each'
                )
            """)

            cur.execute("""
                CREATE TABLE IF NOT EXISTS contractors (
                    id         SERIAL PRIMARY KEY,
                    name       TEXT NOT NULL UNIQUE,
                    short_code TEXT,
                    category   TEXT DEFAULT 'Misc',
                    program    TEXT DEFAULT 'Exploration',
                    sites      TEXT DEFAULT 'Ironbark',
                    expense_gl TEXT,
                    active     BOOLEAN DEFAULT TRUE
                )
            """)
            try:
                cur.execute("ALTER TABLE contractors ADD COLUMN IF NOT EXISTS program TEXT DEFAULT 'Exploration'")
                cur.execute("ALTER TABLE contractors ADD COLUMN IF NOT EXISTS category TEXT DEFAULT 'Misc'")
                cur.execute("ALTER TABLE contractors ADD COLUMN IF NOT EXISTS sites TEXT DEFAULT 'Ironbark'")
                cur.execute("ALTER TABLE contractors ADD COLUMN IF NOT EXISTS expense_gl TEXT")
                cur.execute("UPDATE contractors SET program='Exploration' WHERE program IS NULL OR program=''")
                cur.execute("UPDATE contractors SET category='Misc' WHERE category IS NULL OR category=''")
                cur.execute("UPDATE contractors SET sites='Ironbark' WHERE sites IS NULL OR BTRIM(sites)=''")
                for con_name, con_category in DEFAULT_CONTRACTOR_CATEGORIES.items():
                    cur.execute("""
                        UPDATE contractors
                        SET category=%s
                        WHERE name=%s AND (category IS NULL OR category='' OR category='Misc')
                    """, (con_category, con_name))
                cur.execute("""
                    UPDATE contractors
                    SET expense_gl = CASE
                        WHEN category='Drilling' THEN '4350'
                        WHEN category IN ('Earthworks','Labour') THEN '4250'
                        ELSE '4200'
                    END
                    WHERE expense_gl IS NULL OR BTRIM(expense_gl)=''
                """)
            except Exception:
                conn.rollback()

            cur.execute("""
                CREATE TABLE IF NOT EXISTS projects (
                    id         SERIAL PRIMARY KEY,
                    contractor TEXT NOT NULL DEFAULT 'Allianz Drilling',
                    program    TEXT DEFAULT 'Exploration',
                    name       TEXT NOT NULL,
                    year       TEXT,
                    status     TEXT DEFAULT 'Active',
                    notes      TEXT,
                    UNIQUE(contractor, name)
                )
            """)
            try:
                cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS program TEXT DEFAULT 'Exploration'")
                cur.execute("UPDATE projects SET program='Exploration' WHERE program IS NULL OR program=''")
            except Exception:
                conn.rollback()

            # ── Invoices ──────────────────────────────────────────────────────
            cur.execute("""
                CREATE TABLE IF NOT EXISTS project_budgets (
                    id            SERIAL PRIMARY KEY,
                    contractor    TEXT NOT NULL DEFAULT 'Company',
                    program       TEXT DEFAULT 'Exploration',
                    project       TEXT NOT NULL,
                    section       TEXT NOT NULL,
                    vendor        TEXT,
                    budget_amount FLOAT DEFAULT 0,
                    allocation    TEXT DEFAULT 'Project level',
                    notes         TEXT,
                    updated_at    TIMESTAMP DEFAULT NOW(),
                    UNIQUE(contractor, project, section, vendor)
                )
            """)
            for col, typedef in [
                ("contractor", "TEXT NOT NULL DEFAULT 'Company'"),
                ("program", "TEXT DEFAULT 'Exploration'"),
                ("project", "TEXT"),
                ("section", "TEXT"),
                ("vendor", "TEXT"),
                ("budget_amount", "FLOAT DEFAULT 0"),
                ("allocation", "TEXT DEFAULT 'Project level'"),
                ("notes", "TEXT"),
                ("updated_at", "TIMESTAMP DEFAULT NOW()"),
            ]:
                try:
                    cur.execute(f"ALTER TABLE project_budgets ADD COLUMN IF NOT EXISTS {col} {typedef}")
                except Exception:
                    conn.rollback()

            cur.execute("""
                CREATE TABLE IF NOT EXISTS cost_centre_forecasts (
                    id                SERIAL PRIMARY KEY,
                    site              TEXT NOT NULL,
                    division          TEXT NOT NULL,
                    cost_centre       TEXT NOT NULL,
                    program           TEXT NOT NULL,
                    year              INTEGER NOT NULL,
                    month             INTEGER NOT NULL,
                    category          TEXT NOT NULL,
                    baseline_amount   FLOAT DEFAULT 0,
                    finance_actual    FLOAT DEFAULT 0,
                    manual_accrual    FLOAT DEFAULT 0,
                    forecast_override FLOAT,
                    notes             TEXT,
                    source_file       TEXT,
                    updated_at        TIMESTAMP DEFAULT NOW(),
                    UNIQUE(site, cost_centre, year, month, category)
                )
            """)

            # F07 is a high-level Ironbark Technical Services - Exploration forecast.
            # This view is intentionally limited to the three TECEXP GLs requested by
            # Finance: 4200 Consulting, 4250 Contractors and 4350 Drilling Services.
            # Jan-Jun are Finance actuals and Jul-Dec are the F07 forecast baseline.
            f07_tec_exp = {
                "4200 - Consulting Services": [
                    12029.00, 22771.50, 46547.89, -68316.00, 0, 0,
                    103177.00, 98177.00, 467966.00, 99908.00, 388908.00, 88908.00,
                ],
                "4250 - Contractors": [
                    18974.66, 0, -0.26, 0, 0.38, 44441.55,
                    81510.00, 81510.00, 81510.00, 81510.00, 81510.00, 81510.00,
                ],
                "4350 - Drilling Services": [
                    0, 0, 0, 0, 0, 129343.00,
                    267795.00, 267795.00, 267795.00, 267795.00, 267795.00, 357060.00,
                ],
            }
            cur.execute("""
                DELETE FROM cost_centre_forecasts
                WHERE site='IB' AND cost_centre='TECEXP' AND year=2026
                  AND category NOT IN (
                    '4200 - Consulting Services',
                    '4250 - Contractors',
                    '4350 - Drilling Services'
                  )
            """)
            for category, monthly_amounts in f07_tec_exp.items():
                for month, amount in enumerate(monthly_amounts, start=1):
                    cur.execute("""
                        INSERT INTO cost_centre_forecasts
                        (site, division, cost_centre, program, year, month, category,
                         baseline_amount, finance_actual, manual_accrual,
                         forecast_override, notes, source_file)
                        VALUES ('IB','TEC','TECEXP','Exploration',2026,%s,%s,%s,%s,0,NULL,'','F07 IB TEC.xlsx')
                        ON CONFLICT (site, cost_centre, year, month, category) DO UPDATE SET
                            baseline_amount=EXCLUDED.baseline_amount,
                            finance_actual=EXCLUDED.finance_actual,
                            source_file=EXCLUDED.source_file,
                            forecast_override=CASE
                                WHEN cost_centre_forecasts.notes='Program hold: no forward spend assumed'
                                THEN NULL ELSE cost_centre_forecasts.forecast_override END,
                            notes=CASE
                                WHEN cost_centre_forecasts.notes='Program hold: no forward spend assumed'
                                THEN '' ELSE cost_centre_forecasts.notes END,
                            updated_at=NOW()
                    """, (
                        month,
                        category,
                        amount,
                        amount if month <= 6 else 0,
                    ))

            cur.execute("""
                CREATE TABLE IF NOT EXISTS invoices (
                    id              SERIAL PRIMARY KEY,
                    source_file     TEXT,
                    contractor      TEXT NOT NULL DEFAULT 'Allianz Drilling',
                    invoice_number  TEXT,
                    invoice_date    TEXT,
                    due_date        TEXT,
                    po_reference    TEXT,
                    client          TEXT,
                    abn             TEXT,
                    subtotal        FLOAT DEFAULT 0,
                    gst             FLOAT DEFAULT 0,
                    total_aud       FLOAT DEFAULT 0,
                    amount_paid     FLOAT DEFAULT 0,
                    amount_due      FLOAT DEFAULT 0,
                    status          TEXT DEFAULT 'Unpaid',
                    notes           TEXT,
                    pdf_data        BYTEA,
                    billing_month   TEXT
                )
            """)
            for col, typedef in [("pdf_data", "BYTEA"), ("billing_month", "TEXT"), ("version", "INTEGER DEFAULT 1"), ("query_notes", "TEXT"), ("project", "TEXT")]:
                try:
                    cur.execute(f"ALTER TABLE invoices ADD COLUMN IF NOT EXISTS {col} {typedef}")
                except Exception:
                    conn.rollback()

            # ── Invoice line items ────────────────────────────────────────────
            try:
                cur.execute("""
                    UPDATE invoices
                    SET project='Ironbark'
                    WHERE source_file LIKE 'manual:%%'
                      AND (project IS NULL OR project='' OR project IN ('Exploration','Gas Riser','SIS'))
                """)
            except Exception:
                conn.rollback()

            cur.execute("""
                CREATE TABLE IF NOT EXISTS invoice_lines (
                    id              SERIAL PRIMARY KEY,
                    invoice_id      INTEGER REFERENCES invoices(id) ON DELETE CASCADE,
                    contractor      TEXT NOT NULL DEFAULT 'Allianz Drilling',
                    invoice_number  TEXT,
                    description     TEXT,
                    quantity        FLOAT DEFAULT 0,
                    unit_price      FLOAT DEFAULT 0,
                    gst_rate        TEXT DEFAULT '10%',
                    amount          FLOAT DEFAULT 0,
                    category        TEXT,
                    matched_eos_cost FLOAT,
                    variance        FLOAT,
                    match_status    TEXT DEFAULT 'unmatched'
                )
            """)
            for col, typedef in [
                ("line_date", "TEXT"),
                ("site_name", "TEXT"),
                ("hole_num", "TEXT"),
                ("activity_code", "TEXT"),
                ("unit", "TEXT"),
                ("chargeable", "TEXT"),
                ("source_category", "TEXT"),
            ]:
                try:
                    cur.execute(f"ALTER TABLE invoice_lines ADD COLUMN IF NOT EXISTS {col} {typedef}")
                except Exception:
                    conn.rollback()

            cur.execute("""
                CREATE TABLE IF NOT EXISTS invoice_imports (
                    filename  TEXT,
                    contractor TEXT DEFAULT 'Allianz Drilling',
                    PRIMARY KEY (filename, contractor)
                )
            """)
            for index_sql in [
                "CREATE INDEX IF NOT EXISTS idx_activities_contractor_date ON activities (contractor, date)",
                "CREATE INDEX IF NOT EXISTS idx_consumables_contractor_date ON consumables (contractor, date)",
                "CREATE INDEX IF NOT EXISTS idx_crew_contractor_date ON crew (contractor, date)",
                "CREATE INDEX IF NOT EXISTS idx_report_approvals_contractor_date ON report_approvals (contractor, report_date)",
                "CREATE INDEX IF NOT EXISTS idx_activity_locks_contractor_date ON activity_sheet_locks (contractor, report_date)",
            ]:
                cur.execute(index_sql)
        conn.commit()


init_db()

MCC_SCHEDULE_DATE = "23 April 2026"
MCC_SCHEDULE_RATES = [
    ("MCC_LABOURER", "Labourer", 85.00, "hour", "labour", ["labourer"]),
    ("MCC_CONSTRUCTION_TRADE", "Construction Trade", 100.00, "hour", "labour", ["construction trade", "construc4on trade"]),
    ("MCC_MECHANICAL_TRADE", "Mechanical Trade", 115.00, "hour", "labour", ["mechanical trade"]),
    ("MCC_PUMP_CREW_OPERATOR", "Pump Crew Operator", 95.00, "hour", "labour", ["pump crew operator"]),
    ("MCC_MULTI_SKILLED_OPERATOR", "Multi Skilled Operator", 100.00, "hour", "labour", ["multi skilled operator", "mul4 skilled operator"]),
    ("MCC_SUPERVISOR", "Supervisor", 120.00, "hour", "labour", ["supervisor"]),
    ("MCC_PROJECT_MANAGER", "Project Manager", 140.00, "hour", "labour", ["project manager"]),
    ("MCC_LIGHT_VEHICLE", "Light Vehicle", 105.00, "day", "equipment", ["light vehicle", "light vehicles", "lv"]),
    ("MCC_5T_EXCAVATOR", "5t Excavator", 50.00, "hour", "equipment", ["5t excavator", "pc45 excavator", "komatsu pc45", "ex02"]),
    ("MCC_13T_EXCAVATOR", "13t Excavator", 85.00, "hour", "equipment", ["13t excavator"]),
    ("MCC_BACKHOE", "Backhoe", 85.00, "hour", "equipment", ["backhoe", "caterpillar 432", "caterpillar 432 backhoe", "ld04"]),
    ("MCC_36T_EXCAVATOR", "36t Excavator", 115.00, "hour", "equipment", ["36t excavator"]),
    ("MCC_SKID_STEER", "Skid Steer", 50.00, "hour", "equipment", ["skid steer"]),
    ("MCC_10T_BODY_TIP_TRUCK", "10t Body Tip Truck", 50.00, "hour", "equipment", ["10t body tip truck"]),
    ("MCC_BODY_WATER_TRUCK", "Body Water Truck", 80.00, "hour", "equipment", ["body water truck"]),
    ("MCC_105HP_TRACTOR", "105 Horsepower Tractor", 85.00, "hour", "equipment", ["105 horsepower tractor"]),
    ("MCC_SMALL_TOOL_HIRE", "Small Tool Hire", 50.00, "day", "equipment", ["small tool hire", "chainsaw", "whipper snipper"]),
    ("MCC_355MM_POLYWELDER", "355mm Polywelder", 150.00, "day", "equipment", ["355mm polywelder", "polywelder"]),
    ("MCC_TRAILER_HIRE", "Trailer Hire", 100.00, "day", "equipment", ["trailer hire"]),
    ("MCC_EQUIPMENT_ATTACHMENT", "Attachment for Equipment", 25.00, "hour", "equipment", ["attachment for equipment", "grader", "auger", "rock breaker", "slasher"]),
    ("MCC_120T_EXCAVATOR", "120t Excavator", 220.00, "hour", "equipment", ["120t excavator"]),
    ("MCC_90T_EXCAVATOR", "90t Excavator", 180.00, "hour", "equipment", ["90t excavator"]),
    ("MCC_100T_DUMP_WATER_TRUCK", "100t Dump Truck Class/Water Truck", 165.00, "hour", "equipment", ["100t dump truck", "100t water truck"]),
    ("MCC_40T_ARTICULATED_WATER_TRUCK", "40t Articulated Water Truck", 130.00, "hour", "equipment", ["40t articulated water truck", "40t ar4culated water truck"]),
    ("MCC_IT_LOADER", "IT Loader 15-20t Class", 85.00, "hour", "equipment", ["it loader", "15-20t class"]),
    ("MCC_LOADER_110T", "Loader 110t Class", 230.00, "hour", "equipment", ["loader 110t"]),
    ("MCC_SERVICE_TRUCK", "Service Truck", 80.00, "hour", "equipment", ["service truck"]),
    ("MCC_14_GRADER", "14ft Grader", 125.00, "hour", "equipment", ["14ft grader", "14^ grader", "14 grader"]),
    ("MCC_16_GRADER", "16ft Grader", 145.00, "hour", "equipment", ["16ft grader", "16^ grader", "16 grader"]),
    ("MCC_30T_ARTICULATED_DUMP_TRUCK", "30t Articulated Dump Truck", 105.00, "hour", "equipment", ["30t articulated dump truck", "30t ar4culated dump truck"]),
    ("MCC_40T_ARTICULATED_DUMP_TRUCK", "40t Articulated Dump Truck", 130.00, "hour", "equipment", ["40t articulated dump truck", "40t ar4culated dump truck"]),
    ("MCC_D11_DOZER", "D11 Dozer", 225.00, "hour", "equipment", ["d11 dozer"]),
    ("MCC_D10_DOZER", "D10 Dozer", 185.00, "hour", "equipment", ["d10 dozer"]),
]


def _norm_rate_text(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def mcc_schedule_match(value, group=None):
    haystack = _norm_rate_text(value)
    if not haystack:
        return None
    for code, desc, rate, unit, rate_group, aliases in MCC_SCHEDULE_RATES:
        if group and rate_group != group:
            continue
        for alias in aliases + [desc]:
            needle = _norm_rate_text(alias)
            if needle and (needle == haystack or needle in haystack or haystack in needle):
                return {"code": code, "description": desc, "rate": rate, "unit": unit, "group": rate_group}
    return None


# ── Seed 2025 rates (Allianz Drilling ONLY — other contractors start blank) ───
def note_field_value(text, label):
    prefix = f"{label.lower()}:"
    for part in str(text or "").split("|"):
        item = part.strip()
        if item.lower().startswith(prefix):
            return item.split(":", 1)[1].strip()
    return ""


def mcc_row_hours(row):
    value = str(row.get("total_time") or row.get("duration") or "").strip()
    if not value:
        return 0.0
    m = re.match(r"^(\d+):(\d+)", value)
    if m:
        return int(m.group(1)) + int(m.group(2)) / 60.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def mcc_row_year(row):
    text = " ".join(str(row.get(k) or "") for k in ("date", "source_file", "rate_year"))
    m = re.search(r"(20\d{2})", text)
    return m.group(1) if m else "2026"


def mcc_reprice_from_row(row):
    contractor = str(row.get("contractor") or "")
    source_file = str(row.get("source_file") or "")
    notes = str(row.get("notes") or "")
    if "mcc" not in contractor.lower() and "weeklysheets" not in source_file.lower() and "workstream:" not in notes.lower():
        return None

    code = str(row.get("code") or "")
    match = None
    if code.startswith("MCC_"):
        match = next((
            {"code": c, "description": d, "rate": r, "unit": u, "group": g}
            for c, d, r, u, g, _aliases in MCC_SCHEDULE_RATES
            if c == code
        ), None)
    if match is None:
        equipment = note_field_value(notes, "Equipment") or row.get("drill_rig") or ""
        role = note_field_value(notes, "Role") or ""
        match = mcc_schedule_match(equipment, "equipment") or mcc_schedule_match(role, "labour")
    if match is None:
        return None

    hours = mcc_row_hours(row)
    qty = 1 if match["unit"] == "day" else round(hours, 2)
    if qty <= 0:
        qty = 1
    return {
        "code": match["code"],
        "rate_year": mcc_row_year(row),
        "unit_rate": match["rate"],
        "quantity": qty,
        "line_cost": round(match["rate"] * qty, 2),
        "rate_basis": f"MCC schedule {MCC_SCHEDULE_DATE} - {match['description']} ${match['rate']:,.2f}/{match['unit']} x {qty:g}",
    }


def repair_mcc_weekly_activity_costs():
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT *
                FROM activities
                WHERE (contractor ILIKE 'MCC%%' OR source_file ILIKE '%%WeeklySheets%%' OR notes ILIKE '%%Workstream:%%')
                  AND (line_cost IS NULL OR line_cost = 0)
            """)
            rows = [dict(r) for r in cur.fetchall()]
            repaired = 0
            for row in rows:
                fix = mcc_reprice_from_row(row)
                if not fix:
                    continue
                cur.execute("""
                    UPDATE activities
                    SET code=%s, rate_year=%s, unit_rate=%s, quantity=%s, line_cost=%s, rate_basis=%s
                    WHERE id=%s
                """, (
                    fix["code"], fix["rate_year"], fix["unit_rate"], fix["quantity"],
                    fix["line_cost"], fix["rate_basis"], row["id"],
                ))
                repaired += 1
        conn.commit()
    return repaired


def seed_2025_rates():
    """Only seeds rates for Allianz Drilling. All other contractors start with
    empty rate schedules which must be configured manually."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS n FROM drilling_rates WHERE year='2025' AND contractor='Allianz Drilling'")
            if cur.fetchone()["n"] > 0:
                return
            YEAR = "2025"
            CON  = "Allianz Drilling"
            psycopg2.extras.execute_batch(cur, """
                INSERT INTO drilling_rates (contractor,year,bit_type,depth_from,depth_to,rate)
                VALUES (%s,%s,%s,%s,%s,%s)
            """, contract_drilling_rows(CON, YEAR))

            psycopg2.extras.execute_batch(cur, """
                INSERT INTO hourly_rates (contractor,year,code,description,rate,unit)
                VALUES (%s,%s,%s,%s,%s,%s)
            """, contract_hourly_rows(CON, YEAR))

            # Consumable rates
            consumables = [
                (CON, YEAR, "AMC CR650",                  "AMC CR650 drilling fluid",              127.10, "each"),
                (CON, YEAR, "MUDLOGIC SWELL HIB",         "Mudlogic Swell HIB",                    224.80, "each"),
                (CON, YEAR, "AMC HARD SET",               "AMC Hard Set cement additive",           66.55, "each"),
                (CON, YEAR, "AMC TORQ FREE XTRA 20LTR",  "AMC Torq Free Xtra 20L cube",          156.40, "each"),
                (CON, YEAR, "GP CEMENT",                  "General purpose cement",                 11.47, "bag"),
                (CON, YEAR, "PART A FOAM",                "Part A foam",                            18.86, "each"),
                (CON, YEAR, "PART B FOAM",                "Part B foam",                            18.86, "each"),
                (CON, YEAR, "AMC SUPERLUBE",              "AMC Superlube drilling lubricant",      150.26, "each"),
                (CON, YEAR, "AMC SUPERFOAM",              "AMC Superfoam",                         127.36, "each"),
                (CON, YEAR, "AMC BEN",                    "AMC Bentonite",                           0.00, "each"),
                (CON, YEAR, "AMC GEL",                    "AMC Gel",                                 0.00, "each"),
                (CON, YEAR, "PVC CASING 100MM CLASS 9",   "PVC Casing 100mm Class 9 (18m)",         0.00, "metre"),
            ]
            cur.execute("SELECT COUNT(*) AS n FROM consumable_rates WHERE year=%s AND contractor=%s", (YEAR, CON))
            if cur.fetchone()["n"] == 0:
                psycopg2.extras.execute_batch(cur, """
                    INSERT INTO consumable_rates (contractor,year,product,description,unit_price,unit)
                    VALUES (%s,%s,%s,%s,%s,%s)
                """, consumables)
        conn.commit()


def seed_mcc_2026_rates():
    rows = []
    for contractor in ("MCC Group",):
        for code, desc, rate, unit, group, _aliases in MCC_SCHEDULE_RATES:
            rows.append((contractor, "2026", code, f"{desc} ({group}; MCC schedule {MCC_SCHEDULE_DATE})", rate, unit))
    with get_conn() as conn:
        with conn.cursor() as cur:
            for row in rows:
                cur.execute("""
                    INSERT INTO hourly_rates (contractor, year, code, description, rate, unit)
                    SELECT %s, %s, %s, %s, %s, %s
                    WHERE NOT EXISTS (
                        SELECT 1 FROM hourly_rates
                        WHERE contractor=%s AND year=%s AND code=%s
                    )
                """, (*row, row[0], row[1], row[2]))
        conn.commit()


def remove_legacy_mcc_earthworks_seed():
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM contractors WHERE name='MCC Earthworks'")
            cur.execute("DELETE FROM hourly_rates WHERE contractor='MCC Earthworks'")
        conn.commit()


seed_2025_rates()
seed_mcc_2026_rates()
remove_legacy_mcc_earthworks_seed()
repair_mcc_weekly_activity_costs()
migrate_legacy_drilling_bit_labels()
apply_mitchells_contract_exceptions()


# ── Pricing engine ────────────────────────────────────────────────────────────
DRILLING_METRE_CODES = {"Drill_Core", "Drill_Chip_or_Open_hole"}
ALLIANZ_MIN_SHIFT_COST = 8940.00
ALLIANZ_MIN_SHIFT_ACTIVE_RATE = 745.00
ALLIANZ_MIN_SHIFT_TOPUP_NOTE = "Allianz minimum shift top-up to $8,940"
MINIMUM_SHIFT_RULES = {
    "Allianz Drilling": {
        "cost": 8940.00,
        "active_rate": 745.00,
        "note": ALLIANZ_MIN_SHIFT_TOPUP_NOTE,
        "label": "Allianz",
    },
    "Mitchells Drilling": {
        "cost": 7800.00,
        "active_rate": 650.00,
        "note": "Mitchells minimum shift top-up to $7,800",
        "label": "Mitchells",
    },
}
SUPPORT_EQUIPMENT_CODE_RE = re.compile(r"^D_(Backhoe|Water)", re.I)
# Codes charged at Active rate ($/hr)
ACTIVE_CODES = {
    "H_Tripping_Rods","H_Circulation_Flush","H_Circulation_Lost",
    "H_Reaming","H_Change_Drill_Mthd",
    "H_Casing_Install","H_Rig_Cementing",
}
# Codes charged at Inactive rate ($/hr)
INACTIVE_CODES = {
    "H_Crew_Travel_On","H_Crew_Travel_Off","H_Rig_Move","H_Setup_Packup_Site",
    "H_Safety_Contractor","H_Safety_Prestart","H_Con_Collect_Plan",
    "H_Standby_Sumps","H_Standby_AAC","H_Standby_Logger","H_Standby_Grout",
    "H_Standby_Cement_Set","H_Standby_Cement_set",
    "H_Mud_Mixing","H_Surface_Setup","H_Training","H_Water_Flow_Measure",
}
# Not chargeable
NOT_CHARGEABLE = {"H_Repairs","Crew_Travel"}
MITCHELLS_NOT_CHARGEABLE = {"H_Standby_Fatigue"}
STANDBY_CODES = {
    "H_Standby_Sumps","H_Standby_AAC","H_Standby_Logger",
    "H_Standby_Grout","H_Standby_Cement_Set","H_Standby_Cement_set",
    "H_Standby_Fatigue",
}
DAY_RATE_CODES = {
    "D_Backhoe":                  ("D_Backhoe",            "day"),
    "D_Backhoe_Day_Rate":         ("D_Backhoe",            "day"),
    "D_Backhoe - Day Rate":       ("D_Backhoe",            "day"),
    "D_Backhoe - Standby Rate":   ("D_Backhoe_Standby",    "day"),
    "D_Backhoe_Standby":          ("D_Backhoe_Standby",    "day"),
    "D_WaterCart_Day_Rate":       ("D_Water_Cart",         "day"),
    "D_Water_Cart_Day_Rate":      ("D_Water_Cart",         "day"),
    "D_Water_Cart":               ("D_Water_Cart",         "day"),
    "D_Water Cart - Standby Rate":("D_Water_Cart_Standby", "day"),
    "D_Water_Cart_Standby":       ("D_Water_Cart_Standby", "day"),
}
BIT_TYPE_MAP = {"HQ_HQ3":"HQ_HQ3","HQ":"HQ_HQ3","NQ":"HQ_HQ3","4C":"4C","PCD":PCD_SMALL_LABEL}


def is_not_chargeable_code(code: str, contractor: str = "") -> bool:
    if code in NOT_CHARGEABLE:
        return True
    return contractor == "Mitchells Drilling" and code in MITCHELLS_NOT_CHARGEABLE


def allianz_minimum_shift_group_key(row: dict):
    source_file = row.get("source_file") or ""
    return (
        row.get("contractor") or "Allianz Drilling",
        source_file,
        row.get("date") or "",
        "" if source_file else (row.get("hole_num") or ""),
    )


def minimum_shift_preference_params(contractor: str, report_date: str = "", hole_num: str = "", source_file: str = ""):
    return {
        "contractor": contractor or "Allianz Drilling",
        "report_date": report_date or "",
        "hole_num": hole_num or "",
        "source_file": source_file or "",
    }


def minimum_shift_preference_response(row):
    if not row:
        return {
            "contractor": "",
            "report_date": "",
            "hole_num": "",
            "source_file": "",
            "include_topup": True,
            "reason": "",
            "updated_at": None,
        }
    return {
        "contractor": row.get("contractor") or "",
        "report_date": row.get("report_date") or "",
        "hole_num": row.get("hole_num") or "",
        "source_file": row.get("source_file") or "",
        "include_topup": bool(row.get("include_topup", True)),
        "reason": row.get("reason") or "",
        "updated_at": row.get("updated_at"),
    }


def minimum_shift_rule(contractor: str):
    return MINIMUM_SHIFT_RULES.get(contractor or "")


def is_generated_minimum_shift_topup(row: dict) -> bool:
    return (row.get("code") == "H_Min_Shift") and (row.get("notes") in {r["note"] for r in MINIMUM_SHIFT_RULES.values()})


def build_allianz_minimum_shift_topups(rows: list[dict]) -> list[dict]:
    groups = {}
    for row in rows:
        contractor = row.get("contractor") or ""
        rule = minimum_shift_rule(contractor)
        if not rule:
            continue
        if is_generated_minimum_shift_topup(row):
            continue
        key = allianz_minimum_shift_group_key(row)
        group = groups.setdefault(key, {"base": row, "rule": rule, "drilling_cost": 0.0, "has_drilling_cost": False, "has_min_shift": False})
        code = row.get("code") or ""
        if code == "H_Min_Shift":
            try:
                group["has_min_shift"] = group["has_min_shift"] or float(row.get("line_cost") or 0) > 0
            except (TypeError, ValueError):
                pass
            continue
        if SUPPORT_EQUIPMENT_CODE_RE.search(code):
            continue
        try:
            line_cost = float(row.get("line_cost") or 0)
        except (TypeError, ValueError):
            line_cost = 0
        if line_cost > 0:
            group["has_drilling_cost"] = True
            group["drilling_cost"] += line_cost

    topups = []
    for group in groups.values():
        if not group["has_drilling_cost"] or group["has_min_shift"]:
            continue
        drilling_cost = round(group["drilling_cost"], 2)
        rule = group["rule"]
        topup = round(rule["cost"] - drilling_cost, 2)
        if topup <= 0:
            continue
        base = group["base"]
        active_hours = round(topup / rule["active_rate"], 2) if rule["active_rate"] else 0
        topups.append({
            "source_file": base.get("source_file"),
            "contractor": base.get("contractor"),
            "date": base.get("date"),
            "hole_num": base.get("hole_num"),
            "site_name": base.get("site_name"),
            "location": base.get("location"),
            "drill_rig": base.get("drill_rig"),
            "client": base.get("client"),
            "contract": base.get("contract"),
            "shift": base.get("shift"),
            "time_from": "",
            "time_to": "",
            "total_time": "0:00",
            "bit_type": "",
            "diameter": "",
            "metres_from": None,
            "metres_to": None,
            "total_metres": None,
            "code": "H_Min_Shift",
            "notes": rule["note"],
            "rate_year": extract_year(base.get("date") or ""),
            "unit_rate": rule["active_rate"],
            "quantity": active_hours,
            "line_cost": topup,
            "rate_basis": f"{rule['label']} minimum shift: drilling activity ${drilling_cost:,.2f} + top-up ${topup:,.2f} = ${rule['cost']:,.2f}",
            "po_id": None,
        })
    return topups


def apply_allianz_minimum_shift_topups_to_rows(rows: list[dict]) -> list[dict]:
    if not rows:
        return rows
    return rows + build_allianz_minimum_shift_topups(rows)


def minimum_shift_base_cost(row: dict) -> float:
    basis = row.get("rate_basis") or ""
    basis_match = re.search(r"=\s*\$?([0-9,]+(?:\.\d+)?)", basis)
    if basis_match:
        try:
            return round(float(basis_match.group(1).replace(",", "")), 2)
        except (TypeError, ValueError):
            pass
    try:
        unit_rate = float(row.get("unit_rate") or 0)
    except (TypeError, ValueError):
        unit_rate = 0
    try:
        quantity = float(row.get("quantity") or 0)
    except (TypeError, ValueError):
        quantity = 0
    if unit_rate > 0 and quantity > 0:
        return round(unit_rate * quantity, 2)
    try:
        return round(float(row.get("line_cost") or 0), 2)
    except (TypeError, ValueError):
        return 0


def is_imported_minimum_shift_row(row: dict) -> bool:
    if row.get("code") != "H_Min_Shift":
        return False
    if is_generated_minimum_shift_topup(row):
        return False
    return "minimum shift" in (row.get("notes") or row.get("rate_basis") or "").lower()


def minimum_shift_activity_subtotal(row: dict, rows: list[dict]) -> float:
    key = allianz_minimum_shift_group_key(row)
    total = 0.0
    for other in rows:
        if allianz_minimum_shift_group_key(other) != key:
            continue
        code = other.get("code") or ""
        if code == "H_Min_Shift" or code.startswith("E_") or SUPPORT_EQUIPMENT_CODE_RE.search(code) or code in {"MOB", "DEMOB"}:
            continue
        try:
            line_cost = float(other.get("line_cost") or 0)
        except (TypeError, ValueError):
            line_cost = 0
        if line_cost > 0:
            total += line_cost
    return round(total, 2)


def minimum_shift_non_minimum_total(row: dict, rows: list[dict]) -> float:
    key = allianz_minimum_shift_group_key(row)
    total = 0.0
    for other in rows:
        if allianz_minimum_shift_group_key(other) != key:
            continue
        if (other.get("code") or "") == "H_Min_Shift" or is_generated_minimum_shift_topup(other):
            continue
        try:
            total += float(other.get("line_cost") or 0)
        except (TypeError, ValueError):
            pass
    return round(total, 2)


def minimum_shift_group_totals(rows: list[dict], locked_keys: set | None = None) -> dict:
    totals = {}
    locked_group_keys = set()
    for row in rows or []:
        key = allianz_minimum_shift_group_key(row)
        try:
            totals[key] = totals.get(key, 0.0) + float(row.get("line_cost") or 0)
        except (TypeError, ValueError):
            totals[key] = totals.get(key, 0.0)
        if locked_keys and row_is_in_locked_sheet(row, locked_keys):
            locked_group_keys.add(key)
    if locked_keys is not None:
        totals = {key: round(total, 2) for key, total in totals.items() if key in locked_group_keys}
    else:
        totals = {key: round(total, 2) for key, total in totals.items()}
    return totals


def calculated_coreplan_line_cost(row: dict):
    code = row.get("code") or ""
    if code == "H_Min_Shift" or is_generated_minimum_shift_topup(row):
        return None
    if "coreplan" not in (row.get("rate_basis") or "").lower():
        return None
    try:
        unit_rate = float(row.get("unit_rate") or 0)
        quantity = float(row.get("quantity") or 0)
    except (TypeError, ValueError):
        return None
    if unit_rate <= 0 or quantity <= 0:
        return None
    return round(unit_rate * quantity, 2)


def restore_coreplan_activity_line_costs(rows: list[dict]) -> list[dict]:
    restored = []
    for row in rows or []:
        updated = dict(row)
        calculated = calculated_coreplan_line_cost(updated)
        if calculated is not None:
            try:
                current = float(updated.get("line_cost") or 0)
            except (TypeError, ValueError):
                current = 0
            if current <= 0:
                updated["line_cost"] = calculated
        restored.append(updated)
    return restored


def update_restored_coreplan_activity_line_costs(cur, rows: list[dict]) -> tuple[list[dict], int]:
    original_by_id = {r.get("id"): r for r in rows or [] if r.get("id") is not None}
    restored = restore_coreplan_activity_line_costs(rows)
    updated_count = 0
    for row in restored:
        row_id = row.get("id")
        original = original_by_id.get(row_id)
        if not original:
            continue
        try:
            before = round(float(original.get("line_cost") or 0), 2)
            after = round(float(row.get("line_cost") or 0), 2)
        except (TypeError, ValueError):
            continue
        if after > 0 and before != after:
            cur.execute("UPDATE activities SET line_cost=%s WHERE id=%s", (after, row_id))
            updated_count += cur.rowcount
    return restored, updated_count


def adjust_imported_minimum_shift_rows(rows: list[dict], contractor: str = "", excluded_keys: set | None = None, target_total_by_key: dict | None = None) -> list[dict]:
    if not rows:
        return rows
    excluded_keys = excluded_keys or set()
    target_total_by_key = target_total_by_key or {}
    adjusted = []
    for row in rows:
        updated = dict(row)
        if is_imported_minimum_shift_row(updated):
            minimum_cost = minimum_shift_base_cost(updated)
            rule = minimum_shift_rule(updated.get("contractor") or contractor)
            if rule:
                minimum_cost = rule["cost"]
            key = allianz_minimum_shift_group_key(updated)
            preserve_total = target_total_by_key.get(key)
            activity_cost = minimum_shift_activity_subtotal(updated, rows)
            if preserve_total is not None:
                minimum_cost = round(float(preserve_total or 0), 2)
                activity_cost = minimum_shift_non_minimum_total(updated, rows)
            topup = round(max(0, minimum_cost - activity_cost), 2)
            try:
                unit_rate = float(updated.get("unit_rate") or 0)
            except (TypeError, ValueError):
                unit_rate = 0
            include_topup = allianz_minimum_shift_group_key(updated) not in excluded_keys
            updated["line_cost"] = topup if include_topup else 0
            if unit_rate > 0:
                updated["quantity"] = round((topup if include_topup else 0) / unit_rate, 2)
            if preserve_total is not None:
                basis = f"minimum shift top-up: approved/custom total preserved; other activity ${activity_cost:,.2f} + top-up ${topup:,.2f} = ${minimum_cost:,.2f}"
            else:
                basis = f"minimum shift top-up: drilling activity ${activity_cost:,.2f} + top-up ${topup:,.2f} = ${minimum_cost:,.2f}"
            updated["rate_basis"] = basis if include_topup else f"{basis} (excluded by reviewer)"
        adjusted.append(updated)
    return adjusted


def explicit_pcd_range_label(value: str) -> str:
    text = str(value or "").upper()
    compact = re.sub(r"[\s/\-–]+", "_", text)
    if PCD_LARGE_LABEL.upper() in text or "175_305MM" in compact:
        return PCD_LARGE_LABEL
    if PCD_MEDIUM_LABEL.upper() in text or "125_175MM" in compact:
        return PCD_MEDIUM_LABEL
    if PCD_SMALL_LABEL.upper() in text or "99_125MM" in compact:
        return PCD_SMALL_LABEL
    return ""


def normalise_drilling_bit_key(bit_type: str, code: str = "", notes: str = "") -> str:
    text = f"{bit_type or ''} {code or ''} {notes or ''}".upper()
    compact = re.sub(r"[\s/\-–]+", "_", text)
    tokens = set(re.findall(r"[A-Z0-9.]+", text))
    mm_ranges = [(float(a), float(b)) for a, b in re.findall(r"\b(\d+(?:\.\d+)?)\s*[-–]\s*(\d+(?:\.\d+)?)\s*MM\b", text)]
    mm_values = [float(v) for v in re.findall(r"(\d+(?:\.\d+)?)\s*MM\b", text)]
    inch_values = []
    for whole, numerator, denominator in re.findall(r"\b(\d+)\s+(\d+)/(\d+)\b", text):
        if float(denominator):
            inch_values.append(float(whole) + float(numerator) / float(denominator))
    for value in re.findall(r"\b(\d+(?:\.\d+)?)\s*(?:INCH|IN|\"|\u2033)\b", text):
        inch_values.append(float(value))
    for value in inch_values:
        mm_values.append(value * 25.4)
    if "4C" in text or "101.6" in text:
        return "4C"
    if "PQ" in text:
        return PQ_PQ3_LABEL
    if "HQ" in text or "NQ" in text:
        return "HQ_HQ3"
    if "HAMMER" in text:
        if "175" in text or "HAMMER_L" in compact or "L" in tokens:
            return "HAMMER_L"
        if "125" in text or "HAMMER_M" in compact or "M" in tokens:
            return "HAMMER_M"
        return "HAMMER_S"
    if "BLADE" in text and "PCD" not in text:
        if "175" in text or "BLADE_L" in compact or "L" in tokens:
            return "BLADE_L"
        if "125" in text or "BLADE_M" in compact or "M" in tokens:
            return "BLADE_M"
        return "BLADE_S"
    if "PCD" in text or "CHIP" in text or "OPEN_HOLE" in text:
        explicit = explicit_pcd_range_label(text)
        if explicit:
            return explicit
        if "7_10_INCH" in compact:
            return PCD_LARGE_LABEL
        if "5_7_INCH" in compact:
            return PCD_MEDIUM_LABEL
        if "3_1_2_5_INCH" in compact:
            return PCD_SMALL_LABEL
        if any(lo >= 175 and hi <= 305 for lo, hi in mm_ranges):
            return PCD_LARGE_LABEL
        if any(125 <= lo < 175 and hi <= 175 for lo, hi in mm_ranges):
            return PCD_MEDIUM_LABEL
        if any(99 <= lo < 125 and hi <= 125 for lo, hi in mm_ranges):
            return PCD_SMALL_LABEL
        if any(mm >= 175 for mm in mm_values):
            return PCD_LARGE_LABEL
        if any(125 <= mm < 175 for mm in mm_values):
            return PCD_MEDIUM_LABEL
        if any(99 <= mm < 125 for mm in mm_values):
            return PCD_SMALL_LABEL
        if "305" in text or "PCD_L" in compact or "L" in tokens:
            return PCD_LARGE_LABEL
        if "125" in text or "PCD_M" in compact or "M" in tokens:
            return PCD_MEDIUM_LABEL
        if "PCD_S" in compact or "S" in tokens:
            return PCD_SMALL_LABEL
        if "5-7/8" in text or "5_7_8" in compact:
            return PCD_MEDIUM_LABEL
        return PCD_SMALL_LABEL
    return PCD_SMALL_LABEL if "CHIP" in text else "HQ_HQ3"


def time_str_to_hours(t):
    try:
        h, m = str(t).split(":")
        return int(h) + int(m) / 60
    except:
        return 0.0


def extract_year(date_str):
    m = re.search(r"(\d{4})", str(date_str))
    if m: return m.group(1)
    m2 = re.search(r"/(\d{2})$", str(date_str))
    if m2: return str(2000 + int(m2.group(1)))
    return "2025"


def price_activity(cur, row, contractor):
    code         = row.get("code","") or ""
    total_time   = row.get("total_time","") or ""
    total_metres = row.get("total_metres")
    metres_to    = row.get("metres_to")
    bit_type     = row.get("bit_type","") or ""
    date_str     = row.get("date","") or ""
    year         = extract_year(date_str)
    hours        = time_str_to_hours(total_time)
    unit_rate = quantity = line_cost = rate_basis = None
    mcc_fix = mcc_reprice_from_row(row)

    def get_dr(bit_key, depth):
        # Try exact year first, then fall back to nearest available
        for try_year in [year, str(int(year)-1), str(int(year)+1), "2025"]:
            cur.execute("""
                SELECT rate FROM drilling_rates
                WHERE contractor=%s AND year=%s AND bit_type=%s
                  AND depth_from<=%s AND depth_to>%s
                ORDER BY depth_from LIMIT 1
            """, (contractor, try_year, bit_key, depth, depth))
            r = cur.fetchone()
            if r: return float(r["rate"])
        return None

    def get_hr(code_key):
        # Try exact year first, then fall back to nearest available
        for try_year in [year, str(int(year)-1), str(int(year)+1), "2025"]:
            cur.execute("SELECT rate FROM hourly_rates WHERE contractor=%s AND year=%s AND code=%s",
                        (contractor, try_year, code_key))
            r = cur.fetchone()
            if r: return float(r["rate"])
        return None

    if code in DRILLING_METRE_CODES and total_metres and total_metres > 0:
        cur.execute("SELECT * FROM drilling_rates WHERE contractor=%s", (contractor,))
        priced = drilling_schedule_cost(row, build_rate_context(drilling_rates=[dict(r) for r in cur.fetchall()]))
        if priced is not None:
            unit_rate = priced["unit_rate"]
            quantity = priced["quantity"]
            line_cost = priced["line_cost"]
            rate_basis = priced["rate_basis"]
    elif code in DRILLING_METRE_CODES:
        unit_rate = 0
        quantity = 0
        line_cost = 0
        rate_basis = "drilling time covered by metreage; no metres recorded"
    elif any(k in code for k in DAY_RATE_CODES) or any(code in k for k in DAY_RATE_CODES):
        matched = next((v for k,v in DAY_RATE_CODES.items() if k in code or code in k), None)
        if matched:
            r = get_hr(matched[0])
            if r is not None:
                unit_rate  = r; quantity = 1; line_cost = r
                rate_basis = f"${r:,.2f}/{matched[1]}"
    elif is_not_chargeable_code(code, contractor):
        unit_rate = 0; quantity = round(hours, 2) if hours > 0 else 1
        line_cost = 0; rate_basis = "not chargeable"
    elif mcc_fix:
        code = mcc_fix["code"]
        year = mcc_fix["rate_year"]
        unit_rate = mcc_fix["unit_rate"]
        quantity = mcc_fix["quantity"]
        line_cost = mcc_fix["line_cost"]
        rate_basis = mcc_fix["rate_basis"]
    elif code.startswith("E_"):
        r = get_hr(code)
        try:
            existing_qty = float(row.get("quantity") or 1)
        except (TypeError, ValueError):
            existing_qty = 1
        qty = round(hours, 2) if hours > 0 else existing_qty
        if r is not None:
            unit_rate = r; quantity = qty; line_cost = round(r * qty, 2)
            rate_basis = f"equipment ${r:,.2f}/unit x {qty}"
        elif row.get("line_cost") is not None:
            unit_rate = row.get("unit_rate"); quantity = row.get("quantity") or qty
            line_cost = row.get("line_cost"); rate_basis = row.get("rate_basis") or "equipment charge"
    elif code in STANDBY_CODES or "Standby" in code:
        r = get_hr(code)
        if r is None:
            r = get_hr("H_Inactive")
        if r is not None:
            qty = round(hours, 2) if hours > 0 else 1
            unit_rate  = r; quantity = qty
            line_cost  = round(r * qty, 2)
            rate_basis = f"inactive ${r:,.2f}/hr x {qty}h"
    elif code in INACTIVE_CODES:
        r = get_hr(code)
        if r is None:
            r = get_hr("H_Inactive")
        if r is not None:
            qty = round(hours, 2) if hours > 0 else 1
            unit_rate  = r; quantity = qty
            line_cost  = round(r * qty, 2)
            rate_basis = f"inactive ${r:,.2f}/hr x {qty}h"
    elif hours > 0 and (code in ACTIVE_CODES or "H_" in code):
        r = get_hr(code)
        if r is None:
            r = get_hr("H_Active")
        if r is not None:
            unit_rate  = r; quantity = round(hours,2)
            line_cost  = round(r * hours, 2); rate_basis = f"active ${r:,.2f}/hr x {hours:.2f}h"

    # Fallback: any activity with hours but no match above
    if line_cost is None and hours > 0 and code:
        # Check if we have a specific rate for this code
        r = get_hr(code)
        if r is not None:
            unit_rate = r; quantity = round(hours, 2)
            line_cost = round(r * hours, 2); rate_basis = f"${r:,.2f}/hr x {hours:.2f}h (code match)"
        else:
            r = get_hr("H_Active")
            if r:
                unit_rate = r; quantity = round(hours, 2)
                line_cost = round(r * hours, 2); rate_basis = f"fallback active ${r:,.2f}/hr x {hours:.2f}h"

    # Fallback 2: code present but no hours and no match above
    if line_cost is None and code and not hours:
        r = get_hr(code)
        if r is not None:
            unit_rate = r; quantity = 1
            line_cost = r; rate_basis = f"${r:,.2f} (1 unit, code match)"
        elif "PVC" in code or "Casing" in code or "Cement" in code:
            rate_basis = "consumable - no rate"
        elif "D_" in code:
            rate_basis = "day rate code - check schedule of rates"

    row.update(code=code, rate_year=year, unit_rate=unit_rate, quantity=quantity,
               line_cost=line_cost, rate_basis=rate_basis)
    return row


# ── PDF Parsing ───────────────────────────────────────────────────────────────
def parse_header(text):
    patterns = {
        "client":    r"CLIENT:\s*(.+?)\s+CONTRACT #:",
        "contract":  r"CONTRACT #:\s*(\S+)",
        "date":      r"DATE:\s*(\S+)",
        "hole_num":  r"HOLE #:\s*(\S+)",
        "shift":     r"SHIFT:\s*(\S+)",
        "site_name": r"SITE NAME:\s*(\S+)",
        "location":  r"LOCATION:\s*(\S+)",
        "drill_rig": r"DRILL RIG #\s*(\S+)",
    }
    return {k: (re.search(p, text, re.IGNORECASE).group(1).strip()
                if re.search(p, text, re.IGNORECASE) else "") for k, p in patterns.items()}


def detect_pdf_format(text):
    """Detect whether this is standard EOS, ADR001-decimal, or ADR001-standard format."""
    if "EXPLORATION DAILY SUPERVISORS REPORT" in text:
        # ADR001 template — but check if durations are decimal (March) or H:MM (May+)
        # Look for a CODE column or H:MM style third time field
        lines = text.splitlines()
        for line in lines:
            line = line.strip()
            # If line has 3 H:MM times and a code at end → standard format with ADR001 header
            if re.search(r'\d{1,2}:\d{2}\s+\d{1,2}:\d{2}\s+\d{1,2}:\d{2}\s+\S', line):
                return "standard"
            # If line has decimal duration (0.25, 1.50) → ADR001 decimal format
            if re.search(r'\d{1,2}:\d{2}\s+\d{1,2}:\d{2}\s+\d+\.\d+\s+\d', line):
                return "adr001"
        # If CODE appears in the column header → standard
        if re.search(r'TOTAL\s+METRES\s+CODE', text):
            return "standard"
        return "adr001"  # fallback for old format
    return "standard"


def parse_activities_adr001(text, header, filename, contractor):
    """Parse ADR001 format: decimal durations, full bit type names."""
    # ADR001 header may have hole_num at different position
    m = re.search(r"HOLE #:\s*(\S+)", text, re.IGNORECASE)
    if m and not header.get("hole_num"):
        header["hole_num"] = m.group(1)

    m = re.search(r"DRILL RIG #:\s*(\S+)", text, re.IGNORECASE)
    if m and not header.get("drill_rig"):
        header["drill_rig"] = m.group(1)

    # ADR001 row pattern: notes time_from time_to decimal_duration [bit_type diameter] metres_from metres_to metres_drilled
    # e.g. "flush hole 8:15 8:30 0.25 0 0 0"
    # e.g. "HQ - 211.40m to 214.40m... 16:30 17:00 0.50 HQ / HQ3 (Triple Tube / Wireline) 96 mm 211.4 214.4 3"
    row_re = re.compile(
        r"^(.+?)\s+(\d{1,2}:\d{2})\s+(\d{1,2}:\d{2})\s+(\d+\.?\d*)"
        r"\s+(?:([A-Z][A-Z\s/()]*?(?:Wireline|Tube))\s+(\d+\s*mm)\s+)?"
        r"(\d+\.?\d*)\s+(\d+\.?\d*)\s+(\d+\.?\d*)\s*$",
        re.IGNORECASE
    )

    # Simpler pattern for most lines: notes time_from time_to decimal 0 0 0
    simple_re = re.compile(
        r"^(.+?)\s+(\d{1,2}:\d{2})\s+(\d{1,2}:\d{2})\s+(\d+\.?\d*)\s+(\d+\.?\d*)\s+(\d+\.?\d*)\s+(\d+\.?\d*)\s*$"
    )

    # Pattern for lines with no times: "Backhoe on standby 0:00 0 0 0"
    no_time_re = re.compile(
        r"^(.+?)\s+0:00\s+0\s+0\s+0\s*$"
    )

    rows = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        # Skip header lines, total lines, crew section
        if line.startswith("NOTES") or line.startswith("TOTAL DURATION"):
            continue
        if line.startswith("PEOPLE ON") or line.startswith("Position") or line.startswith("Supervisor") or line.startswith("Driller") or line.startswith("Trainee") or line.startswith("Offsider") or line.startswith("Operator"):
            continue
        if line.startswith("CONSUMABLES") or line.startswith("Consumable"):
            continue
        if line.startswith("CLIENT:") or line.startswith("SHIFT:") or line.startswith("ALLIANZ") or line.startswith("Delay Hours"):
            continue
        if line.startswith("Non Drilling") or "REPRESENTATIVE" in line:
            continue
        if line == "0:00 0 0 0":
            continue

        # Try no-time pattern first (Backhoe on standby, Water cart, etc)
        nm = no_time_re.match(line)
        if nm:
            desc = nm.group(1).strip()
            if not desc or desc in ("0", ""):
                continue
            # Determine code from description
            code = ""
            dl = desc.lower()
            if "backhoe" in dl and "standby" in dl:
                code = "D_Backhoe_Standby"
            elif "backhoe" in dl:
                code = "D_Backhoe"
            elif "water cart" in dl and "standby" in dl:
                code = "D_Water_Cart_Standby"
            elif "water cart" in dl:
                code = "D_Water_Cart"

            if code:
                rows.append({
                    "source_file": filename, "contractor": contractor,
                    "date": header.get("date",""), "hole_num": header.get("hole_num",""),
                    "site_name": header.get("site_name",""), "location": header.get("location",""),
                    "drill_rig": header.get("drill_rig",""), "client": header.get("client",""),
                    "contract": header.get("contract",""), "shift": header.get("shift",""),
                    "time_from": "", "time_to": "", "total_time": "",
                    "bit_type": "", "diameter": "",
                    "metres_from": None, "metres_to": None, "total_metres": None,
                    "code": code, "notes": desc,
                    "rate_year": None, "unit_rate": None, "quantity": None,
                    "line_cost": None, "rate_basis": None, "po_id": None,
                })
            continue

        # Try the full HQ drilling line pattern
        rm = row_re.match(line)
        if rm:
            notes, tf, tt, dur, bt, diam, mf, mt, mto = rm.groups()
            # Convert decimal duration to H:MM
            dur_f = float(dur) if dur else 0
            h = int(dur_f)
            m_val = int((dur_f - h) * 60)
            total_time = f"{h}:{m_val:02d}"

            bit_type = ""
            if bt:
                btl = bt.strip().upper()
                if "HQ" in btl:
                    bit_type = "HQ_HQ3"
                elif "PCD" in btl:
                    bit_type = "PCD"
                elif "NQ" in btl:
                    bit_type = "NQ"

            rows.append({
                "source_file": filename, "contractor": contractor,
                "date": header.get("date",""), "hole_num": header.get("hole_num",""),
                "site_name": header.get("site_name",""), "location": header.get("location",""),
                "drill_rig": header.get("drill_rig",""), "client": header.get("client",""),
                "contract": header.get("contract",""), "shift": header.get("shift",""),
                "time_from": tf, "time_to": tt, "total_time": total_time,
                "bit_type": bit_type, "diameter": (diam or "").strip(),
                "metres_from": float(mf) if mf and float(mf) > 0 else None,
                "metres_to": float(mt) if mt and float(mt) > 0 else None,
                "total_metres": float(mto) if mto and float(mto) > 0 else None,
                "code": "", "notes": notes.strip(),
                "rate_year": None, "unit_rate": None, "quantity": None,
                "line_cost": None, "rate_basis": None, "po_id": None,
            })
            continue

        # Try simple pattern (most common: notes time_from time_to decimal mf mt md)
        sm = simple_re.match(line)
        if sm:
            notes, tf, tt, dur, mf, mt, mto = sm.groups()
            dur_f = float(dur) if dur else 0
            if dur_f == 0 and tf == "0" and tt == "0":
                continue
            h = int(dur_f)
            m_val = int((dur_f - h) * 60)
            total_time = f"{h}:{m_val:02d}"

            rows.append({
                "source_file": filename, "contractor": contractor,
                "date": header.get("date",""), "hole_num": header.get("hole_num",""),
                "site_name": header.get("site_name",""), "location": header.get("location",""),
                "drill_rig": header.get("drill_rig",""), "client": header.get("client",""),
                "contract": header.get("contract",""), "shift": header.get("shift",""),
                "time_from": tf, "time_to": tt, "total_time": total_time,
                "bit_type": "", "diameter": "",
                "metres_from": float(mf) if mf and float(mf) > 0 else None,
                "metres_to": float(mt) if mt and float(mt) > 0 else None,
                "total_metres": float(mto) if mto and float(mto) > 0 else None,
                "code": "", "notes": notes.strip(),
                "rate_year": None, "unit_rate": None, "quantity": None,
                "line_cost": None, "rate_basis": None, "po_id": None,
            })

    return rows


def parse_activities(text, header, filename, contractor):
    time_pat = r"\d{1,2}:\d{2}"

    # Pre-process text: join split lines where a code wraps (e.g. "Drill_Chip_or_Open_hol\n...e")
    lines_raw = text.splitlines()
    lines_joined = []
    i = 0
    while i < len(lines_raw):
        line = lines_raw[i].strip()
        # Check if this line is a partial code that continues on the next line
        # e.g. "Drill_Chip_or_Open_hol" followed by "e" or "drill to 243.55m 8:45..."
        if (line.endswith("_hol") or line.endswith("_Open_hol") or
            line.endswith("_Da") or line.endswith("_Day_Ra") or
            line.endswith("_Standby")):
            if i+1 < len(lines_raw):
                next_line = lines_raw[i+1].strip()
                # If next line starts with the rest of the code or has time data
                if re.match(r'^[a-z_]', next_line) and len(next_line) < 5:
                    # Just the end of a code like "e" or "te"
                    lines_joined.append(line + next_line)
                    i += 2
                    continue
                elif re.search(time_pat, next_line):
                    # Next line has the actual data — prepend the code
                    code_part = line.split()[-1] if line.split() else ""
                    lines_joined.append(next_line + " " + code_part)
                    i += 2
                    continue
        # Check if this line IS a dangling code fragment
        if re.match(r'^[a-z]\w*$', line) and len(line) < 5 and lines_joined:
            # Append to previous line
            lines_joined[-1] = lines_joined[-1] + line
            i += 1
            continue
        lines_joined.append(line)
        i += 1

    # Primary pattern: lines with 3 time fields (time_from, time_to, total_time)
    # Now handles PCD with diameter like: PCD 4" or PCD 5-7/8" or HQ_HQ3
    row_re = re.compile(
        r"^(.*?)(\d{1,2}:\d{2})\s+(\d{1,2}:\d{2})\s+(\d{1,2}:\d{2})"
        r'(?:\s+(HQ_HQ3|HQ|PCD|NQ|PQ)\s+([^\s]+))?'
        r"(?:\s+(\d+\.?\d*))?(?:\s+(\d+\.?\d*))?(?:\s+(\d+\.?\d*))?"
        r"(?:\s+([\w_]+))?\s*$", re.IGNORECASE)

    # Secondary pattern: lines with just a code and maybe numbers (day rates, consumables as activities)
    code_only_re = re.compile(
        r"^(.*?)\s+(D_\w+|H_\w+|PVC[\w\s]+|Cement[\w\s]*)\s*"
        r"(?:(\d+\.?\d*)\s+(\d+\.?\d*)\s+(\d+\.?\d*))?"
        r"\s*$", re.IGNORECASE)

    # Tertiary: lines with a code at the end and optional quantities but no times
    any_code_re = re.compile(
        r"^(.*?)\s+([\w_][\w_\-]+(?:\s*[\w_\-]+)*)\s*$"
    )

    rows = []
    known_codes = {
        "D_WaterCart_Day_Rate", "D_Water_Cart", "D_Water_Cart_Day_Rate",
        "D_Water_Cart_Standby", "D_Backhoe", "D_Backhoe_Day_Rate",
        "D_Backhoe_Standby", "H_Standby_Cement_Set", "H_Standby_Cement_set",
        "H_Standby_Sumps", "H_Standby_AAC", "H_Standby_Fatigue", "H_Standby_Logger",
        "H_Standby_Grout", "H_Water_Flow_Measure",
        "Drill_Core", "Drill_Chip_or_Open_hole",
        "H_Tripping_Rods", "H_Circulation_Flush", "H_Circulation_Lost",
        "H_Reaming", "H_Change_Drill_Mthd", "H_Surface_Setup",
        "H_Casing_Install", "H_Rig_Cementing", "H_Mud_Mixing",
        "H_Repairs", "H_Training", "H_Safety_Prestart",
        "H_Safety_Contractor", "Crew_Travel", "MOB", "DEMOB",
        "H_Con_Collect_Plan", "H_Crew_Travel_On", "H_Crew_Travel_Off",
    }

    for line in lines_joined:
        line = line.strip()
        if not line:
            continue

        # Try primary pattern (with times)
        time_count = len(re.findall(time_pat, line))
        if time_count >= 2:
            m = row_re.match(line)
            if m:
                notes, tf, tt, total, bt, diam, mf, mt, mto, code = m.groups()
                rows.append({
                    "source_file": filename, "contractor": contractor,
                    "date": header.get("date",""), "hole_num": header.get("hole_num",""),
                    "site_name": header.get("site_name",""), "location": header.get("location",""),
                    "drill_rig": header.get("drill_rig",""), "client": header.get("client",""),
                    "contract": header.get("contract",""), "shift": header.get("shift",""),
                    "time_from": tf, "time_to": tt, "total_time": total,
                    "bit_type": bt or "", "diameter": diam or "",
                    "metres_from": float(mf) if mf else None,
                    "metres_to": float(mt) if mt else None,
                    "total_metres": float(mto) if mto else None,
                    "code": code or "", "notes": notes.strip(),
                    "rate_year": None, "unit_rate": None, "quantity": None,
                    "line_cost": None, "rate_basis": None, "po_id": None,
                })
                continue

        # Try secondary: line contains a known code or day-rate pattern
        # Check if any known code appears in the line
        found_code = None
        for kc in known_codes:
            if kc in line or kc.replace("_", " ") in line:
                found_code = kc
                break

        # Also catch lines like "D_Backhoe - Day Rate" or "PVC Casing 100mm Class 9"
        if not found_code:
            if re.search(r"D_\w+|PVC\s+Casing|Water.?Cart|Backhoe", line, re.IGNORECASE):
                # Extract the code-like part
                cm = re.search(r"(D_\w+[\w\s\-]*|PVC\s+Casing[\w\s]*|H_\w+)", line)
                if cm:
                    found_code = cm.group(1).strip()

        if found_code:
            # Extract any numbers from the line
            nums = re.findall(r"(\d+\.?\d*)", line)
            # Try to find times
            times = re.findall(r"(\d{1,2}:\d{2})", line)
            tf = times[0] if len(times) > 0 else ""
            tt = times[1] if len(times) > 1 else ""
            total = times[2] if len(times) > 2 else ""

            # Get metres if present (numbers that aren't part of times)
            non_time_nums = [n for n in nums if ":" not in n and float(n) < 9000]
            mf = float(non_time_nums[0]) if len(non_time_nums) > 0 and float(non_time_nums[0]) > 0 else None
            mt = float(non_time_nums[1]) if len(non_time_nums) > 1 else None
            mto = float(non_time_nums[2]) if len(non_time_nums) > 2 else None

            # Clean up notes - remove the code from the line
            notes = line.replace(found_code, "").strip()
            notes = re.sub(r"\d{1,2}:\d{2}", "", notes).strip()
            notes = re.sub(r"\s+", " ", notes).strip()

            rows.append({
                "source_file": filename, "contractor": contractor,
                "date": header.get("date",""), "hole_num": header.get("hole_num",""),
                "site_name": header.get("site_name",""), "location": header.get("location",""),
                "drill_rig": header.get("drill_rig",""), "client": header.get("client",""),
                "contract": header.get("contract",""), "shift": header.get("shift",""),
                "time_from": tf, "time_to": tt, "total_time": total,
                "bit_type": "", "diameter": "",
                "metres_from": mf, "metres_to": mt, "total_metres": mto,
                "code": found_code, "notes": notes,
                "rate_year": None, "unit_rate": None, "quantity": None,
                "line_cost": None, "rate_basis": None, "po_id": None,
            })

    return rows


def parse_consumables(text, header, filename, contractor):
    rows = []
    m = re.search(r"CONSUMABLES\s*\n(.*?)(?:ALLIANZ REPRESENTATIVE|CONTRACTOR REPRESENTATIVE|$)", text, re.DOTALL|re.IGNORECASE)
    if not m: return rows
    block = m.group(1)
    # Match any consumable line - just grab the product name, set qty=1
    # Patterns: "AMC CR650 drum 1 Ltrs" or "AMC CR650 1" or "Fuel (Diesel) Ltrs"
    for line in block.splitlines():
        line = line.strip()
        if not line or line.lower().startswith('consumable') or line.lower().startswith('type') or line.lower().startswith('quantity'):
            continue
        # Skip lines that are just numbers or units
        if re.match(r'^[\d\s]+$', line) or re.match(r'^(Ltrs?|Kgs?|Mtrs?|drum|bucket|bags?)$', line, re.IGNORECASE):
            continue
        # Extract product name - strip trailing numbers, units, and quantity+unit combos
        product = re.sub(r'\s+(drum|bucket|bags?|tins?|slurry|Kgs?|Ltrs?|Mtrs?|cube|each)\s+\d+\s+\S+\s*$', '', line, flags=re.IGNORECASE).strip()
        product = re.sub(r'\s+(drum|bucket|bags?|tins?|slurry|Kgs?|Ltrs?|Mtrs?|cube|each)\s+\d+\s*$', '', product, flags=re.IGNORECASE).strip()
        product = re.sub(r'\s+(drum|bucket|bags?|tins?|slurry|Kgs?|Ltrs?|Mtrs?|cube|each)\s*$', '', product, flags=re.IGNORECASE).strip()
        product = re.sub(r'\s+\d+\s+(drum|bucket|bags?|tins?|slurry|Kgs?|Ltrs?|Mtrs?|cube|each)\s*$', '', product, flags=re.IGNORECASE).strip()
        product = re.sub(r'\s+\d+\s*$', '', product).strip()
        if not product or len(product) < 3:
            continue
        # Determine unit from original line
        unit_m = re.search(r'(drum|bucket|bags?|tins?|slurry|Kgs?|Ltrs?|Mtrs?|cube|each)', line, re.IGNORECASE)
        unit = unit_m.group(1) if unit_m else 'each'
        rows.append({
            "source_file": filename, "contractor": contractor,
            "date": header.get("date",""), "hole_num": header.get("hole_num",""),
            "site_name": header.get("site_name",""),
            "consumable": product, "type": product, "quantity": "1", "unit": unit,
            "unit_price": None, "line_cost": None,
        })
    return rows


def parse_crew(text, header, filename, contractor):
    rows = []
    for role in ["Rig Manager","Driller","Trainee Driller","Offsider","Operator"]:
        m = re.search(rf"{role}\s+([\w\s\.]+?)\s+(\d+)\s", text, re.IGNORECASE)
        if m:
            rows.append({"source_file":filename,"contractor":contractor,
                         "date":header.get("date",""),"hole_num":header.get("hole_num",""),
                         "site_name":header.get("site_name",""),
                         "role":role,"name":m.group(1).strip(),"hours":m.group(2)})
    return rows


COREPLAN_SECTIONS = {
    "Details","Drilling Intervals","Drilling Events","Minimum Drilling Costs",
    "Down Hole Activities","Survey","Bit Wear Clause","Time Breakdown",
    "Consumables","Equipment","Miscellaneous","People","Checklist"
}

COREPLAN_CATEGORY_CODES = {
    "Awaiting Site Preparation": "H_Standby_AAC",
    "Breakdown": "H_Repairs",
    "Cementing": "H_Rig_Cementing",
    "Circ/Flush_Hole": "H_Circulation_Flush",
    "Circulation - Lost Circulation": "H_Circulation_Lost",
    "Logging": "H_Standby_Logger",
    "Mobilisation": "MOB",
    "Other_Work_Rate": "H_Active",
    "Pack-up / Set-up": "H_Surface_Setup",
    "Pre-Start/Lube_Rig": "H_Safety_Prestart",
    "Reaming": "H_Reaming",
    "Run/Pull_Casing": "H_Casing_Install",
    "Safety/PSI Meeting": "H_Safety_Contractor",
    "Standby": "H_Standby_AAC",
    "Standby - Fatigue Management": "H_Standby_Fatigue",
    "Travel": "H_Crew_Travel_On",
    "Tripping Rods": "H_Tripping_Rods",
}


def coreplan_money(value):
    if value in (None, ""):
        return None
    s = str(value).replace("A$", "").replace("$", "").replace(",", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def coreplan_float(value):
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except Exception:
        return None


def hours_to_hhmm(value):
    hours = coreplan_float(value) or 0
    whole = int(hours)
    mins = int(round((hours - whole) * 60))
    if mins == 60:
        whole += 1
        mins = 0
    return f"{whole}:{mins:02d}"


def add_minutes(clock_minutes, hours):
    mins = int(round((coreplan_float(hours) or 0) * 60))
    return clock_minutes + mins


def clock_from_minutes(clock_minutes):
    mins = clock_minutes % (24 * 60)
    return f"{mins // 60}:{mins % 60:02d}"


def coreplan_bit_type(row):
    text = f"{row.get('type','')} {row.get('drill_bit','')} {row.get('drilling_events','')}".upper()
    if "HQ" in text:
        return "HQ_HQ3"
    if "NQ" in text:
        return "NQ"
    if "PQ" in text:
        return "PQ_PQ3"
    if "PCD" in text:
        return normalise_drilling_bit_key("PCD", notes=text)
    return ""


def coreplan_drill_code(row):
    bit = coreplan_bit_type(row)
    typ = str(row.get("type") or "").upper()
    if bit in ("HQ_HQ3", "NQ", PQ_PQ3_LABEL) or "COR" in typ:
        return "Drill_Core"
    return "Drill_Chip_or_Open_hole"


def coreplan_category_code(category, notes):
    if category == "Travel":
        n = (notes or "").lower()
        if "swipe off" in n or "off site" in n or "from site" in n:
            return "H_Crew_Travel_Off"
    return COREPLAN_CATEGORY_CODES.get(category, "H_Active")


def parse_coreplan_sections(text):
    lines = text.splitlines()
    sections = {}
    i = 0
    while i < len(lines):
        name = lines[i].strip("\ufeff").strip()
        if name in COREPLAN_SECTIONS:
            i += 1
            chunk = []
            while i < len(lines):
                nxt = lines[i].strip("\ufeff").strip()
                if nxt in COREPLAN_SECTIONS:
                    break
                if nxt:
                    chunk.append(lines[i])
                i += 1
            if chunk:
                try:
                    sections[name] = list(csv.DictReader(StringIO("\n".join(chunk))))
                except Exception:
                    sections[name] = []
            continue
        i += 1
    return sections


def parse_coreplan_plod_csv(content, filename, contractor):
    text = content.decode("utf-8-sig", errors="replace")
    sections = parse_coreplan_sections(text)
    details = (sections.get("Details") or [{}])[0]
    plod = details.get("plod") or os.path.splitext(filename)[0]
    report_date = details.get("date") or ""
    rig = details.get("rig") or ""
    shift = details.get("workshift") or ""
    contract = details.get("contract") or ""
    report_notes = details.get("notes") or ""
    total_cost = coreplan_money(details.get("total_cost")) or 0

    default_hole = ""
    for section_name in ("Drilling Intervals", "Time Breakdown", "Consumables", "People"):
        for row in sections.get(section_name) or []:
            default_hole = row.get("hole_name") or default_hole
            if default_hole:
                break
        if default_hole:
            break

    header = {
        "date": report_date,
        "hole_num": default_hole,
        "site_name": default_hole,
        "drill_rig": rig,
        "contract": contract,
        "shift": shift,
        "client": "Fitzroy Coal",
        "location": "",
        "plod": plod,
        "notes": report_notes,
    }

    acts = []
    current_time = 6 * 60

    def base_activity(row, code, notes, duration_hours=None, line_cost=None, unit_rate=None, quantity=None, rate_basis=None):
        nonlocal current_time
        duration_hours = coreplan_float(duration_hours) or 0
        time_from = clock_from_minutes(current_time) if duration_hours else ""
        time_to = clock_from_minutes(add_minutes(current_time, duration_hours)) if duration_hours else ""
        if duration_hours:
            current_time = add_minutes(current_time, duration_hours)
        hole = row.get("hole_name") or default_hole
        return {
            "source_file": filename, "contractor": contractor,
            "date": report_date, "hole_num": hole, "site_name": hole,
            "location": "", "drill_rig": rig, "client": "Fitzroy Coal",
            "contract": contract, "shift": shift,
            "time_from": time_from, "time_to": time_to, "total_time": hours_to_hhmm(duration_hours),
            "bit_type": "", "diameter": "",
            "metres_from": None, "metres_to": None, "total_metres": None,
            "code": code, "notes": notes or "",
            "rate_year": extract_year(report_date),
            "unit_rate": unit_rate, "quantity": quantity,
            "line_cost": line_cost, "rate_basis": rate_basis,
            "po_id": None,
        }

    interval_rows = sorted(sections.get("Drilling Intervals") or [], key=lambda r: coreplan_float(r.get("order")) or 0)
    for row in interval_rows:
        metres_from = coreplan_float(row.get("depth_from"))
        metres_to = coreplan_float(row.get("depth_to"))
        metres = None
        if metres_from is not None and metres_to is not None:
            metres = max(0, round(metres_to - metres_from, 2))
        if not metres and not coreplan_float(row.get("duration_hours")):
            continue
        unit_rate = coreplan_money(row.get("cost_per_m"))
        line_cost = coreplan_money(row.get("cost"))
        code = coreplan_drill_code(row)
        notes = "; ".join(x for x in [
            row.get("type") or "",
            row.get("drill_bit") or "",
            row.get("drilling_events") or "",
        ] if x)
        act = base_activity(
            row, code, notes, row.get("duration_hours"),
            line_cost=line_cost, unit_rate=unit_rate, quantity=metres,
            rate_basis=(f"CorePlan ${unit_rate:,.2f}/m x {metres:.2f}m" if unit_rate is not None and metres is not None else "CorePlan drilling interval")
        )
        act.update({
            "bit_type": coreplan_bit_type(row),
            "metres_from": metres_from,
            "metres_to": metres_to,
            "total_metres": metres,
        })
        acts.append(act)

    time_rows = sections.get("Time Breakdown") or []
    for row in time_rows:
        category = row.get("category") or ""
        notes = row.get("notes") or category
        duration = row.get("duration_hours")
        unit_rate = coreplan_money(row.get("cost_per_hour"))
        line_cost = coreplan_money(row.get("cost"))
        code = coreplan_category_code(category, notes)
        acts.append(base_activity(
            row, code, notes, duration,
            line_cost=line_cost, unit_rate=unit_rate,
            quantity=coreplan_float(duration),
            rate_basis=(f"CorePlan ${unit_rate:,.2f}/hr x {(coreplan_float(duration) or 0):.2f}h" if unit_rate is not None else "CorePlan time breakdown")
        ))

    for row in sections.get("Minimum Drilling Costs") or []:
        duration = row.get("duration_hours")
        unit_rate = coreplan_money(row.get("cost_per_h"))
        line_cost = coreplan_money(row.get("cost"))
        rig_type = row.get("rig_type") or "Rig"
        acts.append(base_activity(
            row, "H_Min_Shift", f"{rig_type} minimum shift charge", 0,
            line_cost=line_cost, unit_rate=unit_rate,
            quantity=coreplan_float(duration),
            rate_basis=(f"CorePlan minimum ${unit_rate:,.2f}/hr x {(coreplan_float(duration) or 0):.2f}h" if unit_rate is not None else "CorePlan minimum drilling cost")
        ))

    for row in sections.get("Miscellaneous") or []:
        name = row.get("name") or "Miscellaneous"
        line_cost = coreplan_money(row.get("cost"))
        qty = coreplan_float(row.get("quantity")) or 1
        if not name and line_cost is None:
            continue
        code = coreplan_category_code(name, row.get("notes") or name)
        unit_rate = round(line_cost / qty, 2) if line_cost is not None and qty else line_cost
        acts.append(base_activity(
            row, code, row.get("notes") or name, 0,
            line_cost=line_cost, unit_rate=unit_rate, quantity=qty,
            rate_basis="CorePlan miscellaneous charge"
        ))

    for row in sections.get("Equipment") or []:
        name = row.get("item_name") or row.get("name") or row.get("equipment") or "Equipment"
        line_cost = coreplan_money(row.get("cost"))
        duration = row.get("duration_hours") or row.get("quantity")
        qty = coreplan_float(duration) or coreplan_float(row.get("quantity")) or 1
        unit_rate = coreplan_money(row.get("cost_per_unit")) or coreplan_money(row.get("cost_per_day")) or coreplan_money(row.get("cost_per_hour"))
        if unit_rate is None and line_cost is not None and qty:
            unit_rate = round(line_cost / qty, 2)
        unit = row.get("unit") or row.get("billing_unit") or ""
        notes = row.get("notes") or name
        acts.append(base_activity(
            row, "E_Equipment", notes, duration,
            line_cost=line_cost, unit_rate=unit_rate, quantity=qty,
            rate_basis=("CorePlan equipment charge" + (f" ({unit})" if unit else ""))
        ))

    cons = []
    for row in sections.get("Consumables") or []:
        qty = coreplan_float(row.get("quantity"))
        unit_price = coreplan_money(row.get("cost_per_unit"))
        line_cost = coreplan_money(row.get("cost"))
        hole = row.get("hole_name") or default_hole
        item = row.get("item_name") or ""
        if not item:
            continue
        cons.append({
            "source_file": filename, "contractor": contractor,
            "date": report_date, "hole_num": hole, "site_name": hole,
            "consumable": item, "type": item,
            "quantity": "" if qty is None else str(qty),
            "unit": row.get("unit") or "",
            "unit_price": unit_price,
            "line_cost": line_cost,
        })

    crew = []
    for row in sections.get("People") or []:
        name = row.get("person_name") or ""
        if not name:
            continue
        hole = row.get("hole_name") or default_hole
        crew.append({
            "source_file": filename, "contractor": contractor,
            "date": report_date, "hole_num": hole, "site_name": hole,
            "role": row.get("job_role") or ("Supervisor" if str(row.get("is_supervisor")).lower() == "true" else "Crew"),
            "name": name,
            "hours": row.get("duration_hours") or "",
        })

    return header, acts, cons, crew, text


# ── API ───────────────────────────────────────────────────────────────────────
MCC_WEEKLY_HEADERS = {
    "Created by",
    "Shift Time Start",
    "Shift Time End",
    "Site Location",
    "Job Description - Role",
    "Job Description - Location",
    "Job Description - Hours",
    "Job Description - Description of Work Performed",
}


def _excel_dt(value):
    if value is None:
        return "", ""
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y"), value.strftime("%H:%M")
    return str(value), ""


def _decimal_hours_to_time(hours):
    try:
        total = int(round(float(hours) * 60))
    except Exception:
        return ""
    return f"{total // 60}:{total % 60:02d}"


def mcc_program_from_location(text):
    s = (text or "").lower()
    if "arg-002" in s or "gas riser" in s:
        return "Gas Riser"
    if "arg-003" in s or "sis" in s:
        return "SIS"
    if "arg-005" in s or "exploration" in s:
        return "Exploration"
    return ""


def mcc_hole_from_text(text):
    s = text or ""
    patterns = [
        r"\bIB[-\s]?(\d{2})[-\s]?(\d{3})\b",
        r"\bIB\s?(\d{2})[-\s]?(\d{2})\b",
        r"\bGR[-\s]?(\d{1,2})\b",
        r"\bSISMG\d{2}[-\s]?\d{2}[A-Z0-9]*\b",
        r"\bMG\d{2}[-\s]?\d{2}[A-Z0-9]*\b",
    ]
    for pat in patterns:
        m = re.search(pat, s, re.I)
        if not m:
            continue
        raw = m.group(0).upper().replace(" ", "-")
        if raw.startswith("IB") and len(m.groups()) >= 2:
            second = m.group(2)
            if len(second) == 2:
                second = second.zfill(3)
            return f"IB-{m.group(1)}-{second}"
        return raw.replace("--", "-")
    return ""


def parse_mcc_weekly_xlsx(content, filename, contractor="MCC Group"):
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise ValueError(f"openpyxl is not available: {e}")

    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    acts, crew, seen, source_lines = [], [], set(), []
    header = {"date": "", "hole_num": "", "site_name": "", "contractor": contractor}

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(v or "").strip() for v in rows[0]]
        if not MCC_WEEKLY_HEADERS.issubset(set(headers)):
            continue
        idx = {h: i for i, h in enumerate(headers)}
        for raw in rows[1:]:
            def cell(name):
                i = idx.get(name, -1)
                return raw[i] if i >= 0 and i < len(raw) else None

            created_by = cell("Created by")
            start = cell("Shift Time Start")
            end = cell("Shift Time End")
            site = cell("Site Location")
            role = cell("Job Description - Role")
            location = cell("Job Description - Location")
            hours = cell("Job Description - Hours")
            description = cell("Job Description - Description of Work Performed")
            equipment = cell("Job Description - Equipment")
            smu_start = cell("Job Description - SMU Start")
            smu_finish = cell("Job Description - SMU Finish")
            smu_total = cell("Job Description - SMU Total")
            if not any([created_by, start, end, site, role, location, hours, description, equipment]):
                continue

            date, time_from = _excel_dt(start)
            _, time_to = _excel_dt(end)
            program = mcc_program_from_location(location)
            hole = mcc_hole_from_text(" ".join([str(description or ""), str(location or "")]))
            source_key = (date, time_from, time_to, str(created_by or ""), str(location or ""), str(description or ""), str(equipment or ""))
            if source_key in seen:
                continue
            seen.add(source_key)

            notes_parts = [
                str(description or "").strip(),
                f"Role: {role}" if role else "",
                f"Program: {program}" if program else "",
                f"Workstream: {location}" if location else "",
                f"Equipment: {equipment}" if equipment else "",
                f"SMU: {smu_start} to {smu_finish} ({smu_total})" if equipment and (smu_start is not None or smu_finish is not None or smu_total is not None) else "",
                f"Created by: {created_by}" if created_by else "",
            ]
            try:
                hours_value = float(hours) if hours not in (None, "") else None
            except Exception:
                hours_value = None
            base_row = {
                "source_file": filename,
                "contractor": contractor,
                "date": date,
                "hole_num": hole,
                "site_name": str(site or "").strip() or hole,
                "program": program,
                "project": str(location or "").strip() or program,
                "location": str(location or "").strip(),
                "drill_rig": str(equipment or "").strip(),
                "client": "ARGO",
                "contract": str(location or "").strip(),
                "shift": "",
                "time_from": time_from,
                "time_to": time_to,
                "total_time": _decimal_hours_to_time(hours),
                "bit_type": "",
                "diameter": "",
                "metres_from": None,
                "metres_to": None,
                "total_metres": None,
                "rate_year": date[-4:] if len(date) >= 4 else "",
                "po_id": None,
            }
            priced_lines = []
            labour_rate = mcc_schedule_match(role, "labour")
            equipment_rate = mcc_schedule_match(equipment, "equipment")
            if labour_rate:
                priced_lines.append(("Labour", labour_rate))
            if equipment_rate:
                priced_lines.append(("Equipment", equipment_rate))
            if not priced_lines:
                row = {
                    **base_row,
                    "code": "H_Active",
                    "notes": " | ".join(p for p in notes_parts if p),
                    "unit_rate": None,
                    "quantity": hours_value,
                    "line_cost": None,
                    "rate_basis": "MCC weekly EOS import - unpriced",
                }
                acts.append(row)
            else:
                for charge_type, rate in priced_lines:
                    quantity = 1 if rate["unit"] == "day" else hours_value
                    line_cost = round(float(quantity or 0) * float(rate["rate"]), 2) if quantity is not None else None
                    row = {
                        **base_row,
                        "code": rate["code"],
                        "notes": " | ".join(p for p in notes_parts + [f"Charge type: {charge_type}", f"Schedule item: {rate['description']}"] if p),
                        "unit_rate": rate["rate"],
                        "quantity": quantity,
                        "line_cost": line_cost,
                        "rate_basis": f"MCC schedule {MCC_SCHEDULE_DATE} - {rate['description']} ({rate['unit']})",
                    }
                    acts.append(row)
            if created_by:
                crew.append({
                    "source_file": filename,
                    "contractor": contractor,
                    "date": date,
                    "hole_num": hole,
                    "site_name": base_row["site_name"],
                    "role": str(role or ""),
                    "name": str(created_by or ""),
                    "hours": str(hours or ""),
                })
            if not header["date"] and date:
                header.update({"date": date, "site_name": base_row["site_name"], "hole_num": hole, "contractor": contractor})
            source_lines.append(f"{date} {created_by or ''} {location or ''} {description or ''}")

    if not acts:
        raise ValueError("No MCC weekly EOS rows found in workbook")
    return header, acts, [], crew, "\n".join(source_lines[:500])


@app.get("/")
def root():
    return {"status": "ok", "app": "DrillOps API v3", "contractors": [c[0] for c in CONTRACTORS]}


@app.get("/contractors")
def get_contractors():
    """Return contractors from DB, ensuring shipped defaults exist."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            for name, code in CONTRACTORS:
                programs = "Exploration,Gas Riser,SIS" if name == "MCC Group" else "Exploration"
                category = DEFAULT_CONTRACTOR_CATEGORIES.get(name, "Misc")
                expense_gl = default_contractor_expense_gl(category)
                cur.execute("""
                    INSERT INTO contractors (name, short_code, category, program, expense_gl)
                    VALUES (%s, %s, %s, %s, %s) ON CONFLICT (name) DO NOTHING
                """, (name, code, category, programs, expense_gl))
            cur.execute("""
                UPDATE contractors
                SET short_code = COALESCE(NULLIF(short_code, ''), 'MCC'),
                    category = COALESCE(NULLIF(category, ''), 'Labour'),
                    program = 'Exploration,Gas Riser,SIS',
                    expense_gl = COALESCE(NULLIF(expense_gl, ''), '4250')
                WHERE name = 'MCC Group'
                  AND (
                    COALESCE(program, '') <> 'Exploration,Gas Riser,SIS'
                    OR COALESCE(expense_gl, '') = ''
                  )
            """)
            cur.execute("SELECT * FROM contractors ORDER BY program, name")
            rows = [dict(r) for r in cur.fetchall()]
        conn.commit()
    return rows


@app.post("/contractors")
async def add_contractor(request: Request):
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON body")
    name = payload.get("name", "").strip()
    code = payload.get("short_code", "").strip().upper() or name[:3].upper()
    category = str(payload.get("category") or "Misc").strip() or "Misc"
    if category not in CONTRACTOR_CATEGORIES:
        category = "Misc"
    expense_gl = str(payload.get("expense_gl") or default_contractor_expense_gl(category)).strip()
    if not re.fullmatch(r"\d{4}", expense_gl):
        raise HTTPException(400, "expense_gl must be a four-digit account code")
    raw_program = payload.get("programs", payload.get("program", "Exploration"))
    if isinstance(raw_program, list):
        program = ",".join(str(p).strip() for p in raw_program if str(p).strip()) or "Exploration"
    else:
        program = str(raw_program or "Exploration").strip() or "Exploration"
    raw_sites = payload.get("sites", payload.get("site", "Ironbark"))
    if isinstance(raw_sites, list):
        sites = ",".join(str(site).strip() for site in raw_sites if str(site).strip()) or "Ironbark"
    else:
        sites = str(raw_sites or "Ironbark").strip() or "Ironbark"
    if not name:
        raise HTTPException(400, "name is required")
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO contractors (name, short_code, category, program, sites, expense_gl, active)
                    VALUES (%s, %s, %s, %s, %s, %s, TRUE)
                    ON CONFLICT (name) DO UPDATE
                    SET short_code=EXCLUDED.short_code,
                        category=EXCLUDED.category,
                        program=EXCLUDED.program,
                        sites=EXCLUDED.sites,
                        expense_gl=EXCLUDED.expense_gl
                    RETURNING *
                """, (name, code, category, program, sites, expense_gl))
                row = dict(cur.fetchone())
            conn.commit()
        row["status"] = "created"
        return row
    except Exception as e:
        raise HTTPException(500, f"Failed to add contractor: {str(e)}")


def rename_contractor_references(cur, old_name: str, new_name: str):
    if not new_name or new_name == old_name:
        return
    for table in CONTRACTOR_REFERENCE_TABLES:
        cur.execute(f"UPDATE {table} SET contractor=%s WHERE contractor=%s", (new_name, old_name))


@app.patch("/contractors/{name}")
async def update_contractor(name: str, request: Request):
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON body")
    safe = {}
    new_name = str(payload.get("name") or "").strip()
    if new_name and new_name != name:
        safe["name"] = new_name
    if "short_code" in payload:
        safe["short_code"] = str(payload.get("short_code") or "").strip().upper()
    if "category" in payload:
        category = str(payload.get("category") or "Misc").strip() or "Misc"
        safe["category"] = category if category in CONTRACTOR_CATEGORIES else "Misc"
    if "expense_gl" in payload:
        expense_gl = str(payload.get("expense_gl") or "").strip()
        if not re.fullmatch(r"\d{4}", expense_gl):
            raise HTTPException(400, "expense_gl must be a four-digit account code")
        safe["expense_gl"] = expense_gl
    if "programs" in payload:
        raw_program = payload.get("programs")
        if isinstance(raw_program, list):
            safe["program"] = ",".join(str(p).strip() for p in raw_program if str(p).strip()) or "Exploration"
        else:
            safe["program"] = str(raw_program or "Exploration").strip() or "Exploration"
    elif "program" in payload:
        safe["program"] = str(payload.get("program") or "Exploration").strip() or "Exploration"
    if "sites" in payload:
        raw_sites = payload.get("sites")
        if isinstance(raw_sites, list):
            safe["sites"] = ",".join(str(site).strip() for site in raw_sites if str(site).strip()) or "Ironbark"
        else:
            safe["sites"] = str(raw_sites or "Ironbark").strip() or "Ironbark"
    elif "site" in payload:
        safe["sites"] = str(payload.get("site") or "Ironbark").strip() or "Ironbark"
    if not safe:
        raise HTTPException(400, "No valid fields")
    set_clause = ", ".join(f"{k}=%({k})s" for k in safe)
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(f"UPDATE contractors SET {set_clause} WHERE name=%(name)s RETURNING *", {**safe, "name": name})
                row = cur.fetchone()
                if not row:
                    raise HTTPException(404, "Contractor not found")
                if safe.get("name"):
                    rename_contractor_references(cur, name, safe["name"])
            conn.commit()
        return dict(row)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Failed to update contractor: {str(e)}")


@app.delete("/contractors/{name}")
def remove_contractor(name: str):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM contractors WHERE name=%s", (name,))
        conn.commit()
    return {"status": "deleted"}


def build_rate_context(hourly_rates=None, drilling_rates=None, consumable_rates=None):
    hourly = {}
    hourly_any = {}
    for r in hourly_rates or []:
        year = str(r.get("year") or "")
        code = r.get("code") or ""
        if not code:
            continue
        hourly[(year, code)] = float(r.get("rate") or 0)
        hourly_any.setdefault(code, float(r.get("rate") or 0))
    drilling = []
    for r in drilling_rates or []:
        drilling.append({
            "year": str(r.get("year") or ""),
            "bit_type": normalise_drilling_bit_key(r.get("bit_type") or ""),
            "depth_from": float(r.get("depth_from") or 0),
            "depth_to": float(r.get("depth_to") or 0),
            "rate": float(r.get("rate") or 0),
        })
    consumables = {}
    for r in consumable_rates or []:
        product = (r.get("product") or "").strip().upper()
        if product:
            consumables[product] = float(r.get("unit_price") or 0)
            consumables[product.replace(" ", "")] = float(r.get("unit_price") or 0)
    return {"hourly": hourly, "hourly_any": hourly_any, "drilling": drilling, "consumables": consumables}


def rate_year_for_row(row):
    for part in str(row.get("date") or "").replace("-", "/").split("/"):
        if len(part) == 4 and part.isdigit():
            return part
    return "2026"


def parse_row_hours(value):
    s = str(value or "").strip()
    if not s:
        return 0.0
    mt = re.match(r"^(\d+):(\d+)", s)
    if mt:
        return int(mt.group(1)) + int(mt.group(2)) / 60.0
    try:
        return float(s)
    except Exception:
        return 0.0


def row_num(value):
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def find_hourly_schedule_rate(code, year, rate_context):
    if not rate_context or not code:
        return None
    years = [year]
    if year.isdigit():
        years += [str(int(year) - 1), str(int(year) + 1)]
    years += ["2026", "2025"]
    for ty in years:
        val = rate_context["hourly"].get((ty, code))
        if val is not None:
            return val
    return rate_context["hourly_any"].get(code)


def drilling_schedule_key(row):
    return normalise_drilling_bit_key(row.get("bit_type"), row.get("code"), row.get("notes"))


def find_drilling_schedule_rate(row, rate_context):
    if not rate_context:
        return None
    mf, mt = row.get("metres_from"), row.get("metres_to")
    if mf is None or mt is None:
        return None
    depth = (row_num(mf) + row_num(mt)) / 2
    bit = drilling_schedule_key(row)
    year = rate_year_for_row(row)
    years = [year]
    if year.isdigit():
        years += [str(int(year) - 1), str(int(year) + 1)]
    years += ["2026", "2025"]
    for ty in years:
        for rate in rate_context["drilling"]:
            if rate["year"] == ty and rate["bit_type"] == bit and rate["depth_from"] <= depth < rate["depth_to"]:
                return rate["rate"]
    return None


def drilling_schedule_segments(row, rate_context):
    if not rate_context:
        return []
    mf, mt = row.get("metres_from"), row.get("metres_to")
    if mf is None or mt is None:
        return []
    start = row_num(mf)
    end = row_num(mt)
    if end <= start:
        return []
    bit = drilling_schedule_key(row)
    year = rate_year_for_row(row)
    years = [year]
    if year.isdigit():
        years += [str(int(year) - 1), str(int(year) + 1)]
    years += ["2026", "2025"]
    for ty in years:
        segments = []
        explicit_bit = explicit_pcd_range_label(f"{row.get('bit_type') or ''} {row.get('diameter') or ''} {row.get('notes') or ''} {row.get('code') or ''}")
        bands = [
            rate for rate in rate_context["drilling"]
            if rate["year"] == ty
            and rate["bit_type"] == bit
            and (not explicit_bit or explicit_pcd_range_label(rate.get("bit_type")) == explicit_bit)
        ]
        for rate in sorted(bands, key=lambda r: (r["depth_from"], r["depth_to"])):
            seg_from = max(start, rate["depth_from"])
            seg_to = min(end, rate["depth_to"])
            if seg_to > seg_from:
                metres = round(seg_to - seg_from, 2)
                segments.append({
                    "year": ty,
                    "bit_type": bit,
                    "depth_from": rate["depth_from"],
                    "depth_to": rate["depth_to"],
                    "metres": metres,
                    "rate": rate["rate"],
                    "cost": round(metres * rate["rate"], 2),
                })
        if segments:
            return segments
    return []


def drilling_schedule_cost(row, rate_context):
    segments = drilling_schedule_segments(row, rate_context)
    if not segments:
        rate = find_drilling_schedule_rate(row, rate_context)
        if rate is None:
            return None
        metres = row_num(row.get("total_metres"))
        return {
            "unit_rate": rate,
            "quantity": round(metres, 2),
            "line_cost": round(rate * metres, 2),
            "rate_basis": f"schedule ${rate:,.2f}/m x {metres:.2f}m ({drilling_schedule_key(row)})",
        }
    quantity = round(sum(s["metres"] for s in segments), 2)
    line_cost = round(sum(s["cost"] for s in segments), 2)
    unit_rate = round(line_cost / quantity, 4) if quantity else None
    parts = [
        f"{s['depth_from']:.0f}-{s['depth_to']:.0f}m ${s['rate']:,.2f}/m x {s['metres']:.2f}m"
        for s in segments
    ]
    return {
        "unit_rate": unit_rate,
        "quantity": quantity,
        "line_cost": line_cost,
        "rate_basis": f"schedule split ({drilling_schedule_key(row)}): " + "; ".join(parts),
    }


def calculate_activity_rate_fix(row, rate_context, suggested_code=None):
    code = suggested_code or row.get("code") or ""
    original_code = row.get("code") or ""
    year = rate_year_for_row(row)
    metres = row_num(row.get("total_metres"))
    hours = parse_row_hours(row.get("total_time"))
    mcc_fix = mcc_reprice_from_row(row)
    updates = {}
    reason = ""

    if metres > 0:
        priced = drilling_schedule_cost({**row, "code": code}, rate_context)
        if priced is None:
            return None
        updates.update({
            "rate_year": year,
            "unit_rate": priced["unit_rate"],
            "quantity": priced["quantity"],
            "line_cost": priced["line_cost"],
            "rate_basis": priced["rate_basis"],
        })
        reason = "Repriced drilled metres from drilling schedule."
    elif code in DRILLING_METRE_CODES:
        updates.update({
            "rate_year": year,
            "unit_rate": 0,
            "quantity": 0,
            "line_cost": 0,
            "rate_basis": "drilling time covered by metreage; no metres recorded",
        })
        reason = "Kept drilling time non-chargeable because drilling is billed by metres."
    elif is_not_chargeable_code(code, row.get("contractor", "")):
        qty = round(hours, 2) if hours > 0 else 1
        updates.update({"rate_year": year, "unit_rate": 0, "quantity": qty, "line_cost": 0, "rate_basis": "not chargeable"})
        reason = "Applied not-chargeable schedule code."
    elif mcc_fix:
        updates.update(mcc_fix)
        reason = "Repriced MCC weekly row from MCC schedule."
    elif code:
        rate = find_hourly_schedule_rate(code, year, rate_context)
        if rate is None:
            return None
        qty = round(hours, 2) if hours > 0 else 1
        updates.update({
            "rate_year": year,
            "unit_rate": rate,
            "quantity": qty,
            "line_cost": round(rate * qty, 2),
            "rate_basis": f"schedule ${rate:,.2f}/hr x {qty:.2f}",
        })
        reason = "Repriced activity from hourly schedule."

    if suggested_code and suggested_code != original_code:
        updates["code"] = suggested_code
        reason = f"Gemini matched code '{original_code}' to schedule code '{suggested_code}'. " + reason

    if not updates:
        return None

    changed = {}
    for key, new_val in updates.items():
        old_val = row.get(key)
        if isinstance(new_val, float):
            if abs(row_num(old_val) - new_val) > 0.01:
                changed[key] = new_val
        elif old_val != new_val:
            changed[key] = new_val
    if not changed:
        return None
    return {"updates": changed, "reason": reason.strip()}


async def gemini_suggest_schedule_codes(contractor, rows, hourly_rates):
    if not os.environ.get("GEMINI_API_KEY") or not rows:
        return {}
    schedule = [
        {"code": r.get("code"), "description": r.get("description"), "rate": r.get("rate"), "unit": r.get("unit")}
        for r in hourly_rates[:160]
    ]
    compact_rows = [
        {
            "id": r.get("id"),
            "current_code": r.get("code"),
            "notes": r.get("notes"),
            "time": r.get("total_time"),
            "metres": r.get("total_metres"),
            "rate_basis": r.get("rate_basis"),
            "line_cost": r.get("line_cost"),
        }
        for r in rows[:120]
    ]
    prompt = f"""You are correcting imported Allianz drilling EOS activity codes against a schedule of rates.

For each row, choose the most reasonable schedule code from the provided schedule. Only suggest a replacement when the current code is missing from the schedule or the notes clearly indicate a better schedule code. Do not invent codes.

Return ONLY valid JSON:
{{"suggestions":[{{"id":123,"suggested_code":"H_Example","confidence":0.0,"reason":"short reason"}}]}}

CONTRACTOR: {contractor}
SCHEDULE:
{json.dumps(schedule, indent=2, default=str)}

ROWS:
{json.dumps(compact_rows, indent=2, default=str)}
"""
    url = gemini_generate_content_url(os.environ.get("GEMINI_API_KEY", ""))
    payload = {"contents": [{"parts": [{"text": prompt}]}], "generationConfig": {"temperature": 0.1, "maxOutputTokens": 4096}}
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(url, json=payload)
        if resp.status_code != 200:
            return {}
        text = resp.json()["candidates"][0]["content"]["parts"][0]["text"].strip()
        if text.startswith("```"):
            text = text.split("\n", 1)[1] if "\n" in text else text[3:]
        if text.endswith("```"):
            text = text[:-3]
        text = text.strip()
        if text.lower().startswith("json"):
            text = text[4:].strip()
        parsed = json.loads(text)
        valid_codes = {r.get("code") for r in hourly_rates}
        out = {}
        for item in parsed.get("suggestions", []):
            code = item.get("suggested_code")
            rid = item.get("id")
            if code in valid_codes and rid is not None and float(item.get("confidence") or 0) >= 0.55:
                out[int(rid)] = {"code": code, "reason": item.get("reason", "")}
        return out
    except Exception:
        return {}


async def gemini_suggest_consumable_matches(contractor, rows, consumable_rates):
    if not os.environ.get("GEMINI_API_KEY") or not rows:
        return {}
    schedule = [
        {"product": r.get("product"), "description": r.get("description"), "unit_price": r.get("unit_price"), "unit": r.get("unit")}
        for r in consumable_rates[:180]
    ]
    compact_rows = [
        {"id": r.get("id"), "item": r.get("consumable") or r.get("type"), "type": r.get("type"), "quantity": r.get("quantity"), "unit": r.get("unit")}
        for r in rows[:120]
    ]
    prompt = f"""You are matching imported Allianz EOS consumables to a consumable schedule of rates.

For each imported consumable, choose an existing schedule product only when it is reasonably the same product despite spelling/casing/unit wording differences. If there is no reasonable match, return null for suggested_product so it can be added as a new consumable schedule item. Do not invent product names except by using the imported item itself when no match exists.

Return ONLY valid JSON:
{{"suggestions":[{{"id":123,"suggested_product":"Existing Product or null","confidence":0.0,"reason":"short reason"}}]}}

CONTRACTOR: {contractor}
CONSUMABLE SCHEDULE:
{json.dumps(schedule, indent=2, default=str)}

IMPORTED CONSUMABLES:
{json.dumps(compact_rows, indent=2, default=str)}
"""
    url = gemini_generate_content_url(os.environ.get("GEMINI_API_KEY", ""))
    payload = {"contents": [{"parts": [{"text": prompt}]}], "generationConfig": {"temperature": 0.1, "maxOutputTokens": 4096}}
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(url, json=payload)
        if resp.status_code != 200:
            return {}
        text = resp.json()["candidates"][0]["content"]["parts"][0]["text"].strip()
        if text.startswith("```"):
            text = text.split("\n", 1)[1] if "\n" in text else text[3:]
        if text.endswith("```"):
            text = text[:-3]
        text = text.strip()
        if text.lower().startswith("json"):
            text = text[4:].strip()
        parsed = json.loads(text)
        products = {(r.get("product") or "").strip().upper(): r for r in consumable_rates}
        out = {}
        for item in parsed.get("suggestions", []):
            rid = item.get("id")
            if rid is None:
                continue
            product = item.get("suggested_product")
            confidence = float(item.get("confidence") or 0)
            if product and product.strip().upper() in products and confidence >= 0.55:
                out[int(rid)] = {"product": products[product.strip().upper()], "reason": item.get("reason", "")}
            elif product is None or confidence < 0.55:
                out[int(rid)] = {"product": None, "reason": item.get("reason", "")}
        return out
    except Exception:
        return {}


def local_import_qa(acts, cons, crew, rate_context=None):
    def _num(value):
        try:
            return float(value or 0)
        except Exception:
            return 0.0

    def _hours(value):
        s = str(value or "").strip()
        if not s:
            return 0.0
        mt = re.match(r"^(\d+):(\d+)", s)
        if mt:
            return int(mt.group(1)) + int(mt.group(2)) / 60.0
        return _num(s)

    def _year(row):
        for part in str(row.get("date") or "").replace("-", "/").split("/"):
            if len(part) == 4 and part.isdigit():
                return part
        return "2026"

    def _hourly_rate(code, year):
        if not rate_context:
            return None
        for ty in [year, str(int(year)-1) if year.isdigit() else year, str(int(year)+1) if year.isdigit() else year, "2025", "2026"]:
            val = rate_context["hourly"].get((ty, code))
            if val is not None:
                return val
        return rate_context["hourly_any"].get(code)

    def _cost_mismatch(row, expected, basis):
        actual = row.get("line_cost")
        if actual is None or expected is None:
            return None
        delta = abs(_num(actual) - expected)
        if delta > 1.0:
            return f"Line cost {basis} mismatch: imported ${_num(actual):,.2f}, expected ${expected:,.2f}"
        return None

    warnings = []
    for i, row in enumerate(acts or [], 1):
        code = row.get("code") or ""
        metres = _num(row.get("total_metres"))
        mf = row.get("metres_from")
        mt = row.get("metres_to")
        line_cost = row.get("line_cost")
        if not row.get("date"):
            warnings.append({"severity": "critical", "row": i, "code": code, "issue": "Missing activity date", "recommendation": "Check the report header/date extraction."})
        if not row.get("hole_num") and not row.get("site_name"):
            warnings.append({"severity": "warning", "row": i, "code": code, "issue": "Missing both hole and site", "recommendation": "Confirm the hole/site mapping before relying on borehole summaries."})
        if metres > 0:
            try:
                interval = abs(_num(mt) - _num(mf)) if mf is not None and mt is not None else None
            except Exception:
                interval = None
            if interval is not None and abs(interval - metres) > 0.05:
                warnings.append({"severity": "critical", "row": i, "code": code, "issue": f"Metres interval mismatch: from/to gives {interval:.2f}m but total is {metres:.2f}m", "recommendation": "Check metres from, metres to, and total metres."})
            if not row.get("bit_type"):
                warnings.append({"severity": "warning", "row": i, "code": code, "issue": "Drilled metres row has no bit type", "recommendation": "Check the drilling line and rate card mapping."})
            if rate_context:
                priced = drilling_schedule_cost(row, rate_context)
                if priced is None:
                    warnings.append({"severity": "critical", "section": "rates", "row": i, "code": code, "issue": f"No drilling schedule rate found for {row.get('bit_type') or 'blank bit'} at {row.get('metres_from')} - {row.get('metres_to')} m", "recommendation": "Check the bit type/depth band against the drilling schedule of rates."})
                else:
                    mismatch = _cost_mismatch(row, priced["line_cost"], f"against {priced['rate_basis']}")
                    if mismatch:
                        warnings.append({"severity": "critical", "section": "rates", "row": i, "code": code, "issue": mismatch, "recommendation": "Recalculate the drilling charge from the schedule of rates."})
        if code.startswith("H_") and row.get("total_time") and line_cost is None:
            warnings.append({"severity": "warning", "row": i, "code": code, "issue": "Chargeable hourly-looking activity has no calculated cost", "recommendation": "Check the code against the hourly rate schedule."})
        if rate_context and metres <= 0 and code and not is_not_chargeable_code(code, row.get("contractor", "")):
            hours = _hours(row.get("total_time"))
            schedule_rate = _hourly_rate(code, _year(row))
            if schedule_rate is None and (code.startswith("H_") or line_cost is not None):
                warnings.append({"severity": "critical", "section": "rates", "row": i, "code": code, "issue": "Activity code is not found in the hourly schedule of rates", "recommendation": "Check whether the imported code is wrong or add the code to the rate schedule."})
            elif schedule_rate is not None:
                unit_rate = row.get("unit_rate")
                if unit_rate is None or abs(_num(unit_rate) - schedule_rate) > 0.01:
                    warnings.append({"severity": "critical", "section": "rates", "row": i, "code": code, "issue": f"Hourly unit rate does not match schedule: imported {unit_rate}, schedule ${schedule_rate:,.2f}", "recommendation": "Check the activity code and schedule year."})
                quantity = _num(row.get("quantity")) or hours or 1
                mismatch = _cost_mismatch(row, schedule_rate * quantity, f"against quantity {quantity:.2f} x ${schedule_rate:,.2f}")
                if mismatch:
                    warnings.append({"severity": "critical", "section": "rates", "row": i, "code": code, "issue": mismatch, "recommendation": "Recalculate the activity charge from the schedule of rates."})
        if metres > 300:
            warnings.append({"severity": "warning", "row": i, "code": code, "issue": f"Very large drilled metres value: {metres:.2f}m", "recommendation": "Check whether a depth was imported as metres drilled."})
    for i, row in enumerate(cons or [], 1):
        if (row.get("consumable") or row.get("type")) and row.get("line_cost") is None:
            warnings.append({"severity": "info", "section": "consumables", "row": i, "issue": "Consumable imported but not priced", "recommendation": "Check consumable rate setup or product name spelling."})
        if rate_context:
            product = (row.get("consumable") or row.get("type") or "").strip().upper()
            if product:
                schedule_rate = rate_context["consumables"].get(product) or rate_context["consumables"].get(product.replace(" ", ""))
                if schedule_rate is None:
                    warnings.append({"severity": "warning", "section": "rates", "row": i, "issue": f"Consumable '{product}' is not found in the consumable schedule of rates", "recommendation": "Check product spelling or add a consumable rate."})
                elif row.get("unit_price") is None or abs(_num(row.get("unit_price")) - schedule_rate) > 0.01:
                    warnings.append({"severity": "critical", "section": "rates", "row": i, "issue": f"Consumable unit price does not match schedule: imported {row.get('unit_price')}, schedule ${schedule_rate:,.2f}", "recommendation": "Check the consumable rate schedule."})
    return warnings


async def gemini_import_qa(filename, contractor, header, source_text, acts, cons, crew, rate_context=None):
    local_warnings = local_import_qa(acts, cons, crew, rate_context)
    if not os.environ.get("GEMINI_API_KEY"):
        return {"status": "unavailable", "summary": "Gemini import QA not run because GEMINI_API_KEY is not configured.", "warnings": local_warnings}

    compact_acts = [
        {k: r.get(k) for k in ["date","hole_num","site_name","time_from","time_to","total_time","bit_type","diameter","metres_from","metres_to","total_metres","code","notes","unit_rate","quantity","line_cost","rate_basis"]}
        for r in (acts or [])[:90]
    ]
    compact_cons = [
        {k: r.get(k) for k in ["date","hole_num","site_name","consumable","type","quantity","unit","unit_price","line_cost"]}
        for r in (cons or [])[:40]
    ]
    compact_crew = [
        {k: r.get(k) for k in ["date","hole_num","site_name","role","name","hours"]}
        for r in (crew or [])[:30]
    ]
    prompt = f"""You are checking an Allianz drilling EOS PDF import before it is trusted in DrillOps.

Look for extraction/parsing/rating errors in the parsed data. Focus on:
- missing or wrong date, hole, site, rig, or shift
- time rows out of sequence, impossible durations, duplicate rows
- metres_from/metres_to/total_metres mismatches
- activity codes that do not appear to match a schedule-of-rates code
- drilled metres rows where bit type/depth band/rate/cost do not match the schedule of rates
- standby/day-rate/consumable rows where the charge does not match the schedule of rates
- line_cost that looks inconsistent with quantity, metres, hours, unit_rate, or rate_basis
- notes that suggest a code should be different

Return ONLY valid JSON:
{{
  "status": "ok or needs_review",
  "summary": "short summary",
  "warnings": [
    {{"severity":"critical|warning|info","section":"activities|consumables|crew|header","row":1,"code":"optional","issue":"what looks wrong","recommendation":"what the user should check"}}
  ]
}}

FILENAME: {filename}
CONTRACTOR: {contractor}
HEADER:
{json.dumps(header, default=str)}

PDF TEXT EXCERPT:
{source_text[:5000]}

PARSED ACTIVITIES:
{json.dumps(compact_acts, indent=2, default=str)}

PARSED CONSUMABLES:
{json.dumps(compact_cons, indent=2, default=str)}

PARSED CREW:
{json.dumps(compact_crew, indent=2, default=str)}
"""
    url = gemini_generate_content_url(os.environ.get("GEMINI_API_KEY", ""))
    payload = {"contents": [{"parts": [{"text": prompt}]}], "generationConfig": {"temperature": 0.1, "maxOutputTokens": 4096}}
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(url, json=payload)
        if resp.status_code != 200:
            return {"status": "partial", "summary": f"Gemini import QA failed: HTTP {resp.status_code}", "warnings": local_warnings}
        text = resp.json()["candidates"][0]["content"]["parts"][0]["text"].strip()
        if text.startswith("```"):
            text = text.split("\n", 1)[1] if "\n" in text else text[3:]
        if text.endswith("```"):
            text = text[:-3]
        text = text.strip()
        if text.lower().startswith("json"):
            text = text[4:].strip()
        result = json.loads(text)
        result["warnings"] = (result.get("warnings") or []) + local_warnings
        if result.get("warnings") and result.get("status") == "ok":
            result["status"] = "needs_review"
        return result
    except Exception as e:
        return {"status": "partial", "summary": f"Gemini import QA could not parse a result: {str(e)}", "warnings": local_warnings}


def infer_report_contractor(selected_contractor: str, filename: str = "", header: dict = None, source_text: str = "") -> str:
    header = header or {}
    report_haystack = " ".join([
        filename or "",
        header.get("drill_rig") or "",
        header.get("contract") or "",
        header.get("client") or "",
        source_text[:4000] if source_text else "",
    ]).upper()
    filename_upper = str(filename or "").upper().strip()
    rig = str(header.get("drill_rig") or "").upper().strip()
    # Report identity is stronger evidence than the upload card selected in the UI.
    # In particular, ADR001 PDFs must never be filed under Mitchells accidentally.
    if filename_upper.startswith("ADR001") or rig.startswith(("ADR", "ALZ")) or "ALLIANZ" in report_haystack:
        return "Allianz Drilling"
    if rig in {"IB652C"} or "MITCHELL" in report_haystack or "COREPLAN" in report_haystack:
        return "Mitchells Drilling"
    return selected_contractor or "Allianz Drilling"


def apply_import_activity_scope(rows: list[dict], contractor: str, program: str = "") -> list[dict]:
    """Ensure imported activity rows participate in program/project filtering."""
    default_program = str(program or "").strip()
    if not default_program:
        default_program = "Exploration" if contractor != "MCC Group" else ""
    for row in rows:
        if not str(row.get("program") or "").strip():
            row["program"] = default_program
        if "project" not in row:
            row["project"] = None
    return rows


@app.post("/import")
async def import_pdf(
    file: UploadFile = File(...),
    contractor: str = Form(default="Allianz Drilling"),
):
    filename = file.filename
    content = await file.read()
    if filename.lower().endswith(".xlsx"):
        try:
            contractor = "MCC Group"
            header, acts, cons, crew, source_text = parse_mcc_weekly_xlsx(content, filename, contractor)
        except Exception as e:
            raise HTTPException(400, f"Could not read MCC weekly XLSX: {e}")

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM imported_files WHERE filename=%s AND contractor=%s",
                            (filename, contractor))
                if cur.fetchone():
                    return {"status":"skipped","filename":filename,"rows":0,"contractor":contractor}
                if acts:
                    psycopg2.extras.execute_batch(cur, """
                        INSERT INTO activities
                        (source_file,contractor,date,hole_num,site_name,program,project,location,drill_rig,
                         client,contract,shift,time_from,time_to,total_time,bit_type,diameter,
                         metres_from,metres_to,total_metres,code,notes,
                         rate_year,unit_rate,quantity,line_cost,rate_basis,po_id)
                        VALUES
                        (%(source_file)s,%(contractor)s,%(date)s,%(hole_num)s,%(site_name)s,%(program)s,%(project)s,
                         %(location)s,%(drill_rig)s,%(client)s,%(contract)s,%(shift)s,
                         %(time_from)s,%(time_to)s,%(total_time)s,%(bit_type)s,%(diameter)s,
                         %(metres_from)s,%(metres_to)s,%(total_metres)s,%(code)s,%(notes)s,
                         %(rate_year)s,%(unit_rate)s,%(quantity)s,%(line_cost)s,%(rate_basis)s,%(po_id)s)
                    """, acts)
                if crew:
                    psycopg2.extras.execute_batch(cur, """
                        INSERT INTO crew (source_file,contractor,date,hole_num,site_name,role,name,hours)
                        VALUES (%(source_file)s,%(contractor)s,%(date)s,%(hole_num)s,%(site_name)s,%(role)s,%(name)s,%(hours)s)
                    """, crew)
                cur.execute("INSERT INTO imported_files (filename,contractor) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                            (filename, contractor))
                cur.execute("""
                    INSERT INTO source_files (filename, contractor, file_type, pdf_data)
                    VALUES (%s, %s, 'mcc_weekly_xlsx', %s) ON CONFLICT (filename, contractor) DO NOTHING
                """, (filename, contractor, psycopg2.Binary(content)))
            conn.commit()

        return {"status":"imported","filename":filename,"rows":len(acts),
                "contractor":contractor,
                "total_cost":round(sum(float(a.get("line_cost") or 0) for a in acts), 2),
                "consumables":0,"crew":len(crew),
                "import_check":{"status":"ok","summary":"MCC Group weekly end-of-shift XLSX imported. Labour and equipment rows are priced from the MCC schedule where matched, and separated by ARG workstream/program.","warnings":[]}}

    if filename.lower().endswith(".csv"):
        try:
            header, acts, cons, crew, source_text = parse_coreplan_plod_csv(content, filename, "Mitchells Drilling")
            contractor = infer_report_contractor(contractor, filename, header, source_text)
            header, acts, cons, crew, source_text = parse_coreplan_plod_csv(content, filename, contractor)
        except Exception as e:
            raise HTTPException(400, f"Could not read CorePlan CSV: {e}")

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM imported_files WHERE filename=%s AND contractor=%s",
                            (filename, contractor))
                if cur.fetchone():
                    return {"status":"skipped","filename":filename,"rows":0,"contractor":contractor}

        acts = restore_coreplan_activity_line_costs(acts)
        acts = adjust_imported_minimum_shift_rows(acts, contractor)
        acts = apply_allianz_minimum_shift_topups_to_rows(acts)
        acts = apply_import_activity_scope(acts, contractor)

        with get_conn() as conn:
            with conn.cursor() as cur:
                if acts:
                    psycopg2.extras.execute_batch(cur, """
                        INSERT INTO activities
                        (source_file,contractor,date,hole_num,site_name,program,project,location,drill_rig,
                         client,contract,shift,time_from,time_to,total_time,bit_type,diameter,
                         metres_from,metres_to,total_metres,code,notes,
                         rate_year,unit_rate,quantity,line_cost,rate_basis,po_id)
                        VALUES
                        (%(source_file)s,%(contractor)s,%(date)s,%(hole_num)s,%(site_name)s,%(program)s,%(project)s,
                         %(location)s,%(drill_rig)s,%(client)s,%(contract)s,%(shift)s,
                         %(time_from)s,%(time_to)s,%(total_time)s,%(bit_type)s,%(diameter)s,
                         %(metres_from)s,%(metres_to)s,%(total_metres)s,%(code)s,%(notes)s,
                         %(rate_year)s,%(unit_rate)s,%(quantity)s,%(line_cost)s,%(rate_basis)s,%(po_id)s)
                    """, acts)
                if cons:
                    psycopg2.extras.execute_batch(cur, """
                        INSERT INTO consumables (source_file,contractor,date,hole_num,site_name,consumable,type,quantity,unit,unit_price,line_cost)
                        VALUES (%(source_file)s,%(contractor)s,%(date)s,%(hole_num)s,%(site_name)s,%(consumable)s,%(type)s,%(quantity)s,%(unit)s,%(unit_price)s,%(line_cost)s)
                    """, cons)
                if crew:
                    psycopg2.extras.execute_batch(cur, """
                        INSERT INTO crew (source_file,contractor,date,hole_num,site_name,role,name,hours)
                        VALUES (%(source_file)s,%(contractor)s,%(date)s,%(hole_num)s,%(site_name)s,%(role)s,%(name)s,%(hours)s)
                    """, crew)
                cur.execute("INSERT INTO imported_files (filename,contractor) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                            (filename, contractor))
                cur.execute("""
                    INSERT INTO source_files (filename, contractor, file_type, pdf_data)
                    VALUES (%s, %s, 'coreplan_csv', %s) ON CONFLICT (filename, contractor) DO NOTHING
                """, (filename, contractor, psycopg2.Binary(content)))
            conn.commit()

        return {"status":"imported","filename":filename,"rows":len(acts),
                "contractor":contractor,
                "total_cost":round(sum(r["line_cost"] for r in acts if r["line_cost"]),2),
                "consumables":len(cons),"crew":len(crew),
                "import_check":{"status":"ok","summary":"CorePlan CSV imported using Mitchells/CorePlan structured export.","warnings":[]}}

    try:
        with pdfplumber.open(BytesIO(content)) as pdf:
            text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    except Exception as e:
        raise HTTPException(400, f"Could not read PDF: {e}")

    header = parse_header(text)
    fmt = detect_pdf_format(text)
    import_check = None
    contractor = infer_report_contractor(contractor, filename, header, text)
    if fmt == "adr001":
        acts = parse_activities_adr001(text, header, filename, contractor)
    else:
        acts = parse_activities(text, header, filename, contractor)
    cons = parse_consumables(text, header, filename, contractor)
    crew = parse_crew(text, header, filename, contractor)
    acts = apply_import_activity_scope(acts, contractor)

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT 1 FROM imported_files WHERE filename=%s AND contractor=%s",
                        (filename, contractor))
            if cur.fetchone():
                return {"status":"skipped","filename":filename,"rows":0,"contractor":contractor}

    with get_conn() as conn:
        with conn.cursor() as cur:
            # Load all rates into memory for fast pricing (same as reprice)
            cur.execute("SELECT * FROM drilling_rates WHERE contractor=%s", (contractor,))
            all_dr = [dict(r) for r in cur.fetchall()]
            cur.execute("SELECT * FROM hourly_rates WHERE contractor=%s", (contractor,))
            all_hr = [dict(r) for r in cur.fetchall()]
            cur.execute("SELECT * FROM consumable_rates WHERE contractor=%s", (contractor,))
            all_cr = [dict(r) for r in cur.fetchall()]

        hr_lookup = {}
        for r in all_hr:
            hr_lookup[(r["year"], r["code"])] = float(r["rate"])
        dr_lookup = {}
        for r in all_dr:
            key = (r["year"], normalise_drilling_bit_key(r["bit_type"]))
            if key not in dr_lookup: dr_lookup[key] = []
            dr_lookup[key].append((float(r["depth_from"]), float(r["depth_to"]), float(r["rate"])))
        cr_lookup = {}
        for r in all_cr:
            cr_lookup[r["product"].strip().upper()] = float(r["unit_price"])
            cr_lookup[r["product"].strip().upper().replace(" ","")] = float(r["unit_price"])
        rate_context = build_rate_context(all_hr, all_dr, all_cr)

        def _get_hr(code, year):
            for ty in [year, str(int(year)-1) if year.isdigit() else year, str(int(year)+1) if year.isdigit() else year, "2025"]:
                r = hr_lookup.get((ty, code))
                if r is not None: return r
            return None

        def _get_dr(bit_key, depth, year):
            for ty in [year, str(int(year)-1) if year.isdigit() else year, str(int(year)+1) if year.isdigit() else year, "2025"]:
                for frm, to, rate in dr_lookup.get((ty, bit_key), []):
                    if frm <= depth < to: return rate
            return None

        def _price_row(row):
            code = row.get("code","") or ""
            total_time = row.get("total_time","") or ""
            hours = 0
            mt = re.match(r"(\d+):(\d+)", str(total_time))
            if mt: hours = int(mt.group(1)) + int(mt.group(2))/60.0
            else:
                try: hours = float(total_time)
                except: pass

            bit_type = row.get("bit_type","") or ""
            metres = 0
            try: metres = float(row.get("total_metres",0) or 0)
            except: pass
            depth = None
            mf, mto = row.get("metres_from"), row.get("metres_to")
            if mf is not None and mto is not None:
                try: depth = (float(mf) + float(mto)) / 2
                except: pass

            date_str = row.get("date","") or ""
            year = "2026"
            for p in date_str.replace("-","/").split("/"):
                if len(p)==4 and p.isdigit(): year = p; break

            lc = None; ur = None; qty = None; rb = None

            # Drilling metres
            if metres > 0 and bit_type and depth is not None:
                priced = drilling_schedule_cost(row, rate_context)
                if priced is not None:
                    ur = priced["unit_rate"]
                    qty = priced["quantity"]
                    lc = priced["line_cost"]
                    rb = priced["rate_basis"]
            elif code in DRILLING_METRE_CODES:
                ur = 0; qty = 0; lc = 0; rb = "drilling time covered by metreage; no metres recorded"

            # Day rates
            if lc is None:
                for dk, (dk_code, dk_unit) in DAY_RATE_CODES.items():
                    if dk in code or code in dk:
                        r = _get_hr(dk_code, year)
                        if r is not None: ur = r; qty = 1; lc = r; rb = f"${r:,.2f}/{dk_unit}"
                        break

            # Not chargeable
            if lc is None and is_not_chargeable_code(code, contractor):
                ur = 0; qty = round(hours,2) if hours > 0 else 1; lc = 0; rb = "not chargeable"

            # Equipment packages
            if lc is None and code.startswith("E_"):
                r = _get_hr(code, year)
                q = round(hours,2) if hours > 0 else (coreplan_float(row.get("quantity")) or 1)
                if r is not None:
                    ur = r; qty = q; lc = round(r*q,2); rb = f"equipment ${r:,.2f}/unit x {q}"
                elif row.get("line_cost") is not None:
                    ur = row.get("unit_rate"); qty = row.get("quantity") or q
                    lc = row.get("line_cost"); rb = row.get("rate_basis") or "equipment charge"

            # Standby/inactive
            if lc is None and (code in STANDBY_CODES or "Standby" in code or code in INACTIVE_CODES):
                r = _get_hr(code, year)
                if r is None:
                    r = _get_hr("H_Inactive", year)
                if r is not None:
                    q = round(hours,2) if hours > 0 else 1
                    ur = r; qty = q; lc = round(r*q,2); rb = f"inactive ${r:,.2f}/hr x {q}"

            # Active
            if lc is None and hours > 0 and (code in ACTIVE_CODES or code.startswith("H_")):
                r = _get_hr(code, year)
                if r is None:
                    r = _get_hr("H_Active", year)
                if r is not None:
                    ur = r; qty = round(hours,2); lc = round(r*hours,2); rb = f"active ${r:,.2f}/hr x {hours:.2f}h"

            # Fallback
            if lc is None and hours > 0 and code:
                r = _get_hr(code, year)
                if r is None:
                    r = _get_hr("H_Active", year)
                if r is not None:
                    ur = r; qty = round(hours,2); lc = round(r*hours,2); rb = f"fallback ${r:,.2f}/hr x {hours:.2f}h"

            row["rate_year"] = year; row["unit_rate"] = ur; row["quantity"] = qty
            row["line_cost"] = lc; row["rate_basis"] = rb
            return row

        # Price all activities in memory
        acts = [_price_row(row) for row in acts]
        acts = adjust_imported_minimum_shift_rows(acts, contractor)
        acts = apply_allianz_minimum_shift_topups_to_rows(acts)

        # Price consumables
        for c in cons:
            if contractor == "Mitchells Drilling" and c.get("line_cost") is not None:
                continue
            product = (c.get("consumable") or c.get("type") or "").strip().upper()
            price = cr_lookup.get(product) or cr_lookup.get(product.replace(" ",""))
            if price is None:
                for rk, rv in cr_lookup.items():
                    if rk in product or product in rk:
                        price = rv; break
            if price is not None and price > 0:
                qty = 1
                try: qty = float(c.get("quantity") or 1)
                except: pass
                c["unit_price"] = price
                c["line_cost"] = round(price * qty, 2)
            elif c.get("line_cost") is None:
                c["unit_price"] = None
                c["line_cost"] = None

        if contractor == "Allianz Drilling":
            import_check = await gemini_import_qa(filename, contractor, header, text, acts, cons, crew, rate_context)

        with conn.cursor() as cur:
            if acts:
                psycopg2.extras.execute_batch(cur, """
                    INSERT INTO activities
                    (source_file,contractor,date,hole_num,site_name,program,project,location,drill_rig,
                     client,contract,shift,time_from,time_to,total_time,bit_type,diameter,
                     metres_from,metres_to,total_metres,code,notes,
                     rate_year,unit_rate,quantity,line_cost,rate_basis,po_id)
                    VALUES
                    (%(source_file)s,%(contractor)s,%(date)s,%(hole_num)s,%(site_name)s,%(program)s,%(project)s,
                     %(location)s,%(drill_rig)s,%(client)s,%(contract)s,%(shift)s,
                     %(time_from)s,%(time_to)s,%(total_time)s,%(bit_type)s,%(diameter)s,
                     %(metres_from)s,%(metres_to)s,%(total_metres)s,%(code)s,%(notes)s,
                     %(rate_year)s,%(unit_rate)s,%(quantity)s,%(line_cost)s,%(rate_basis)s,%(po_id)s)
                """, acts)
            if cons:
                psycopg2.extras.execute_batch(cur, """
                    INSERT INTO consumables (source_file,contractor,date,hole_num,site_name,consumable,type,quantity,unit,unit_price,line_cost)
                    VALUES (%(source_file)s,%(contractor)s,%(date)s,%(hole_num)s,%(site_name)s,%(consumable)s,%(type)s,%(quantity)s,%(unit)s,%(unit_price)s,%(line_cost)s)
                """, cons)
            if crew:
                psycopg2.extras.execute_batch(cur, """
                    INSERT INTO crew (source_file,contractor,date,hole_num,site_name,role,name,hours)
                    VALUES (%(source_file)s,%(contractor)s,%(date)s,%(hole_num)s,%(site_name)s,%(role)s,%(name)s,%(hours)s)
                """, crew)
            cur.execute("INSERT INTO imported_files (filename,contractor) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                        (filename, contractor))
            cur.execute("""
                INSERT INTO source_files (filename, contractor, file_type, pdf_data)
                VALUES (%s, %s, 'eos', %s) ON CONFLICT (filename, contractor) DO NOTHING
            """, (filename, contractor, psycopg2.Binary(content)))
        conn.commit()

    return {"status":"imported","filename":filename,"rows":len(acts),
            "contractor":contractor,
            "total_cost":round(sum(r["line_cost"] for r in acts if r["line_cost"]),2),
            "import_check":import_check}


@app.post("/imports/qa-existing")
async def qa_existing_imports(request: Request):
    payload = await request.json()
    contractor = payload.get("contractor", "Allianz Drilling")
    limit = int(payload.get("limit") or 25)
    use_gemini = bool(payload.get("use_gemini", True))
    if limit < 1:
        limit = 1
    if limit > 500:
        limit = 500

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM hourly_rates WHERE contractor=%s", (contractor,))
            all_hr = [dict(r) for r in cur.fetchall()]
            cur.execute("SELECT * FROM drilling_rates WHERE contractor=%s", (contractor,))
            all_dr = [dict(r) for r in cur.fetchall()]
            cur.execute("SELECT * FROM consumable_rates WHERE contractor=%s", (contractor,))
            all_cr = [dict(r) for r in cur.fetchall()]
            rate_context = build_rate_context(all_hr, all_dr, all_cr)

            cur.execute("""
                SELECT source_file, MAX(date) AS report_date, COUNT(*) AS rows
                FROM activities
                WHERE contractor=%s AND COALESCE(source_file,'') <> ''
                GROUP BY source_file
                ORDER BY MAX(date) DESC NULLS LAST, source_file
                LIMIT %s
            """, (contractor, limit))
            files = [dict(r) for r in cur.fetchall()]

            results = []
            for item in files:
                filename = item["source_file"]
                cur.execute("SELECT * FROM activities WHERE contractor=%s AND source_file=%s ORDER BY date,time_from,id", (contractor, filename))
                acts = [dict(r) for r in cur.fetchall()]
                cur.execute("SELECT * FROM consumables WHERE contractor=%s AND source_file=%s ORDER BY id", (contractor, filename))
                cons = [dict(r) for r in cur.fetchall()]
                cur.execute("SELECT * FROM crew WHERE contractor=%s AND source_file=%s ORDER BY id", (contractor, filename))
                crew = [dict(r) for r in cur.fetchall()]
                cur.execute("SELECT pdf_data FROM source_files WHERE contractor=%s AND filename=%s", (contractor, filename))
                source = cur.fetchone()

                text = ""
                if source and source.get("pdf_data"):
                    try:
                        with pdfplumber.open(BytesIO(bytes(source["pdf_data"]))) as pdf:
                            text = "\n".join(p.extract_text() or "" for p in pdf.pages)
                    except Exception:
                        text = ""

                header = parse_header(text) if text else {
                    "date": acts[0].get("date") if acts else "",
                    "hole_num": acts[0].get("hole_num") if acts else "",
                    "site_name": acts[0].get("site_name") if acts else "",
                }
                check = await gemini_import_qa(filename, contractor, header, text, acts, cons, crew, rate_context) if use_gemini else {
                    "status": "local",
                    "summary": "Local schedule-of-rates checks only.",
                    "warnings": local_import_qa(acts, cons, crew, rate_context),
                }
                results.append({
                    "filename": filename,
                    "date": item.get("report_date"),
                    "rows": item.get("rows"),
                    "check": check,
                })

    issue_count = sum(len((r.get("check") or {}).get("warnings") or []) for r in results)
    review_count = sum(1 for r in results if ((r.get("check") or {}).get("warnings") or []))
    return {
        "status": "ok",
        "contractor": contractor,
        "checked": len(results),
        "needs_review": review_count,
        "issues": issue_count,
        "results": results,
    }


@app.post("/imports/ai-fix-rates")
async def ai_fix_import_rates(request: Request):
    payload = await request.json()
    contractor = payload.get("contractor", "Allianz Drilling")
    limit = int(payload.get("limit") or 500)
    apply_changes = bool(payload.get("apply", True))
    if limit < 1:
        limit = 1
    if limit > 1000:
        limit = 1000

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM hourly_rates WHERE contractor=%s", (contractor,))
            all_hr = [dict(r) for r in cur.fetchall()]
            cur.execute("SELECT * FROM drilling_rates WHERE contractor=%s", (contractor,))
            all_dr = [dict(r) for r in cur.fetchall()]
            cur.execute("SELECT * FROM consumable_rates WHERE contractor=%s", (contractor,))
            all_cr = [dict(r) for r in cur.fetchall()]
            rate_context = build_rate_context(all_hr, all_dr, all_cr)
            locked_keys = locked_activity_sheet_keys(cur, contractor)

            cur.execute("""
                SELECT * FROM activities
                WHERE contractor=%s
                ORDER BY date DESC NULLS LAST, source_file, time_from, id
                LIMIT %s
            """, (contractor, limit))
            activities = [dict(r) for r in cur.fetchall()]
            locked_activity_count = sum(1 for r in activities if row_is_in_locked_sheet(r, locked_keys))
            activities = [r for r in activities if not row_is_in_locked_sheet(r, locked_keys)]

            hourly_codes = set(rate_context["hourly_any"].keys()) | set(NOT_CHARGEABLE) | set(MITCHELLS_NOT_CHARGEABLE)
            fuzzy_activity_rows = [
                r for r in activities
                if row_num(r.get("total_metres")) <= 0
                and (r.get("code") or "")
                and (r.get("code") or "") not in hourly_codes
                and ((r.get("line_cost") is not None) or (r.get("notes") or ""))
            ]
            code_suggestions = await gemini_suggest_schedule_codes(contractor, fuzzy_activity_rows, all_hr)

            activity_changes = []
            for row in activities:
                suggestion = code_suggestions.get(int(row["id"]))
                suggested_code = suggestion["code"] if suggestion else None
                fix = calculate_activity_rate_fix(row, rate_context, suggested_code)
                if not fix:
                    continue
                if suggested_code and suggestion and suggestion.get("reason"):
                    fix["reason"] = fix["reason"] + " " + suggestion["reason"]
                activity_changes.append({"id": row["id"], "source_file": row.get("source_file"), "date": row.get("date"), "old_code": row.get("code"), "updates": fix["updates"], "reason": fix["reason"]})

            cur.execute("""
                SELECT * FROM consumables
                WHERE contractor=%s
                ORDER BY date DESC NULLS LAST, source_file, id
                LIMIT %s
            """, (contractor, limit))
            consumables = [dict(r) for r in cur.fetchall()]
            locked_consumable_count = sum(1 for r in consumables if row_is_in_locked_sheet(r, locked_keys))
            consumables = [r for r in consumables if not row_is_in_locked_sheet(r, locked_keys)]

            product_lookup = {}
            for r in all_cr:
                key = (r.get("product") or "").strip().upper()
                if key:
                    product_lookup[key] = r
                    product_lookup[key.replace(" ", "")] = r

            fuzzy_consumables = []
            for row in consumables:
                product = (row.get("consumable") or row.get("type") or "").strip().upper()
                if product and product not in product_lookup and product.replace(" ", "") not in product_lookup:
                    fuzzy_consumables.append(row)
            consumable_matches = await gemini_suggest_consumable_matches(contractor, fuzzy_consumables, all_cr)

            consumable_changes = []
            new_consumable_rates = []
            created_products = set(product_lookup.keys())

            for row in consumables:
                raw_product = (row.get("consumable") or row.get("type") or "").strip()
                if not raw_product:
                    continue
                key = raw_product.upper()
                rate_row = product_lookup.get(key) or product_lookup.get(key.replace(" ", ""))
                match = consumable_matches.get(int(row["id"]))
                reason = "Matched imported consumable to schedule."
                if match and match.get("product"):
                    rate_row = match["product"]
                    reason = "Gemini matched consumable to schedule product. " + (match.get("reason") or "")
                elif not rate_row:
                    unit_price = row_num(row.get("unit_price"))
                    year = rate_year_for_row(row)
                    new_product = raw_product
                    product_key = new_product.upper()
                    if product_key not in created_products:
                        created_products.add(product_key)
                        new_consumable_rates.append({
                            "contractor": contractor,
                            "year": year,
                            "product": new_product,
                            "description": "Added from imported EOS consumable audit",
                            "unit_price": unit_price,
                            "unit": row.get("unit") or "each",
                        })
                    rate_row = {"product": new_product, "unit_price": unit_price, "unit": row.get("unit") or "each"}
                    reason = "No reasonable schedule match found; added imported consumable to the consumable rate list for review."

                unit_price = row_num(rate_row.get("unit_price"))
                qty = row_num(row.get("quantity")) or 1
                line_cost = round(unit_price * qty, 2)
                updates = {}
                if row.get("unit_price") is None or abs(row_num(row.get("unit_price")) - unit_price) > 0.01:
                    updates["unit_price"] = unit_price
                if row.get("line_cost") is None or abs(row_num(row.get("line_cost")) - line_cost) > 0.01:
                    updates["line_cost"] = line_cost
                if match and match.get("product"):
                    product_name = rate_row.get("product")
                    if product_name and product_name != row.get("consumable"):
                        updates["consumable"] = product_name
                        updates["type"] = product_name
                if updates:
                    consumable_changes.append({"id": row["id"], "source_file": row.get("source_file"), "date": row.get("date"), "old_consumable": raw_product, "updates": updates, "reason": reason})

            if apply_changes:
                for change in activity_changes:
                    updates = change["updates"]
                    set_clause = ", ".join(f"{k}=%s" for k in updates)
                    vals = list(updates.values()) + [change["id"]]
                    cur.execute(f"UPDATE activities SET {set_clause} WHERE id=%s", vals)
                for rate in new_consumable_rates:
                    cur.execute("""
                        INSERT INTO consumable_rates (contractor,year,product,description,unit_price,unit)
                        VALUES (%(contractor)s,%(year)s,%(product)s,%(description)s,%(unit_price)s,%(unit)s)
                        ON CONFLICT DO NOTHING
                    """, rate)
                for change in consumable_changes:
                    updates = change["updates"]
                    set_clause = ", ".join(f"{k}=%s" for k in updates)
                    vals = list(updates.values()) + [change["id"]]
                    cur.execute(f"UPDATE consumables SET {set_clause} WHERE id=%s", vals)
                conn.commit()

    return {
        "status": "updated" if apply_changes else "preview",
        "contractor": contractor,
        "activities_changed": len(activity_changes),
        "consumables_changed": len(consumable_changes),
        "consumable_rates_added": len(new_consumable_rates),
        "locked_activities_skipped": locked_activity_count,
        "locked_consumables_skipped": locked_consumable_count,
        "activity_changes": activity_changes[:100],
        "consumable_changes": consumable_changes[:100],
        "new_consumable_rates": new_consumable_rates[:100],
    }


def cleanup_coreplan_doubleups_for_contractor(contractor: str):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                DELETE FROM activities a
                WHERE a.contractor=%s
                  AND a.notes='CorePlan report total adjustment'
                  AND NOT EXISTS (
                    SELECT 1 FROM activity_sheet_locks l
                    WHERE l.contractor=a.contractor
                      AND l.report_date=COALESCE(a.date,'')
                      AND l.hole_num=COALESCE(a.hole_num,'')
                      AND l.source_file=COALESCE(a.source_file,'')
                      AND l.locked=TRUE
                  )
            """, (contractor,))
            deleted_adjustments = cur.rowcount

            cur.execute("""
                UPDATE activities a
                SET total_time='0:00', time_from='', time_to=''
                WHERE a.contractor=%s
                  AND a.code='H_Min_Shift'
                  AND COALESCE(a.notes,'') ILIKE '%%minimum shift charge%%'
                  AND COALESCE(a.total_time,'') NOT IN ('', '0', '0:00', '00:00')
                  AND NOT EXISTS (
                    SELECT 1 FROM activity_sheet_locks l
                    WHERE l.contractor=a.contractor
                      AND l.report_date=COALESCE(a.date,'')
                      AND l.hole_num=COALESCE(a.hole_num,'')
                      AND l.source_file=COALESCE(a.source_file,'')
                      AND l.locked=TRUE
                  )
            """, (contractor,))
            zeroed_minimum_shift_rows = cur.rowcount
        conn.commit()
    return {
        "deleted_adjustments": deleted_adjustments,
        "zeroed_minimum_shift_rows": zeroed_minimum_shift_rows,
    }


def minimum_shift_excluded_keys(cur, contractor: str):
    cur.execute(
        """
        SELECT contractor, report_date, hole_num, source_file
        FROM minimum_shift_topup_preferences
        WHERE contractor=%s AND include_topup=FALSE
        """,
        (contractor,),
    )
    return {
        (r.get("contractor") or "", r.get("source_file") or "", r.get("report_date") or "", r.get("hole_num") or "")
        for r in cur.fetchall()
    }


def sync_allianz_minimum_shift_topups(contractor: str):
    if not minimum_shift_rule(contractor):
        return {"deleted": 0, "inserted": 0}

    with get_conn() as conn:
        with conn.cursor() as cur:
            locked_keys = locked_activity_sheet_keys(cur, contractor)
            excluded_keys = minimum_shift_excluded_keys(cur, contractor)
            cur.execute("SELECT * FROM activities WHERE contractor=%s", (contractor,))
            rows = [dict(r) for r in cur.fetchall()]
            locked_target_totals = minimum_shift_group_totals(rows, locked_keys)
            rows, restored_coreplan_rows = update_restored_coreplan_activity_line_costs(cur, rows)

            unlocked_rows = [r for r in rows if not row_is_in_locked_sheet(r, locked_keys)]
            existing_topups = [
                r for r in unlocked_rows
                if is_generated_minimum_shift_topup(r)
            ]
            for row in existing_topups:
                cur.execute("DELETE FROM activities WHERE id=%s", (row["id"],))

            base_rows = [r for r in rows if not is_generated_minimum_shift_topup(r)]
            adjusted_rows = adjust_imported_minimum_shift_rows(base_rows, contractor, excluded_keys, locked_target_totals)
            adjusted = 0
            for row in adjusted_rows:
                if not is_imported_minimum_shift_row(row) or row.get("id") is None:
                    continue
                cur.execute(
                    """
                    UPDATE activities
                    SET quantity=%s, line_cost=%s, rate_basis=%s
                    WHERE id=%s
                    """,
                    (row.get("quantity"), row.get("line_cost"), row.get("rate_basis"), row["id"]),
                )
                adjusted += cur.rowcount

            source_rows = [
                r for r in adjusted_rows
                if not is_generated_minimum_shift_topup(r)
                and allianz_minimum_shift_group_key(r) not in excluded_keys
                and not row_is_in_locked_sheet(r, locked_keys)
            ]
            topups = build_allianz_minimum_shift_topups(source_rows)
            if topups:
                psycopg2.extras.execute_batch(cur, """
                    INSERT INTO activities
                    (source_file,contractor,date,hole_num,site_name,location,drill_rig,
                     client,contract,shift,time_from,time_to,total_time,bit_type,diameter,
                     metres_from,metres_to,total_metres,code,notes,
                     rate_year,unit_rate,quantity,line_cost,rate_basis,po_id)
                    VALUES
                    (%(source_file)s,%(contractor)s,%(date)s,%(hole_num)s,%(site_name)s,
                     %(location)s,%(drill_rig)s,%(client)s,%(contract)s,%(shift)s,
                     %(time_from)s,%(time_to)s,%(total_time)s,%(bit_type)s,%(diameter)s,
                     %(metres_from)s,%(metres_to)s,%(total_metres)s,%(code)s,%(notes)s,
                     %(rate_year)s,%(unit_rate)s,%(quantity)s,%(line_cost)s,%(rate_basis)s,%(po_id)s)
                """, topups)
        conn.commit()

    return {"deleted": len(existing_topups), "inserted": len(topups), "adjusted": adjusted, "restored_coreplan_rows": restored_coreplan_rows}


@app.on_event("startup")
def repair_minimum_shift_rows_on_startup():
    for contractor in MINIMUM_SHIFT_RULES:
        try:
            sync_allianz_minimum_shift_topups(contractor)
        except Exception as exc:
            print(f"minimum shift repair failed for {contractor}: {exc}")


@app.post("/imports/cleanup-coreplan-doubleups")
async def cleanup_coreplan_doubleups(request: Request):
    payload = await request.json()
    contractor = payload.get("contractor", "Allianz Drilling")
    result = cleanup_coreplan_doubleups_for_contractor(contractor)
    return {"status": "cleaned", "contractor": contractor, **result}


@app.post("/rates/minimum-shift-topups")
async def refresh_minimum_shift_topups(request: Request):
    payload = await request.json()
    contractor = payload.get("contractor", "Allianz Drilling")
    result = sync_allianz_minimum_shift_topups(contractor)
    return {"status": "refreshed", "contractor": contractor, "minimum_shift_topups": result}


@app.get("/minimum-shift-topups/preference")
def get_minimum_shift_topup_preference(
    contractor: str = Query(...),
    date: str = Query(""),
    hole: str = Query(""),
    source: str = Query(""),
):
    key = minimum_shift_preference_params(contractor, date, hole, source)
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT * FROM minimum_shift_topup_preferences
                WHERE contractor=%(contractor)s
                  AND report_date=%(report_date)s
                  AND hole_num=%(hole_num)s
                  AND source_file=%(source_file)s
                """,
                key,
            )
            row = cur.fetchone()
    if not row:
        return {**key, "include_topup": True, "reason": "", "updated_at": None}
    return minimum_shift_preference_response(row)


@app.post("/minimum-shift-topups/preference")
async def save_minimum_shift_topup_preference(request: Request):
    payload = await request.json()
    key = minimum_shift_preference_params(
        payload.get("contractor") or "Allianz Drilling",
        payload.get("date") or payload.get("report_date") or "",
        payload.get("hole") or payload.get("hole_num") or "",
        payload.get("source") or payload.get("source_file") or "",
    )
    include_topup = bool(payload.get("include_topup", True))
    reason = (payload.get("reason") or "").strip()
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO minimum_shift_topup_preferences
                    (contractor, report_date, hole_num, source_file, include_topup, reason, updated_at)
                VALUES
                    (%(contractor)s, %(report_date)s, %(hole_num)s, %(source_file)s, %(include_topup)s, %(reason)s, NOW())
                ON CONFLICT (contractor, report_date, hole_num, source_file)
                DO UPDATE SET include_topup=EXCLUDED.include_topup, reason=EXCLUDED.reason, updated_at=NOW()
                RETURNING *
                """,
                {**key, "include_topup": include_topup, "reason": reason},
            )
            saved = dict(cur.fetchone())
        conn.commit()
    result = sync_allianz_minimum_shift_topups(key["contractor"])
    return {
        "status": "saved",
        "preference": minimum_shift_preference_response(saved),
        "minimum_shift_topups": result,
    }


def normalize_report_approval_status(status: str):
    status = (status or "").strip().lower()
    return status if status in {"approved", "query", "rejected"} else ""


def report_approval_params(contractor: str, report_date: str = "", hole_num: str = "", source_file: str = ""):
    return {
        "contractor": contractor or "Allianz Drilling",
        "report_date": report_date or "",
        "hole_num": hole_num or "",
        "source_file": source_file or "",
    }


def report_approval_response(row):
    if not row:
        return {"status": "", "reason": "", "log": []}
    return {
        "contractor": row.get("contractor") or "",
        "report_date": row.get("report_date") or "",
        "hole_num": row.get("hole_num") or "",
        "source_file": row.get("source_file") or "",
        "status": row.get("status") or "",
        "reason": row.get("reason") or "",
        "log": row.get("log") or [],
    }


def activity_sheet_lock_params(contractor: str, report_date: str = "", hole_num: str = "", source_file: str = ""):
    return {
        "contractor": contractor or "Allianz Drilling",
        "report_date": report_date or "",
        "hole_num": hole_num or "",
        "source_file": source_file or "",
    }


def activity_sheet_lock_response(row):
    if not row:
        return {"locked": False, "reason": "", "updated_at": ""}
    return {
        "contractor": row.get("contractor") or "",
        "report_date": row.get("report_date") or "",
        "hole_num": row.get("hole_num") or "",
        "source_file": row.get("source_file") or "",
        "locked": bool(row.get("locked")),
        "reason": row.get("reason") or "",
        "updated_at": str(row.get("updated_at") or ""),
    }


def save_activity_sheet_lock_record(cur, key: dict, locked: bool, reason: str):
    cur.execute(
        """
        INSERT INTO activity_sheet_locks (contractor, report_date, hole_num, source_file, locked, reason, updated_at)
        VALUES (%(contractor)s, %(report_date)s, %(hole_num)s, %(source_file)s, %(locked)s, %(reason)s, NOW())
        ON CONFLICT (contractor, report_date, hole_num, source_file)
        DO UPDATE SET locked=EXCLUDED.locked, reason=EXCLUDED.reason, updated_at=NOW()
        RETURNING *
        """,
        {**key, "locked": locked, "reason": reason},
    )
    return dict(cur.fetchone())


def activity_row_lock_key(row: dict):
    return activity_sheet_lock_params(
        row.get("contractor") or "Allianz Drilling",
        row.get("date") or "",
        row.get("hole_num") or "",
        row.get("source_file") or "",
    )


def activity_sheet_is_locked(cur, contractor: str, report_date: str = "", hole_num: str = "", source_file: str = ""):
    key = activity_sheet_lock_params(contractor, report_date, hole_num, source_file)
    cur.execute(
        """
        SELECT locked FROM activity_sheet_locks
        WHERE contractor=%(contractor)s
          AND report_date=%(report_date)s
          AND hole_num=%(hole_num)s
          AND source_file=%(source_file)s
        """,
        key,
    )
    row = cur.fetchone()
    return bool(row and row.get("locked"))


def locked_activity_sheet_keys(cur, contractor: str):
    cur.execute(
        """
        SELECT contractor, report_date, hole_num, source_file
        FROM activity_sheet_locks
        WHERE contractor=%s AND locked=TRUE
        """,
        (contractor,),
    )
    return {
        (r.get("contractor") or "", r.get("report_date") or "", r.get("hole_num") or "", r.get("source_file") or "")
        for r in cur.fetchall()
    }


def row_is_in_locked_sheet(row: dict, locked_keys: set):
    key = (
        row.get("contractor") or "",
        row.get("date") or "",
        row.get("hole_num") or "",
        row.get("source_file") or "",
    )
    return key in locked_keys


@app.get("/activity-sheet-locks")
def get_activity_sheet_locks(
    contractor: str = Query(...),
    dates: Optional[str] = Query(None),
    date: Optional[str] = Query(None),
    hole: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
):
    conds = ["contractor=%(contractor)s"]
    params = {"contractor": contractor}
    if dates and dates.strip():
        dl = [d.strip() for d in dates.split(",") if d.strip()]
        if dl:
            conds.append("report_date=ANY(%(dates)s)")
            params["dates"] = dl
    if date is not None:
        conds.append("report_date=%(date)s")
        params["date"] = date or ""
    if hole is not None:
        conds.append("hole_num=%(hole)s")
        params["hole"] = hole or ""
    if source is not None:
        conds.append("source_file=%(source)s")
        params["source"] = source or ""
    q = f"SELECT * FROM activity_sheet_locks WHERE {' AND '.join(conds)} ORDER BY updated_at DESC"
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(q, params)
            rows = [dict(r) for r in cur.fetchall()]
    if date is not None or hole is not None or source is not None:
        return activity_sheet_lock_response(rows[0] if rows else None)
    return [activity_sheet_lock_response(r) for r in rows]


@app.post("/activity-sheet-locks")
async def save_activity_sheet_lock(request: Request):
    payload = await request.json()
    key = activity_sheet_lock_params(
        payload.get("contractor") or "Allianz Drilling",
        payload.get("date") or payload.get("report_date") or "",
        payload.get("hole") or payload.get("hole_num") or "",
        payload.get("source") or payload.get("source_file") or "",
    )
    locked = bool(payload.get("locked", True))
    reason = (payload.get("reason") or "").strip()
    with get_conn() as conn:
        with conn.cursor() as cur:
            saved = save_activity_sheet_lock_record(cur, key, locked, reason)
        conn.commit()
    return activity_sheet_lock_response(saved)


@app.get("/report-approvals")
def get_report_approvals(
    contractor: str = Query(...),
    dates: Optional[str] = Query(None),
    date: Optional[str] = Query(None),
    hole: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
):
    conds = ["contractor=%(contractor)s"]
    params = {"contractor": contractor}
    if dates and dates.strip():
        dl = [d.strip() for d in dates.split(",") if d.strip()]
        if dl:
            conds.append("report_date=ANY(%(dates)s)")
            params["dates"] = dl
    if date is not None:
        conds.append("report_date=%(date)s")
        params["date"] = date or ""
    if hole is not None:
        conds.append("hole_num=%(hole)s")
        params["hole"] = hole or ""
    if source is not None:
        conds.append("source_file=%(source)s")
        params["source"] = source or ""
    q = f"SELECT * FROM report_approvals WHERE {' AND '.join(conds)} ORDER BY updated_at DESC"
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(q, params)
            rows = [dict(r) for r in cur.fetchall()]
    if date is not None or hole is not None or source is not None:
        return report_approval_response(rows[0] if rows else None)
    return [report_approval_response(r) for r in rows]


@app.post("/report-approvals")
async def save_report_approval(request: Request):
    payload = await request.json()
    contractor = payload.get("contractor") or "Allianz Drilling"
    status = normalize_report_approval_status(payload.get("status"))
    if not status:
        raise HTTPException(400, "status must be approved, query, or rejected")
    reason = (payload.get("reason") or "").strip()
    if status in {"query", "rejected"} and not reason:
        raise HTTPException(400, "reason is required for query or rejected decisions")
    key = report_approval_params(
        contractor,
        payload.get("date") or payload.get("report_date") or "",
        payload.get("hole") or payload.get("hole_num") or "",
        payload.get("source") or payload.get("source_file") or "",
    )
    entry = {
        "status": status,
        "reason": reason,
        "at": payload.get("at") or "",
        "by": payload.get("by") or "Client",
    }
    if not entry["at"]:
        from datetime import datetime, timezone
        entry["at"] = datetime.now(timezone.utc).isoformat()
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO report_approvals (contractor, report_date, hole_num, source_file, status, reason, log, updated_at)
                VALUES (%(contractor)s, %(report_date)s, %(hole_num)s, %(source_file)s, %(status)s, %(reason)s, %(log)s::jsonb, NOW())
                ON CONFLICT (contractor, report_date, hole_num, source_file)
                DO UPDATE SET
                    status=EXCLUDED.status,
                    reason=EXCLUDED.reason,
                    log=EXCLUDED.log || report_approvals.log,
                    updated_at=NOW()
                RETURNING *
            """, {**key, "status": status, "reason": reason, "log": json.dumps([entry])})
            saved = dict(cur.fetchone())
            lock_saved = None
            if status == "approved" and bool(payload.get("lock_on_approval", False)):
                lock_reason = (payload.get("lock_reason") or "Locked on client approval to protect reviewed/custom changes.").strip()
                lock_saved = save_activity_sheet_lock_record(cur, activity_sheet_lock_params(**key), True, lock_reason)
        conn.commit()
    response = report_approval_response(saved)
    if lock_saved:
        response["lock"] = activity_sheet_lock_response(lock_saved)
    return response


@app.delete("/report-approvals")
def delete_report_approval(
    contractor: str = Query(...),
    date: str = Query(""),
    hole: str = Query(""),
    source: str = Query(""),
):
    key = report_approval_params(contractor, date, hole, source)
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                DELETE FROM report_approvals
                WHERE contractor=%(contractor)s
                  AND report_date=%(report_date)s
                  AND hole_num=%(hole_num)s
                  AND source_file=%(source_file)s
            """, key)
        conn.commit()
    return {"status": "deleted"}


@app.get("/activities")
def get_activities(
    contractor: str = Query(...),
    dates:  Optional[str] = Query(None),
    holes:  Optional[str] = Query(None),
    sites:  Optional[str] = Query(None),
    codes:  Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    missing_codes: bool = Query(False),
):
    conds = ["contractor=%(contractor)s"]
    params = {"contractor": contractor}
    if dates and dates.strip():
        dl = [d.strip() for d in dates.split(",") if d.strip()]
        if dl: conds.append("date=ANY(%(dates)s)"); params["dates"] = dl
    if holes and holes.strip():
        hl = [h.strip() for h in holes.split(",") if h.strip()]
        if hl: conds.append("hole_num=ANY(%(holes)s)"); params["holes"] = hl
    if sites and sites.strip():
        sl = [s.strip() for s in sites.split(",") if s.strip()]
        if sl: conds.append("site_name=ANY(%(sites)s)"); params["sites"] = sl
    if codes and codes.strip():
        cl = [c.strip() for c in codes.split(",") if c.strip()]
        if cl: conds.append("code=ANY(%(codes)s)"); params["codes"] = cl
    if missing_codes:
        conds.append("(code IS NULL OR TRIM(code)='')")
    if search:
        conds.append("(notes ILIKE %(search)s OR code ILIKE %(search)s)")
        params["search"] = f"%{search}%"
    q = f"SELECT * FROM activities WHERE {' AND '.join(conds)} ORDER BY date,time_from"
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(q, params)
            return [dict(r) for r in cur.fetchall()]


@app.get("/activity-report-data")
def get_activity_report_data(
    contractor: str = Query(...),
    dates: Optional[str] = Query(None),
    holes: Optional[str] = Query(None),
    sites: Optional[str] = Query(None),
    codes: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    missing_codes: bool = Query(False),
):
    """Load the complete Activity Reports view with one database round trip."""
    params = {"contractor": contractor}
    activity_conds = ["a.contractor=%(contractor)s"]
    related_conds = ["contractor=%(contractor)s"]
    approval_conds = ["contractor=%(contractor)s"]

    date_list = [d.strip() for d in (dates or "").split(",") if d.strip()]
    if date_list:
        params["dates"] = date_list
        activity_conds.append("a.date=ANY(%(dates)s)")
        related_conds.append("date=ANY(%(dates)s)")
        approval_conds.append("report_date=ANY(%(dates)s)")

    hole_list = [h.strip() for h in (holes or "").split(",") if h.strip()]
    if hole_list:
        params["holes"] = hole_list
        activity_conds.append("a.hole_num=ANY(%(holes)s)")

    site_list = [s.strip() for s in (sites or "").split(",") if s.strip()]
    if site_list:
        params["sites"] = site_list
        activity_conds.append("a.site_name=ANY(%(sites)s)")

    code_list = [c.strip() for c in (codes or "").split(",") if c.strip()]
    if code_list:
        params["codes"] = code_list
        activity_conds.append("a.code=ANY(%(codes)s)")

    if missing_codes:
        activity_conds.append("(a.code IS NULL OR TRIM(a.code)='')")
    if search:
        params["search"] = f"%{search}%"
        activity_conds.append("(a.notes ILIKE %(search)s OR a.code ILIKE %(search)s)")

    query = f"""
        SELECT
          COALESCE((
            SELECT jsonb_agg(to_jsonb(a) ORDER BY a.date, a.time_from, a.id)
            FROM activities a
            WHERE {' AND '.join(activity_conds)}
          ), '[]'::jsonb) AS activities,
          COALESCE((
            SELECT jsonb_agg(to_jsonb(c) ORDER BY c.date, c.id)
            FROM consumables c
            WHERE {' AND '.join('c.' + condition if condition.startswith('contractor') or condition.startswith('date') else condition for condition in related_conds)}
          ), '[]'::jsonb) AS consumables,
          COALESCE((
            SELECT jsonb_agg(to_jsonb(cr) ORDER BY cr.date, cr.id)
            FROM crew cr
            WHERE {' AND '.join('cr.' + condition if condition.startswith('contractor') or condition.startswith('date') else condition for condition in related_conds)}
          ), '[]'::jsonb) AS crew,
          COALESCE((
            SELECT jsonb_agg(to_jsonb(ra) ORDER BY ra.updated_at DESC)
            FROM report_approvals ra
            WHERE {' AND '.join('ra.' + condition if condition.startswith('contractor') or condition.startswith('report_date') else condition for condition in approval_conds)}
          ), '[]'::jsonb) AS report_approvals,
          COALESCE((
            SELECT jsonb_agg(to_jsonb(al) ORDER BY al.updated_at DESC)
            FROM activity_sheet_locks al
            WHERE {' AND '.join('al.' + condition if condition.startswith('contractor') or condition.startswith('report_date') else condition for condition in approval_conds)}
          ), '[]'::jsonb) AS activity_sheet_locks
    """

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(query, params)
            bundle = dict(cur.fetchone())

    bundle["report_approvals"] = [
        report_approval_response(row) for row in bundle.get("report_approvals", [])
    ]
    bundle["activity_sheet_locks"] = [
        activity_sheet_lock_response(row) for row in bundle.get("activity_sheet_locks", [])
    ]
    return bundle


@app.post("/activity-reports/delete")
@app.delete("/activity-reports")
async def delete_activity_reports(request: Request):
    payload = await request.json()
    contractor = payload.get("contractor")
    reports = payload.get("reports") or []
    if not contractor:
        raise HTTPException(400, "contractor is required")
    if not isinstance(reports, list) or not reports:
        raise HTTPException(400, "reports are required")

    totals = {
        "reports": 0,
        "activities": 0,
        "consumables": 0,
        "crew": 0,
        "approvals": 0,
        "locks": 0,
        "topup_preferences": 0,
        "imported_files": 0,
        "source_files": 0,
    }

    with get_conn() as conn:
        with conn.cursor() as cur:
            for report in reports:
                report_date = report.get("date") or report.get("report_date") or ""
                hole_num = report.get("hole") or report.get("hole_num") or ""
                source_file = report.get("source") or report.get("source_file") or ""
                key = {
                    "contractor": contractor,
                    "report_date": report_date,
                    "hole_num": hole_num,
                    "source_file": source_file,
                }
                if not (report_date or hole_num or source_file):
                    continue

                row_params = {
                    "contractor": contractor,
                    "date": report_date,
                    "hole_num": hole_num,
                    "source_file": source_file,
                }
                if source_file:
                    where = """
                        contractor=%(contractor)s
                        AND COALESCE(source_file,'')=%(source_file)s
                    """
                else:
                    where = """
                        contractor=%(contractor)s
                        AND COALESCE(date,'')=%(date)s
                        AND COALESCE(hole_num,'')=%(hole_num)s
                        AND COALESCE(source_file,'')=%(source_file)s
                    """
                for table, counter in (
                    ("activities", "activities"),
                    ("consumables", "consumables"),
                    ("crew", "crew"),
                ):
                    cur.execute(f"DELETE FROM {table} WHERE {where}", row_params)
                    totals[counter] += max(cur.rowcount, 0)

                if source_file:
                    meta_where = """
                        contractor=%(contractor)s
                        AND COALESCE(source_file,'')=%(source_file)s
                    """
                else:
                    meta_where = """
                        contractor=%(contractor)s
                        AND COALESCE(report_date,'')=%(report_date)s
                        AND COALESCE(hole_num,'')=%(hole_num)s
                        AND COALESCE(source_file,'')=%(source_file)s
                    """
                for table, counter in (
                    ("report_approvals", "approvals"),
                    ("activity_sheet_locks", "locks"),
                    ("minimum_shift_topup_preferences", "topup_preferences"),
                ):
                    cur.execute(f"DELETE FROM {table} WHERE {meta_where}", key)
                    totals[counter] += max(cur.rowcount, 0)

                if source_file:
                    cur.execute(
                        """
                        SELECT EXISTS (
                          SELECT 1 FROM activities WHERE contractor=%(contractor)s AND source_file=%(source_file)s
                          UNION ALL
                          SELECT 1 FROM consumables WHERE contractor=%(contractor)s AND source_file=%(source_file)s
                          UNION ALL
                          SELECT 1 FROM crew WHERE contractor=%(contractor)s AND source_file=%(source_file)s
                        ) AS still_used
                        """,
                        {"contractor": contractor, "source_file": source_file},
                    )
                    still_used = bool(cur.fetchone()["still_used"])
                    if not still_used:
                        cur.execute(
                            "DELETE FROM imported_files WHERE contractor=%s AND filename=%s",
                            (contractor, source_file),
                        )
                        totals["imported_files"] += max(cur.rowcount, 0)
                        cur.execute(
                            "DELETE FROM source_files WHERE contractor=%s AND filename=%s",
                            (contractor, source_file),
                        )
                        totals["source_files"] += max(cur.rowcount, 0)
                totals["reports"] += 1
        conn.commit()
    return {"status": "deleted", **totals}


@app.post("/activities")
async def create_activity(request: Request):
    payload = await request.json()
    contractor = payload.get("contractor")
    if not contractor:
        raise HTTPException(400, "contractor is required")
    safe = {"source_file","contractor","date","hole_num","site_name","program","project","location","drill_rig","client","contract","shift",
            "time_from","time_to","total_time","bit_type","diameter",
            "metres_from","metres_to","total_metres","code","notes",
            "rate_year","unit_rate","quantity","line_cost","rate_basis","po_id"}
    row = {k: v for k, v in payload.items() if k in safe}
    row.setdefault("source_file", "Manual entry")
    cols = list(row.keys())
    placeholders = ",".join(f"%({c})s" for c in cols)
    col_names = ",".join(cols)
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(f"""
                INSERT INTO activities ({col_names})
                VALUES ({placeholders})
                RETURNING *
            """, row)
            created = dict(cur.fetchone())
        conn.commit()
    return created


@app.patch("/activities/{row_id}")
async def update_activity(row_id: int, request: Request):
    payload = await request.json()
    safe = {"date","hole_num","site_name","program","project","location","drill_rig","client","contract","shift",
            "time_from","time_to","total_time","bit_type","diameter",
            "metres_from","metres_to","total_metres","code","notes",
            "rate_year","unit_rate","quantity","line_cost","rate_basis","po_id"}
    updates = {k:v for k,v in payload.items() if k in safe}
    if not updates: raise HTTPException(400,"No valid fields")
    set_clause = ",".join(f"{k}=%({k})s" for k in updates)
    updates["row_id"] = row_id
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(f"UPDATE activities SET {set_clause} WHERE id=%(row_id)s", updates)
        conn.commit()
    return {"status":"updated"}


@app.post("/activities/{row_id}/reprice")
def reprice_activity_row(row_id: int):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM activities WHERE id=%s", (row_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Activity not found")
            if activity_sheet_is_locked(
                cur,
                row.get("contractor") or "Allianz Drilling",
                row.get("date") or "",
                row.get("hole_num") or "",
                row.get("source_file") or "",
            ):
                locked_row = dict(row)
                locked_row["_locked"] = True
                return locked_row
            priced = price_activity(cur, dict(row), row["contractor"])
            updates = {
                "rate_year": priced.get("rate_year"),
                "unit_rate": priced.get("unit_rate"),
                "quantity": priced.get("quantity"),
                "line_cost": priced.get("line_cost"),
                "rate_basis": priced.get("rate_basis"),
            }
            cur.execute("""
                UPDATE activities
                SET rate_year=%(rate_year)s,
                    unit_rate=%(unit_rate)s,
                    quantity=%(quantity)s,
                    line_cost=%(line_cost)s,
                    rate_basis=%(rate_basis)s
                WHERE id=%(id)s
                RETURNING *
            """, {**updates, "id": row_id})
            updated = dict(cur.fetchone())
        conn.commit()
    return updated


@app.get("/consumables")
def get_consumables(contractor: str = Query(...), dates: Optional[str] = Query(None)):
    q = "SELECT * FROM consumables WHERE contractor=%(contractor)s"
    p = {"contractor": contractor}
    if dates and dates.strip():
        dl = [d.strip() for d in dates.split(",") if d.strip()]
        if dl:
            q += " AND date=ANY(%(dates)s)"; p["dates"] = dl
    q += " ORDER BY date"
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(q,p); return [dict(r) for r in cur.fetchall()]


@app.patch("/consumables/{row_id}")
async def update_consumable(row_id: int, request: Request):
    payload = await request.json()
    safe = {"date","hole_num","site_name","consumable","type","quantity","unit","unit_price","line_cost"}
    updates = {k:v for k,v in payload.items() if k in safe}
    if not updates: raise HTTPException(400, "No valid fields")
    # Auto-recalculate line_cost if unit_price or quantity changed
    if "unit_price" in updates or "quantity" in updates:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM consumables WHERE id=%s", (row_id,))
                row = dict(cur.fetchone()) if cur.rowcount else {}
        qty = updates.get("quantity", row.get("quantity", 1))
        up = updates.get("unit_price", row.get("unit_price", 0))
        try:
            updates["line_cost"] = round(float(up or 0) * float(qty or 1), 2)
        except (ValueError, TypeError):
            pass
    set_clause = ",".join(f"{k}=%({k})s" for k in updates)
    updates["row_id"] = row_id
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(f"UPDATE consumables SET {set_clause} WHERE id=%(row_id)s", updates)
        conn.commit()
    return {"status": "updated"}


@app.get("/crew")
def get_crew(contractor: str = Query(...), dates: Optional[str] = Query(None)):
    q = "SELECT * FROM crew WHERE contractor=%(contractor)s"
    p = {"contractor": contractor}
    if dates and dates.strip():
        dl = [d.strip() for d in dates.split(",") if d.strip()]
        if dl:
            q += " AND date=ANY(%(dates)s)"; p["dates"] = dl
    q += " ORDER BY date"
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(q,p); return [dict(r) for r in cur.fetchall()]


@app.get("/filters")
def get_filters(contractor: str = Query(...)):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT DISTINCT date FROM activities WHERE contractor=%s AND date IS NOT NULL ORDER BY date", (contractor,))
            dates = [r["date"] for r in cur.fetchall()]
            cur.execute("SELECT DISTINCT hole_num FROM activities WHERE contractor=%s AND hole_num IS NOT NULL ORDER BY hole_num", (contractor,))
            holes = [r["hole_num"] for r in cur.fetchall()]
            cur.execute("SELECT DISTINCT site_name FROM activities WHERE contractor=%s AND site_name IS NOT NULL ORDER BY site_name", (contractor,))
            sites = [r["site_name"] for r in cur.fetchall()]
            cur.execute("SELECT DISTINCT code FROM activities WHERE contractor=%s AND code!='' ORDER BY code", (contractor,))
            codes = [r["code"] for r in cur.fetchall()]
            cur.execute("SELECT COUNT(*) AS n FROM activities WHERE contractor=%s", (contractor,))
            total = cur.fetchone()["n"]
    return {"dates": dates, "holes": holes, "sites": sites, "codes": codes, "total_rows": total}


@app.get("/analytics")
def get_analytics(contractor: str = Query(...), hole: Optional[str] = Query(None)):
    try:
        q = "SELECT * FROM activities WHERE contractor=%(contractor)s"
        p = {"contractor": contractor}
        if hole and hole != "all":
            q += " AND hole_num=%(hole)s"; p["hole"] = hole

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(q, p)
                df = pd.DataFrame(cur.fetchall())

        if df.empty:
            return {"kpis":{},"daily_categories":[],"drill_runs":[],"anomalies":[],"heatmap":[]}

        def toh(t):
            try: h,m=str(t).split(":"); return int(h)+int(m)/60
            except: return 0.0

        df["hours"] = df["total_time"].apply(toh)

        def cat(c):
            if not c: return "Other"
            if any(x in c for x in ["Drill_Core","Drill_Chip"]): return "Productive Drilling"
            if "Repair" in c: return "Repairs"
            if any(x in c for x in ["Standby","Grout","Cement_Set","AAC","Logger","Sumps"]): return "Standby / Delays"
            if "Circulation" in c: return "Circulation"
            if "Travel" in c: return "Travel"
            if any(x in c for x in ["Safety","Training","Prestart"]): return "Safety & Admin"
            if "Tripping" in c: return "Tripping Rods"
            return "Other"

        df["category"] = df["code"].apply(cat)
        th = df["hours"].sum()
        dh = df[df["category"]=="Productive Drilling"]["hours"].sum()
        rh = df[df["category"]=="Repairs"]["hours"].sum()
        sh = df[df["category"]=="Standby / Delays"]["hours"].sum()
        tm = df["total_metres"].dropna().sum()
        tc = float(df["line_cost"].dropna().sum())

        daily = df.groupby(["date","category"])["hours"].sum().reset_index()

        runs = df[df["total_metres"].notna() & (df["total_metres"]>0)].copy()
        runs = runs.sort_values(["date","time_from"])
        runs["cumulative"] = runs.groupby("hole_num")["total_metres"].cumsum()

        anomalies = []
        for date, day in df.groupby("date"):
            d2 = day[day["category"]=="Productive Drilling"]["hours"].sum()
            r2 = day[day["category"]=="Repairs"]["hours"].sum()
            s2 = day[day["category"]=="Standby / Delays"]["hours"].sum()
            h2 = day["hole_num"].iloc[0] if len(day) else ""
            if d2==0: anomalies.append({"date":date,"hole":h2,"type":"No Drilling","severity":"critical","detail":"Zero productive drilling hours"})
            if r2>=2: anomalies.append({"date":date,"hole":h2,"type":"High Repairs","severity":"warning","detail":f"{r2:.1f}h repairs"})
            if s2>=3: anomalies.append({"date":date,"hole":h2,"type":"High Standby","severity":"caution","detail":f"{s2:.1f}h standby"})
        for _,r in df[df["total_metres"].notna()&(df["total_metres"]<1)&(df["total_metres"]>0)].iterrows():
            anomalies.append({"date":r["date"],"hole":r["hole_num"],"type":"Short Run","severity":"info","detail":f"{r['total_metres']}m"})
        for _,r in df[df["code"].str.contains("Circulation_Lost",na=False)].iterrows():
            anomalies.append({"date":r["date"],"hole":r["hole_num"],"type":"Lost Circulation","severity":"critical","detail":r["notes"] or "Lost returns"})

        npt = df[df["category"].isin(["Repairs","Standby / Delays","Circulation"])].groupby(["date","category"])["hours"].sum().reset_index()

        run_cols = [c for c in ["date","hole_num","time_from","total_metres","cumulative","notes"] if c in runs.columns]

        return {
            "kpis": {
                "total_hours": round(th,1), "drill_hours": round(dh,1),
                "repair_hours": round(rh,1), "delay_hours": round(sh,1),
                "total_metres": round(tm,1),
                "efficiency": round(dh/th*100,1) if th else 0,
                "total_cost": round(tc,2),
            },
            "daily_categories": daily.to_dict(orient="records"),
            "drill_runs": runs[run_cols].to_dict(orient="records"),
            "anomalies": anomalies,
            "heatmap": npt.to_dict(orient="records"),
        }
    except Exception as e:
        raise HTTPException(500, f"Analytics error: {str(e)}")


@app.get("/costing")
def get_costing(contractor: str = Query(...), holes: Optional[str]=Query(None), dates: Optional[str]=Query(None)):
    try:
        conds = ["contractor=%(contractor)s","line_cost IS NOT NULL"]
        p = {"contractor": contractor}
        if holes: conds.append("hole_num=ANY(%(holes)s)"); p["holes"]=holes.split(",")
        if dates: conds.append("date=ANY(%(dates)s)");     p["dates"]=dates.split(",")
        where = " AND ".join(conds)
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(f"""
                    SELECT
                        COALESCE(NULLIF(hole_num,''), NULLIF(project,''), NULLIF(program,''), 'Unallocated') AS hole_num,
                        COALESCE(program, '') AS program,
                        COALESCE(project, '') AS project,
                        SUM(line_cost) AS total_cost,
                        SUM(CASE WHEN code IN ('Drill_Core','Drill_Chip_or_Open_hole') THEN line_cost ELSE 0 END) AS drilling_cost,
                        SUM(CASE WHEN code NOT IN ('Drill_Core','Drill_Chip_or_Open_hole') THEN line_cost ELSE 0 END) AS non_drilling_cost,
                        SUM(total_metres) AS total_metres, COUNT(*) AS activity_count
                    FROM activities WHERE {where}
                    GROUP BY COALESCE(NULLIF(hole_num,''), NULLIF(project,''), NULLIF(program,''), 'Unallocated'), COALESCE(program, ''), COALESCE(project, '')
                    ORDER BY program, project, hole_num
                """, p)
                by_hole = [dict(r) for r in cur.fetchall()]
                cur.execute(f"""
                    SELECT
                        date,
                        COALESCE(NULLIF(hole_num,''), NULLIF(project,''), NULLIF(program,''), 'Unallocated') AS hole_num,
                        COALESCE(program, '') AS program,
                        COALESCE(project, '') AS project,
                        SUM(line_cost) AS total_cost,
                        SUM(total_metres) AS total_metres
                    FROM activities WHERE {where}
                    GROUP BY date, COALESCE(NULLIF(hole_num,''), NULLIF(project,''), NULLIF(program,''), 'Unallocated'), COALESCE(program, ''), COALESCE(project, '')
                    ORDER BY date, program, project, hole_num
                """, p)
                by_date = [dict(r) for r in cur.fetchall()]
                cur.execute(f"SELECT SUM(line_cost) AS g FROM activities WHERE {where}", p)
                grand = float(cur.fetchone()["g"] or 0)
        return {"by_hole": by_hole, "by_date": by_date, "grand_total": grand}
    except Exception as e:
        raise HTTPException(500, f"Costing error: {str(e)}")


@app.get("/rates/years")
def get_rate_years(contractor: str = Query(...)):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT DISTINCT year FROM drilling_rates WHERE contractor=%s ORDER BY year",(contractor,))
            dy = [r["year"] for r in cur.fetchall()]
            cur.execute("SELECT DISTINCT year FROM hourly_rates WHERE contractor=%s ORDER BY year",(contractor,))
            hy = [r["year"] for r in cur.fetchall()]
    return {"years": sorted(set(dy+hy))}


@app.get("/rates/drilling")
def get_drilling_rates(contractor: str = Query(...), year: Optional[str]=Query(None)):
    q = "SELECT * FROM drilling_rates WHERE contractor=%(contractor)s"
    p = {"contractor":contractor}
    if year: q += " AND year=%(year)s"; p["year"]=year
    q += " ORDER BY year,bit_type,depth_from"
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(q,p); return [dict(r) for r in cur.fetchall()]


@app.get("/rates/hourly")
def get_hourly_rates(contractor: str = Query(...), year: Optional[str]=Query(None)):
    q = "SELECT * FROM hourly_rates WHERE contractor=%(contractor)s"
    p = {"contractor":contractor}
    if year: q += " AND year=%(year)s"; p["year"]=year
    q += " ORDER BY year,code"
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(q,p); return [dict(r) for r in cur.fetchall()]


@app.post("/rates/allianz-contract-sync")
async def sync_allianz_contract_rates(request: Request):
    payload = await request.json()
    contractor = payload.get("contractor", "Allianz Drilling")
    year = str(payload.get("year") or "2025")
    if contractor not in {"Allianz Drilling", "Mitchells Drilling"}:
        raise HTTPException(400, "Contract rate sync is only available for Allianz Drilling and Mitchells Drilling.")
    if not re.match(r"^\d{4}$", year):
        raise HTTPException(400, "Year must be a four digit value.")

    drilling_rows = contract_drilling_rows(contractor, year)
    hourly_rows = contract_hourly_rows(contractor, year)
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM drilling_rates WHERE contractor=%s AND year=%s", (contractor, year))
            cur.execute("DELETE FROM hourly_rates WHERE contractor=%s AND year=%s", (contractor, year))
            psycopg2.extras.execute_batch(cur, """
                INSERT INTO drilling_rates (contractor,year,bit_type,depth_from,depth_to,rate)
                VALUES (%s,%s,%s,%s,%s,%s)
            """, drilling_rows)
            psycopg2.extras.execute_batch(cur, """
                INSERT INTO hourly_rates (contractor,year,code,description,rate,unit)
                VALUES (%s,%s,%s,%s,%s,%s)
            """, hourly_rows)
        conn.commit()
    return {
        "status": "synced",
        "contractor": contractor,
        "year": year,
        "drilling_rates": len(drilling_rows),
        "hourly_rates": len(hourly_rows),
        "source": "DrillOps contract rate template; review and edit against the signed schedule of rates.",
    }


@app.put("/rates/drilling/{rid}")
@app.patch("/rates/drilling/{rid}")
def update_drilling_rate(rid: int, payload: dict):
    safe = {"year","bit_type","depth_from","depth_to","rate"}
    u = {k:v for k,v in payload.items() if k in safe}
    if not u: raise HTTPException(400,"No valid fields")
    u["rid"]=rid
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(f"UPDATE drilling_rates SET {','.join(f'{k}=%('+k+')s' for k in u if k!='rid')} WHERE id=%(rid)s", u)
        conn.commit()
    return {"status":"updated"}


@app.put("/rates/hourly/{rid}")
@app.patch("/rates/hourly/{rid}")
def update_hourly_rate(rid: int, payload: dict):
    safe = {"year","code","description","rate","unit"}
    u = {k:v for k,v in payload.items() if k in safe}
    if not u: raise HTTPException(400,"No valid fields")
    u["rid"]=rid
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(f"UPDATE hourly_rates SET {','.join(f'{k}=%('+k+')s' for k in u if k!='rid')} WHERE id=%(rid)s", u)
        conn.commit()
    return {"status":"updated"}


@app.post("/rates/drilling")
def add_drilling_rate(payload: dict):
    req = {"contractor","year","bit_type","depth_from","depth_to","rate"}
    if not req.issubset(payload): raise HTTPException(400,f"Required: {req}")
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO drilling_rates (contractor,year,bit_type,depth_from,depth_to,rate) VALUES (%(contractor)s,%(year)s,%(bit_type)s,%(depth_from)s,%(depth_to)s,%(rate)s) RETURNING id", payload)
            nid = cur.fetchone()["id"]
        conn.commit()
    return {"status":"created","id":nid}


@app.post("/rates/hourly")
def add_hourly_rate(payload: dict):
    req = {"contractor","year","code","rate","unit"}
    if not req.issubset(payload): raise HTTPException(400,f"Required: {req}")
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO hourly_rates (contractor,year,code,description,rate,unit) VALUES (%(contractor)s,%(year)s,%(code)s,%(description)s,%(rate)s,%(unit)s) RETURNING id",
                        {**{"description":""},**payload})
            nid = cur.fetchone()["id"]
        conn.commit()
    return {"status":"created","id":nid}


@app.delete("/rates/drilling/{rid}")
def del_drilling_rate(rid: int):
    with get_conn() as conn:
        with conn.cursor() as cur: cur.execute("DELETE FROM drilling_rates WHERE id=%s",(rid,))
        conn.commit()
    return {"status":"deleted"}


@app.delete("/rates/hourly/{rid}")
def del_hourly_rate(rid: int):
    with get_conn() as conn:
        with conn.cursor() as cur: cur.execute("DELETE FROM hourly_rates WHERE id=%s",(rid,))
        conn.commit()
    return {"status":"deleted"}


# ── Consumable Rates ──────────────────────────────────────────────────────────

@app.get("/rates/consumables")
def get_consumable_rates(contractor: str = Query(...), year: Optional[str]=Query(None)):
    with get_conn() as conn:
        with conn.cursor() as cur:
            if year:
                cur.execute("SELECT * FROM consumable_rates WHERE contractor=%s AND year=%s ORDER BY product", (contractor, year))
            else:
                cur.execute("SELECT * FROM consumable_rates WHERE contractor=%s ORDER BY year,product", (contractor,))
            return [dict(r) for r in cur.fetchall()]


@app.put("/rates/consumables/{rid}")
def update_consumable_rate(rid: int, payload: dict):
    safe = {"product","description","unit_price","unit","year"}
    u = {k:v for k,v in payload.items() if k in safe}
    if not u: raise HTTPException(400,"No valid fields")
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(f"UPDATE consumable_rates SET {','.join(f'{k}=%({k})s' for k in u)} WHERE id=%(id)s", {**u,"id":rid})
        conn.commit()
    return {"status":"updated"}


@app.post("/rates/consumables")
def add_consumable_rate(payload: dict):
    req = {"contractor","year","product","unit_price"}
    if not req.issubset(payload): raise HTTPException(400, f"Required: {req}")
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""INSERT INTO consumable_rates (contractor,year,product,description,unit_price,unit)
                VALUES (%(contractor)s,%(year)s,%(product)s,%(description)s,%(unit_price)s,%(unit)s) RETURNING id""",
                {**{"description":"","unit":"each"},**payload})
            nid = cur.fetchone()["id"]
        conn.commit()
    return {"status":"created","id":nid}


@app.delete("/rates/consumables/{rid}")
def del_consumable_rate(rid: int):
    with get_conn() as conn:
        with conn.cursor() as cur: cur.execute("DELETE FROM consumable_rates WHERE id=%s",(rid,))
        conn.commit()
    return {"status":"deleted"}


@app.get("/purchase_orders")
def get_pos(contractor: Optional[str] = Query(None)):
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                if contractor:
                    cur.execute("""
                        SELECT id, po_number, contractor, description,
                               project, issue_date, expiry_date, po_value, status, notes
                        FROM purchase_orders
                        WHERE contractor=%s
                        ORDER BY issue_date DESC
                    """, (contractor,))
                else:
                    cur.execute("""
                        SELECT id, po_number, contractor, description,
                               project, issue_date, expiry_date, po_value, status, notes
                        FROM purchase_orders
                        ORDER BY issue_date DESC
                    """)
                pos = [dict(r) for r in cur.fetchall()]

                # Add spent from invoices matched to this PO
                for po in pos:
                    cur.execute("""
                        SELECT COALESCE(SUM(total_aud),0) AS spent
                        FROM invoices
                        WHERE po_reference LIKE %s AND contractor=%s
                    """, (f"%{po['po_number']}%", po["contractor"]))
                    spent = float(cur.fetchone()["spent"] or 0)
                    po["spent_to_date"] = spent
                    po["remaining"] = (po["po_value"] or 0) - spent

                return pos
    except Exception as e:
        raise HTTPException(500, f"PO error: {str(e)}")


@app.post("/purchase_orders")
async def add_po(request: Request):
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON body")
    if "po_number" not in payload: raise HTTPException(400,"po_number required")
    if "contractor" not in payload: raise HTTPException(400,"contractor required")
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO purchase_orders (po_number,contractor,project,description,issue_date,expiry_date,po_value,status,notes)
                    VALUES (%(po_number)s,%(contractor)s,%(project)s,%(description)s,%(issue_date)s,%(expiry_date)s,%(po_value)s,%(status)s,%(notes)s)
                    RETURNING id
                """, {"po_number":payload["po_number"],"contractor":payload["contractor"],
                      "project":payload.get("project",""),
                      "description":payload.get("description",""),"issue_date":payload.get("issue_date",""),
                      "expiry_date":payload.get("expiry_date",""),"po_value":payload.get("po_value",0),
                      "status":payload.get("status","Active"),"notes":payload.get("notes","")})
                nid = cur.fetchone()["id"]
            conn.commit()
        return {"status":"created","id":nid}
    except Exception as e:
        raise HTTPException(500, f"Failed to create PO: {str(e)}")


@app.patch("/purchase_orders/{po_id}")
async def update_po(po_id: int, request: Request):
    payload = await request.json()
    safe = {"po_number","project","description","issue_date","expiry_date","po_value","status","notes"}
    u = {k:v for k,v in payload.items() if k in safe}
    if not u: raise HTTPException(400,"No valid fields")
    u["po_id"]=po_id
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(f"UPDATE purchase_orders SET {','.join(f'{k}=%('+k+')s' for k in u if k!='po_id')} WHERE id=%(po_id)s", u)
        conn.commit()
    return {"status":"updated"}


@app.delete("/purchase_orders/{po_id}")
def delete_po(po_id: int):
    with get_conn() as conn:
        with conn.cursor() as cur: cur.execute("DELETE FROM purchase_orders WHERE id=%s",(po_id,))
        conn.commit()
    return {"status":"deleted"}



# ── Invoice PDF Parser ────────────────────────────────────────────────────────

# Map invoice line descriptions to EOS activity categories for matching
INV_CATEGORY_MAP = [
    # Drilling metres
    ("HQ/HQ3",          "drilling_metres"),
    ("HQ_HQ3",          "drilling_metres"),
    ("PCD",             "drilling_metres"),
    ("Hammer",          "drilling_metres"),
    # Active rate items — match both spaced and stripped
    ("tripping",        "active"),
    ("casing",          "active"),
    ("changingdrilling","active"),
    ("changing drilling","active"),
    ("flushing",        "active"),
    ("circulation",     "active"),
    ("reaming",         "active"),
    ("cementing",       "active"),
    ("mixingdrilling",  "active"),
    ("mixing drilling", "active"),
    ("measuringand",    "active"),
    ("measuring and",   "active"),
    ("T-Piece",         "active"),
    ("blooey",          "active"),
    ("repairs",         "active"),
    ("unplanned",       "active"),
    # Standby / inactive
    ("standby",         "standby"),
    ("waiting",         "standby"),
    # Travel / setup
    ("travel",          "travel"),
    ("movingbetween",   "travel"),
    ("moving between",  "travel"),
    ("settingup",       "setup"),
    ("packingup",       "setup"),
    ("setting up",      "setup"),
    ("packing up",      "setup"),
    # Safety / admin
    ("safety",          "safety"),
    ("pre-start",       "safety"),
    ("prestart",        "safety"),
    ("induction",       "safety"),
    ("training",        "safety"),
    ("toolbox",         "safety"),
    ("authorisation",   "safety"),
    ("consumables from local", "active"),
    ("consumablesfromlocal",   "active"),
    # Equipment day rates
    ("watercart",       "equipment"),
    ("water cart",      "equipment"),
    ("backhoe",         "equipment"),
    ("excavator",       "equipment"),
    ("grader",          "equipment"),
    ("vac truck",       "equipment"),
    ("water truck",     "equipment"),
    ("light vehicle",   "equipment"),
    ("lv - light",      "equipment"),
    ("trade labour",    "labour"),
    ("labour",          "labour"),
    # Consumables
    ("AMC",             "consumable"),
    ("cement",          "consumable"),
    ("PVC",             "consumable"),
    ("MUDLOGIC",        "consumable"),
    ("foam",            "consumable"),
    ("slurry",          "consumable"),
    ("GRIP",            "consumable"),
    ("SUPERLUBE",       "consumable"),
    ("SUPERFOAM",       "consumable"),
    ("HARD SET",        "consumable"),
    ("HARDSET",         "consumable"),
    ("TORQ",            "consumable"),
    ("SWELL",           "consumable"),
    ("CR650",           "consumable"),
    # Mobilisation
    ("mobilisation",    "mobilisation"),
    ("mobilization",    "mobilisation"),
    ("demobilisation",  "mobilisation"),
    ("compliance",      "compliance"),
]

def categorise_invoice_line(description: str) -> str:
    dl = description.lower()
    for keyword, cat in INV_CATEGORY_MAP:
        if keyword.lower() in dl:
            return cat
    return "other"


def parse_invoice_breakdown_date(value: str) -> str:
    months = {
        "jan": "01", "feb": "02", "mar": "03", "apr": "04",
        "may": "05", "jun": "06", "jul": "07", "aug": "08",
        "sep": "09", "oct": "10", "nov": "11", "dec": "12",
    }
    m = re.match(r"(\d{1,2})-([A-Za-z]{3})-(\d{4})", str(value or "").strip())
    if not m:
        return value or ""
    return f"{int(m.group(1)):02d}/{months.get(m.group(2).lower(), '01')}/{m.group(3)}"


def parse_invoice_breakdown_money(value: str) -> float:
    try:
        return float(str(value or "0").replace("$", "").replace(",", "").strip())
    except Exception:
        return 0.0


def parse_invoice_breakdown_hours(value: str) -> float:
    s = str(value or "").strip()
    if ":" in s:
        h, m = s.split(":", 1)
        try:
            return int(h) + int(m) / 60.0
        except Exception:
            return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def invoice_breakdown_activity_match(prefix: str):
    pattern = re.compile(
        r"\b(Drill_Core|Drill_Chip_or_Open_hole|Crew_Travel|MOB|DEMOB|D_[A-Za-z0-9_]+|H_[A-Za-z0-9_]+|HQ/HQ3\s+.*?\)|PCD\s+.*?\))\b"
    )
    match = pattern.search(prefix)
    if not match:
        return None
    return match


def invoice_breakdown_location_fields(left: str):
    left = re.sub(r"\s+", " ", left or "").strip()
    site = ""
    hole = ""
    location = left
    site_matches = list(re.finditer(r"CD-\d{2}-\d{3}(?:\s*/\s*\d+)?", left))
    if site_matches:
        site_match = site_matches[-1]
        site = site_match.group(0).strip()
        before = left[:site_match.start()].strip()
        after = left[site_match.end():].strip()
        ddr = re.search(r"\b([A-Z]{1,4}\d{3,5}[A-Z]?)\b", after)
        if ddr:
            hole = ddr.group(1)
        elif before:
            parts = before.split()
            if len(parts) > 2:
                location = " ".join(parts[:2]) if parts[0].upper() == "CDM" else parts[0]
                hole = " ".join(parts[2:]) if parts[0].upper() == "CDM" else " ".join(parts[1:])
            else:
                location = before
        if not hole:
            hole = site
    return location, site, hole


def invoice_breakdown_match_key(line: dict) -> str:
    code = line.get("activity_code") or ""
    source_cat = line.get("source_category") or line.get("category") or ""
    if source_cat.lower() == "drilling" and not code.startswith("Drill_"):
        return "DRILLING_METRES"
    return code or categorise_invoice_line(line.get("description") or "")


def parse_monthly_breakdown_invoice_pdf(text: str, filename: str) -> dict:
    rows = []
    row_re = re.compile(r"(?=(\d{1,2}-[A-Za-z]{3}-\d{4})\s+\d+\s+)")
    starts = [m.start() for m in row_re.finditer(text)]
    starts.append(len(text))
    for i in range(len(starts) - 1):
        raw = " ".join(text[starts[i]:starts[i+1]].split())
        suffix = re.search(r"\s+([\d:]+(?:\.\d+)?)\s+(hrs|m)\s+\$([\d,]+\.\d{2})\s+(Yes|No)\s+(.+)$", raw)
        if not suffix:
            continue
        head = raw[:suffix.start()].strip()
        quantity_raw, unit, amount_raw, chargeable, source_category = suffix.groups()
        dm = re.match(r"(\d{1,2}-[A-Za-z]{3}-\d{4})\s+(\d+)\s+(.+)$", head)
        if not dm:
            continue
        line_date_raw, day, rest = dm.groups()
        activity_match = invoice_breakdown_activity_match(rest)
        if not activity_match:
            continue
        left = rest[:activity_match.start()].strip()
        activity_code = activity_match.group(1).strip()
        description = rest[activity_match.end():].strip()
        location, site, hole = invoice_breakdown_location_fields(left)
        amount = parse_invoice_breakdown_money(amount_raw)
        quantity = parse_invoice_breakdown_hours(quantity_raw) if unit == "hrs" else coreplan_float(quantity_raw) or 0
        unit_price = round(amount / quantity, 2) if quantity else amount
        rows.append({
            "line_date": parse_invoice_breakdown_date(line_date_raw),
            "site_name": site,
            "hole_num": hole,
            "activity_code": activity_code,
            "description": f"{activity_code} - {description}",
            "quantity": quantity,
            "unit": unit,
            "unit_price": unit_price,
            "gst_rate": "10%",
            "amount": amount,
            "chargeable": chargeable,
            "source_category": source_category,
            "category": categorise_invoice_line(activity_code + " " + description + " " + source_category),
        })
    subtotal = round(sum(r["amount"] for r in rows), 2)
    return {
        "invoice_number": os.path.splitext(filename)[0],
        "invoice_date": "31/05/2026" if "May" in filename else "",
        "due_date": "",
        "po_reference": "",
        "client": "Fitzroy Coal",
        "abn": "",
        "subtotal": subtotal,
        "gst": 0.0,
        "total_aud": subtotal,
        "amount_paid": 0.0,
        "amount_due": subtotal,
        "status": "Unpaid",
        "lines": rows,
        "breakdown_type": "monthly_task",
    }


def parse_daily_breakdown_invoice_pdf(text: str, filename: str) -> dict:
    rows = []
    line_re = re.compile(
        r"^(\d{1,2}-[A-Za-z]{3}-\d{4})\s+(\d+)\s+(.+?)\s+(\d{1,2}:\d{2})\s+\$([\d,]+\.\d{2})\s+([\d.]+)\s+\$([\d,]+\.\d{2})\s+\$([\d,]+\.\d{2})$",
        re.MULTILINE
    )
    for m in line_re.finditer(text):
        line_date_raw, day, left, total_hours, hourly_cost, drilling_qty, drilling_cost, daily_total = m.groups()
        location, site, hole = invoice_breakdown_location_fields(left)
        amount = parse_invoice_breakdown_money(daily_total)
        rows.append({
            "line_date": parse_invoice_breakdown_date(line_date_raw),
            "site_name": site,
            "hole_num": hole,
            "activity_code": "DAILY_TOTAL",
            "description": f"Daily total - hourly {hourly_cost}, drilling {drilling_qty}m / {drilling_cost}",
            "quantity": 1,
            "unit": "day",
            "unit_price": amount,
            "gst_rate": "10%",
            "amount": amount,
            "chargeable": "Yes",
            "source_category": "Daily Summary",
            "category": "daily_summary",
        })
    subtotal = round(sum(r["amount"] for r in rows), 2)
    return {
        "invoice_number": os.path.splitext(filename)[0],
        "invoice_date": "31/05/2026" if "May" in filename else "",
        "due_date": "",
        "po_reference": "",
        "client": "Fitzroy Coal",
        "abn": "",
        "subtotal": subtotal,
        "gst": 0.0,
        "total_aud": subtotal,
        "amount_paid": 0.0,
        "amount_due": subtotal,
        "status": "Unpaid",
        "lines": rows,
        "breakdown_type": "daily_summary",
    }


def parse_king_konstruct_invoice_pdf(text: str, filename: str) -> dict:
    """Parse King Konstruct tax invoices."""

    def find(pattern, default=""):
        m = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        return m.group(1).strip() if m else default

    def amount(pattern):
        raw = find(pattern)
        if not raw:
            return 0.0
        try:
            return float(raw.replace(",", ""))
        except ValueError:
            return 0.0

    invoice_number = find(r"INVOICE\s+(\d{5})") or find(r"Invoice_(\d{5})", filename)
    invoice_date = find(r"\bDATE\s+(\d{1,2}/\d{1,2}/\d{4})")
    due_date = find(r"DUE DATE\s+(\d{1,2}/\d{1,2}/\d{4})")
    po_reference = find(r"PURCHASE ORDER\s*\n?\s*(C\d+)")
    client = find(r"(Fitzroy Coal Management Pty Ltd)")
    abn = find(r"ABN\s+([\d\s]{10,})")

    subtotal = amount(r"^SUBTOTAL\s+([\d,]+\.\d{2})")
    gst = amount(r"^GST TOTAL\s+([\d,]+\.\d{2})")
    total_aud = amount(r"^TOTAL\s+([\d,]+\.\d{2})")
    amount_due = amount(r"A\$([\d,]+\.\d{2})\s+BALANCE DUE")
    if not amount_due:
        amount_due = total_aud

    lines = []
    source_lines = [line.strip() for line in text.splitlines() if line.strip()]
    try:
        start = next(i for i, line in enumerate(source_lines) if "ACTIVITY QTY RATE AMOUNT" in line) + 1
    except StopIteration:
        start = len(source_lines)
    end = next((i for i, line in enumerate(source_lines[start:], start) if line.startswith("SUBTOTAL")), len(source_lines))

    line_re = re.compile(r"^(.+?)\s+(\d+(?:\.\d+)?)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})$")
    for line in source_lines[start:end]:
        m = line_re.match(line)
        if not m:
            continue
        description = m.group(1).strip()
        try:
            quantity = float(m.group(2).replace(",", ""))
            unit_price = float(m.group(3).replace(",", ""))
            line_amount = float(m.group(4).replace(",", ""))
        except ValueError:
            continue
        lines.append({
            "description": description,
            "quantity": quantity,
            "unit_price": unit_price,
            "gst_rate": "10%",
            "amount": line_amount,
            "category": categorise_invoice_line(description),
        })

    if subtotal == 0 and lines:
        subtotal = round(sum(l["amount"] for l in lines), 2)
    if gst == 0 and subtotal:
        gst = round(subtotal * 0.1, 2)
    if total_aud == 0 and subtotal:
        total_aud = round(subtotal + gst, 2)
    if amount_due == 0:
        amount_due = total_aud

    return {
        "invoice_number": invoice_number,
        "invoice_date": invoice_date,
        "due_date": due_date,
        "po_reference": po_reference,
        "client": client,
        "abn": abn,
        "subtotal": subtotal,
        "gst": gst,
        "total_aud": total_aud,
        "amount_paid": 0.0,
        "amount_due": amount_due,
        "status": "Unpaid" if amount_due else "Paid",
        "lines": lines,
    }


def invoice_project_from_text(text: str, filename: str = "") -> str:
    haystack = " ".join([text or "", filename or ""]).upper()
    if "ARG-EXP" in haystack or "ARGEXP" in haystack:
        return "Exploration"
    if "ARG-002" in haystack or "GAS RISER" in haystack:
        return "Gas Riser"
    if "ARG-003" in haystack or "SIS" in haystack:
        return "SIS"
    if "ARG-005" in haystack:
        return "Exploration"
    return ""


def parse_mcc_group_invoice_pdf(text: str, filename: str) -> dict:
    def find(pattern, default=""):
        m = re.search(pattern, text, re.IGNORECASE)
        return m.group(1).strip() if m else default

    def amount(pattern):
        raw = find(pattern)
        try:
            return float(raw.replace(",", "")) if raw else 0.0
        except Exception:
            return 0.0

    invoice_number = find(r"\bINVOICE\s+(ARG[-A-Z0-9]+)") or find(r"(ARG[-A-Z]+[-]?\d+)", os.path.splitext(filename)[0])
    invoice_date = find(r"\bDATE\s+(\d{1,2}/\d{1,2}/\d{4})")
    due_date = find(r"\bDUE DATE\s+(\d{1,2}/\d{1,2}/\d{4})")
    po_reference = find(r"\bPO\s*\n\s*([A-Z]?\d{4,})")
    client = find(r"INVOICE TO\s+(.+?)\s+DATE") or "Argo Coal Management Pty Ltd"
    abn = find(r"\bABN[:\s]+(\d{11})")
    subtotal = amount(r"\bSUBTOTAL\s+([\d,]+\.\d{2})")
    gst = amount(r"\bGST TOTAL\s+([\d,]+\.\d{2})")
    raw_lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    total_aud = 0.0
    for raw_line in raw_lines:
        m_total = re.match(r"^TOTAL\s+([\d,]+\.\d{2})$", raw_line, re.I)
        if m_total:
            total_aud = float(m_total.group(1).replace(",", ""))
            break
    if not total_aud:
        total_aud = amount(r"A\$([\d,]+\.\d{2})")
    lines = []
    line_re = re.compile(r"^(Labour Services|Hire - [A-Za-z ]+)\s+([\d,]+(?:\.\d+)?)\s+([\d,]+(?:\.\d{2})?)\s+GST\s+([\d,]+(?:\.\d{2})?)$", re.I)
    for idx, line in enumerate(raw_lines):
        m = line_re.match(line)
        if not m:
            continue
        item = m.group(1).strip()
        detail = raw_lines[idx + 1].strip() if idx + 1 < len(raw_lines) else ""
        desc = f"{item} - {detail}" if detail and not re.search(r"\b(SUBTOTAL|GST|TOTAL|DATE|PO)\b", detail, re.I) else item
        qty = float(m.group(2).replace(",", ""))
        rate = float(m.group(3).replace(",", ""))
        amt = float(m.group(4).replace(",", ""))
        lines.append({
            "description": desc,
            "quantity": qty,
            "unit_price": rate,
            "gst_rate": "10%",
            "amount": amt,
            "category": categorise_invoice_line(desc),
        })

    if subtotal == 0 and lines:
        subtotal = round(sum(l["amount"] for l in lines), 2)
    if gst == 0 and subtotal:
        gst = round(subtotal * 0.1, 2)
    if total_aud == 0 and subtotal:
        total_aud = round(subtotal + gst, 2)

    return {
        "invoice_number": invoice_number or os.path.splitext(filename)[0],
        "invoice_date": invoice_date,
        "due_date": due_date,
        "po_reference": po_reference,
        "project": invoice_project_from_text(text, filename),
        "client": client,
        "abn": abn,
        "subtotal": subtotal,
        "gst": gst,
        "total_aud": total_aud,
        "amount_paid": 0,
        "amount_due": total_aud,
        "status": "Unpaid",
        "lines": lines,
    }


def parse_chms_quote_invoice_pdf(text: str, filename: str) -> dict:
    def find(pattern, default=""):
        m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        return m.group(1).strip() if m else default

    def money(pattern):
        raw = find(pattern)
        try:
            return float(raw.replace(",", "")) if raw else 0.0
        except Exception:
            return 0.0

    header = re.search(r"Quote number\s+Issue date\s+Expiry date\s+(\d+)\s+(\d{1,2}/\d{1,2}/\d{4})\s+(\d{1,2}/\d{1,2}/\d{4})", text, re.I)
    quote_number = header.group(1) if header else find(r"Quote no:\s*(\d+)")
    issue_date = header.group(2) if header else ""
    expiry_date = header.group(3) if header else ""
    invoice_number = f"Quote-{quote_number}" if quote_number else os.path.splitext(filename)[0]
    client = find(r"Bill to\s+Ship to\s+(.+?)\s+Peak Downs", "Argo Coal Management Pty Ltd")
    abn = find(r"\bABN:\s*([\d\s]+)")
    subtotal = money(r"Subtotal\s*\(exc\. tax\)\s*\$([\d,]+\.\d{2})")
    gst = money(r"\bTax\s*\$([\d,]+\.\d{2})")
    total_aud = money(r"Total amount\s*\$([\d,]+\.\d{2})")
    if not total_aud:
        total_aud = money(r"Total amount:\s*\$([\d,]+\.\d{2})")

    raw_lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    line_re = re.compile(
        r"^(?P<item>[A-Z]+(?:\s*-\s*[A-Z]+)?)\s+(?P<desc>.+?)\s+Qty\s+"
        r"(?P<qty>[\d,]+(?:\.\d+)?)\s+(?P<unit>[\d,]+\.\d{2})\s+GST\s+(?P<amount>[\d,]+\.\d{2})$",
        re.I,
    )
    lines = []
    for idx, raw in enumerate(raw_lines):
        m = line_re.match(raw)
        if not m:
            continue
        item = m.group("item").strip()
        desc = m.group("desc").strip()
        if idx + 1 < len(raw_lines):
            continuation = raw_lines[idx + 1].strip()
            if continuation and not re.search(r"\b(Subtotal|Tax|Total|Quote|Bill to|Item ID)\b", continuation, re.I) and " Qty " not in continuation:
                desc = f"{desc} {continuation}"
        qty = float(m.group("qty").replace(",", ""))
        unit_price = float(m.group("unit").replace(",", ""))
        amount = float(m.group("amount").replace(",", ""))
        full_desc = f"{item} - {desc}"
        category = "labour" if item.upper().startswith("LAB") else ("equipment" if item.upper().startswith("EQUIP") else categorise_invoice_line(full_desc))
        lines.append({
            "description": full_desc,
            "quantity": qty,
            "unit_price": unit_price,
            "gst_rate": "10%",
            "amount": amount,
            "category": category,
            "unit": "Qty",
            "source_category": item,
        })

    if subtotal == 0 and lines:
        subtotal = round(sum(l["amount"] for l in lines), 2)
    if gst == 0 and subtotal:
        gst = round(subtotal * 0.1, 2)
    if total_aud == 0 and subtotal:
        total_aud = round(subtotal + gst, 2)

    return {
        "invoice_number": invoice_number,
        "invoice_date": issue_date,
        "due_date": expiry_date,
        "po_reference": "",
        "project": invoice_project_from_text(text, filename),
        "client": client,
        "abn": abn,
        "subtotal": subtotal,
        "gst": gst,
        "total_aud": total_aud,
        "amount_paid": 0,
        "amount_due": total_aud,
        "status": "Unpaid",
        "lines": lines,
    }


def parse_invoice_pdf(text: str, filename: str, contractor: str) -> dict:
    """Parse an Allianz-style tax invoice PDF.
    Note: pdfplumber strips spaces from words so 'Invoice Number' becomes 'InvoiceNumber'.
    """

    if "Daily Total" in text and "Hourly Cost" in text and "Drilling Qty" in text:
        return parse_daily_breakdown_invoice_pdf(text, filename)

    if "Activity Code" in text and "Chargeable Category" in text:
        return parse_monthly_breakdown_invoice_pdf(text, filename)

    if "KING KONSTRUCT" in text.upper() or re.search(r"Invoice_\d{5}_from_KING_KONSTRUCT", filename, re.IGNORECASE):
        return parse_king_konstruct_invoice_pdf(text, filename)

    if "MCC Group Pty Ltd" in text or re.search(r"from_MCC_Group", filename, re.IGNORECASE):
        return parse_mcc_group_invoice_pdf(text, filename)

    if "Central Highlands Mining Services" in text or re.search(r"Quote-00001532", filename, re.IGNORECASE):
        return parse_chms_quote_invoice_pdf(text, filename)

    def find(pattern, default=""):
        m = re.search(pattern, text, re.IGNORECASE)
        return m.group(1).strip() if m else default

    # Header — pdfplumber strips spaces, layout is non-standard
    # "InvoiceNumber NOOSAVILLEQLD4566\nFitzroyAustraliaResourcesPtyLtd INV-0509"
    invoice_number = find(r"(INV-\d+)")  # just find it anywhere
    invoice_date   = find(r"Invoice\s*Date\s*\n?\s*(\d+\s*\w+\s*\d{4})")
    if not invoice_date:
        # format: "InvoiceDate\n3Apr2026" or "24Feb2026"
        invoice_date = find(r"InvoiceDate\s*\n?\s*(\d+\w+\d{4})")
    due_date       = find(r"Due\s*Date[:\s]+(\d+\s+\w+\s+\d{4})")
    po_reference   = (find(r"Reference\s*\n?\s*(Purchase\s*Order\s*\S+)") or
                      find(r"Reference\s*\n?\s*(PO\s+\S+)") or
                      find(r"PurchaseOrder(\S+)") or
                      find(r"(C\d{6,}|F\d{6,})"))
    client         = find(r"(Fitzroy[\w\s]+?(?:Pty\s*Ltd|Resources\s*Pty\s*Ltd))")
    abn_raw        = find(r"ABN\s*\n?\s*([\d\s]{10,})")
    abn            = abn_raw.strip() if abn_raw else ""

    # Totals — handle both "Subtotal" and no-space versions
    def find_amount(pattern):
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            try: return float(m.group(1).replace(",",""))
            except: return 0.0
        return 0.0

    subtotal    = find_amount(r"Subtotal\s+([\d,]+\.?\d*)")
    gst         = find_amount(r"TOTAL\s*GST\s*10%\s+([\d,]+\.?\d*)")
    total_aud   = find_amount(r"TOTAL\s*AUD\s+([\d,]+\.?\d*)")
    amount_paid = find_amount(r"Less\s*Amount\s*Paid\s+([\d,]+\.?\d*)")
    amount_due  = find_amount(r"AMOUNT\s*DUE\s*AUD\s+([\d,]+\.?\d*)")

    # Determine status
    if amount_paid > 0 and amount_due == 0:
        status = "Paid"
    elif amount_due > 0 and amount_paid > 0:
        status = "Partial"
    elif amount_due > 0:
        status = "Unpaid"
    else:
        status = "Paid"

    # ── Line items ─────────────────────────────────────────────────────────────
    # Each line: description qty unit_price 10% amount
    # In stripped text spaces between words are removed so we match on
    # the numeric columns which are preserved: qty unit_price 10% amount
    # Pattern: any text ending with qty price 10% amount on the same line
    line_re = re.compile(
        r"^(.+?)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+10%\s+([\d,]+\.\d{2})\s*$",
        re.MULTILINE
    )

    lines = []
    skip_words = {"description", "subtotal", "quantity", "unitprice", "gst", "amountaud"}

    for m in line_re.finditer(text):
        desc  = m.group(1).strip()
        desc_lower = desc.lower().replace(" ","")

        if any(s in desc_lower for s in skip_words):
            continue

        try:
            qty   = float(m.group(2).replace(",",""))
            price = float(m.group(3).replace(",",""))
            amt   = float(m.group(4).replace(",",""))
        except ValueError:
            continue

        # Sanity check with 2% tolerance
        if qty > 0 and price > 0:
            expected = qty * price
            if expected > 0 and abs(expected - amt) / expected > 0.02:
                continue

        lines.append({
            "description": desc,
            "quantity":    qty,
            "unit_price":  price,
            "gst_rate":    "10%",
            "amount":      amt,
            "category":    categorise_invoice_line(desc),
        })

    # If subtotal wasn't found, calculate from lines
    if subtotal == 0 and lines:
        subtotal = round(sum(l["amount"] for l in lines), 2)
    if gst == 0 and subtotal:
        gst = round(subtotal * 0.1, 2)
    if total_aud == 0 and subtotal:
        total_aud = round(subtotal * 1.1, 2)

    return {
        "invoice_number": invoice_number,
        "invoice_date":   invoice_date,
        "due_date":       due_date,
        "po_reference":   po_reference,
        "client":         client,
        "abn":            abn,
        "subtotal":       subtotal,
        "gst":            gst,
        "total_aud":      total_aud,
        "amount_paid":    amount_paid,
        "amount_due":     amount_due,
        "status":         status,
        "lines":          lines,
    }


def match_invoice_to_eos(cur, invoice_id: int, contractor: str, po_reference: str):
    """
    Compare invoice line totals by category against EOS activity costs.
    Updates match_status and variance on each invoice line.
    """
    cur.execute("SELECT * FROM invoice_lines WHERE invoice_id=%s", (invoice_id,))
    inv_lines = [dict(r) for r in cur.fetchall()]
    if not inv_lines:
        return
    cur.execute("SELECT project FROM invoices WHERE id=%s", (invoice_id,))
    invoice_project = ((cur.fetchone() or {}).get("project") or "").strip()

    def add_invoice_project_filter(where, params):
        if not invoice_project:
            return
        if invoice_project in {"Exploration", "Gas Riser", "SIS"}:
            where.append("(program=%s OR project=%s)")
            params.extend([invoice_project, invoice_project])
        else:
            where.append("project=%s")
            params.append(invoice_project)

    if any(line.get("line_date") or line.get("activity_code") for line in inv_lines):
        groups = {}
        for line in inv_lines:
            key = (
                line.get("line_date") or "",
                line.get("site_name") or "",
                line.get("hole_num") or "",
                invoice_breakdown_match_key(line),
            )
            groups.setdefault(key, []).append(line)

        eos_by_key = {}
        for key in groups:
            line_date, site_name, hole_num, match_key = key
            params = [contractor]
            where = ["contractor=%s", "line_cost IS NOT NULL"]
            add_invoice_project_filter(where, params)
            if line_date:
                iso = line_date
                m = re.match(r"(\d{2})/(\d{2})/(\d{4})", line_date)
                if m:
                    iso = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
                where.append("(date=%s OR date=%s)")
                params.extend([line_date, iso])
            if hole_num and hole_num != site_name:
                where.append("(hole_num=%s OR drill_rig=%s)")
                params.extend([hole_num, hole_num])
            elif site_name:
                where.append("site_name=%s")
                params.append(site_name)
            elif hole_num:
                where.append("hole_num=%s")
                params.append(hole_num)
            if match_key == "DRILLING_METRES":
                where.append("COALESCE(total_metres,0)>0")
            elif match_key == "DAILY_TOTAL":
                pass
            elif match_key:
                where.append("code=%s")
                params.append(match_key)
            cur.execute(f"""
                SELECT COALESCE(SUM(line_cost),0) AS eos_total
                FROM activities
                WHERE {' AND '.join(where)}
            """, tuple(params))
            eos_by_key[key] = float((cur.fetchone() or {}).get("eos_total") or 0)

        for key, lines in groups.items():
            eos_total = eos_by_key.get(key, 0)
            inv_total = sum(float(line.get("amount") or 0) for line in lines)
            for line in lines:
                amount = float(line.get("amount") or 0)
                matched_eos = round(eos_total * (amount / inv_total), 2) if inv_total else (round(eos_total / len(lines), 2) if lines else 0)
                variance = round(amount - matched_eos, 2)
                if eos_total == 0 and amount == 0:
                    match_status = "exact_match"
                elif eos_total == 0:
                    match_status = "no_eos_data"
                elif abs(variance) <= 1:
                    match_status = "exact_match"
                elif amount and abs(variance) / abs(amount) <= 0.05:
                    match_status = "close_match"
                elif variance > 0:
                    match_status = "invoice_over_eos"
                else:
                    match_status = "invoice_under_eos"
                cur.execute("""
                    UPDATE invoice_lines
                    SET matched_eos_cost=%s, variance=%s, match_status=%s
                    WHERE id=%s
                """, (matched_eos, variance, match_status, line["id"]))
        return

    cur.execute("""
        SELECT
            CASE
                WHEN code IN ('Drill_Core','Drill_Chip_or_Open_hole') THEN 'drilling_metres'
                WHEN code LIKE '%Standby%' OR code LIKE '%standby%' THEN 'standby'
                WHEN code LIKE '%Travel%' OR code LIKE '%travel%' THEN 'travel'
                WHEN code LIKE '%Safety%' OR code LIKE '%Repair%' OR code LIKE '%Training%'
                  OR code LIKE '%Prestart%' THEN 'safety'
                WHEN code LIKE 'MCC_%' AND notes ILIKE '%Charge type: Labour%' THEN 'labour'
                WHEN code LIKE 'MCC_%' AND notes ILIKE '%Charge type: Equipment%' THEN 'equipment'
                WHEN code LIKE 'D_Backhoe%' OR code LIKE 'D_Water%' THEN 'equipment'
                WHEN code LIKE '%Setup%' OR code LIKE '%Surface%' THEN 'setup'
                ELSE 'active'
            END AS category,
            SUM(line_cost) AS eos_total
        FROM activities
        WHERE contractor=%s AND line_cost IS NOT NULL
          AND (%s='' OR program=%s OR project=%s)
        GROUP BY 1
    """, (contractor, invoice_project, invoice_project, invoice_project))
    eos_by_cat = {r["category"]: float(r["eos_total"] or 0) for r in cur.fetchall()}

    inv_by_cat = {}
    for line in inv_lines:
        cat = line["category"]
        inv_by_cat.setdefault(cat, 0)
        inv_by_cat[cat] += float(line["amount"] or 0)

    for line in inv_lines:
        cat = line["category"]
        eos_total = eos_by_cat.get(cat, 0)
        inv_total = inv_by_cat.get(cat, 0)
        # Pro-rate EOS cost to this line's share
        line_share = (float(line["amount"] or 0) / inv_total) if inv_total else 0
        matched_eos = round(eos_total * line_share, 2) if eos_total else None
        variance = round(float(line["amount"] or 0) - matched_eos, 2) if matched_eos is not None else None

        if matched_eos is None:
            match_status = "no_eos_data"
        elif abs(variance) < 0.01:
            match_status = "exact_match"
        elif abs(variance) / float(line["amount"]) < 0.05 if float(line["amount"]) else False:
            match_status = "close_match"
        elif variance > 0:
            match_status = "invoice_over_eos"
        else:
            match_status = "invoice_under_eos"

        cur.execute("""
            UPDATE invoice_lines
            SET matched_eos_cost=%s, variance=%s, match_status=%s
            WHERE id=%s
        """, (matched_eos, variance, match_status, line["id"]))


# ── Invoice API endpoints ─────────────────────────────────────────────────────

@app.post("/invoices/test")
async def test_invoice_parse(
    file: UploadFile = File(...),
    contractor: str = Form(default="Allianz Drilling"),
):
    """Test endpoint — parses invoice and returns result without saving."""
    content = await file.read()
    try:
        with pdfplumber.open(BytesIO(content)) as pdf:
            text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    except Exception as e:
        return {"error": f"PDF read failed: {e}"}
    try:
        inv = parse_invoice_pdf(text, file.filename, contractor)
        return {"parsed": inv, "line_count": len(inv.get("lines",[]))}
    except Exception as e:
        return {"error": f"Parse failed: {e}", "text_preview": text[:500]}


@app.post("/invoices/import")
async def import_invoice(
    file: UploadFile = File(...),
    contractor: str = Form(default="Allianz Drilling"),
):
    filename = file.filename

    # Check if already imported — handle both old single-col and new dual-col table
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM invoice_imports WHERE filename=%s AND contractor=%s",
                            (filename, contractor))
                if cur.fetchone():
                    return {"status": "skipped", "filename": filename}
    except Exception:
        # Table may have old schema — try migration
        with get_conn() as conn:
            with conn.cursor() as cur:
                try:
                    cur.execute("ALTER TABLE invoice_imports ADD COLUMN IF NOT EXISTS contractor TEXT DEFAULT 'Allianz Drilling'")
                    cur.execute("SELECT 1 FROM invoice_imports WHERE filename=%s", (filename,))
                    if cur.fetchone():
                        return {"status": "skipped", "filename": filename}
                except Exception:
                    pass
            conn.commit()

    content = await file.read()
    try:
        with pdfplumber.open(BytesIO(content)) as pdf:
            text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    except Exception as e:
        raise HTTPException(400, f"Could not read PDF: {e}")

    if "MCC Group Pty Ltd" in text or re.search(r"from_MCC_Group", filename, re.IGNORECASE):
        contractor = "MCC Group"
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM invoice_imports WHERE filename=%s AND contractor=%s",
                            (filename, contractor))
                if cur.fetchone():
                    return {"status": "skipped", "filename": filename}

    if "Central Highlands Mining Services" in text or re.search(r"Quote-00001532", filename, re.IGNORECASE):
        contractor = "CHMS"
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM invoice_imports WHERE filename=%s AND contractor=%s",
                            (filename, contractor))
                if cur.fetchone():
                    return {"status": "skipped", "filename": filename}

    try:
        inv = parse_invoice_pdf(text, filename, contractor)
    except Exception as e:
        raise HTTPException(422, f"Could not parse invoice: {e}")

    lines = inv.pop("lines", [])

    match_summary = {}
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO invoices
                    (source_file,contractor,invoice_number,invoice_date,due_date,po_reference,
                     project,client,abn,subtotal,gst,total_aud,amount_paid,amount_due,status,pdf_data)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    RETURNING id
                """, (filename, contractor,
                      inv.get("invoice_number",""), inv.get("invoice_date",""),
                      inv.get("due_date",""), inv.get("po_reference",""),
                      inv.get("project",""), inv.get("client",""), inv.get("abn",""),
                      inv.get("subtotal",0), inv.get("gst",0), inv.get("total_aud",0),
                      inv.get("amount_paid",0), inv.get("amount_due",0), inv.get("status","Unpaid"),
                      psycopg2.Binary(content)))
                invoice_id = cur.fetchone()["id"]

                if lines:
                    psycopg2.extras.execute_batch(cur, """
                        INSERT INTO invoice_lines
                        (invoice_id,contractor,invoice_number,description,quantity,unit_price,gst_rate,amount,category,
                         line_date,site_name,hole_num,activity_code,unit,chargeable,source_category)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, [(invoice_id, contractor, inv.get("invoice_number",""),
                           l["description"], l["quantity"], l["unit_price"],
                           l["gst_rate"], l["amount"], l["category"],
                           l.get("line_date"), l.get("site_name"), l.get("hole_num"),
                           l.get("activity_code"), l.get("unit"), l.get("chargeable"),
                           l.get("source_category")) for l in lines])

                # Run matching — don't let this crash the import
                try:
                    match_invoice_to_eos(cur, invoice_id, contractor, inv.get("po_reference",""))
                except Exception:
                    pass

                cur.execute("""
                    SELECT
                        COUNT(*) AS total,
                        SUM(CASE WHEN match_status='exact_match' THEN 1 ELSE 0 END) AS exact,
                        SUM(CASE WHEN match_status='close_match' THEN 1 ELSE 0 END) AS close,
                        SUM(CASE WHEN match_status='no_eos_data' THEN 1 ELSE 0 END) AS no_eos,
                        SUM(CASE WHEN match_status='invoice_over_eos' THEN 1 ELSE 0 END) AS over,
                        SUM(CASE WHEN match_status='invoice_under_eos' THEN 1 ELSE 0 END) AS under,
                        SUM(COALESCE(amount,0)) AS invoiced,
                        SUM(COALESCE(matched_eos_cost,0)) AS matched_eos,
                        SUM(COALESCE(variance,0)) AS variance
                    FROM invoice_lines WHERE invoice_id=%s
                """, (invoice_id,))
                match_summary = dict(cur.fetchone() or {})

                # Record import — use INSERT OR IGNORE equivalent
                try:
                    cur.execute("INSERT INTO invoice_imports (filename,contractor) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                                (filename, contractor))
                except Exception:
                    cur.execute("INSERT INTO invoice_imports (filename) VALUES (%s) ON CONFLICT DO NOTHING", (filename,))
            conn.commit()
    except Exception as e:
        raise HTTPException(500, f"Database error: {str(e)}")

    return {
        "status": "imported",
        "filename": filename,
        "invoice_number": inv.get("invoice_number", filename),
        "project": inv.get("project", ""),
        "total_aud": inv.get("total_aud", 0),
        "line_count": len(lines),
        "match_summary": match_summary,
    }


@app.post("/invoices/manual")
async def create_manual_invoice(request: Request):
    payload = await request.json()
    contractor = str(payload.get("contractor") or "").strip() or "Allianz Drilling"
    invoice_number = str(payload.get("invoice_number") or "").strip()
    description = str(payload.get("description") or "").strip()
    if not invoice_number:
        raise HTTPException(400, "invoice_number is required")
    if not description:
        raise HTTPException(400, "description is required")

    def num(value, default=0.0):
        try:
            if value is None or value == "":
                return default
            return float(str(value).replace("$", "").replace(",", "").strip())
        except Exception:
            return default

    quantity = num(payload.get("quantity"), 1.0) or 1.0
    unit_price = num(payload.get("unit_price"), 0.0)
    amount = num(payload.get("amount"), 0.0)
    if amount == 0 and unit_price:
        amount = round(quantity * unit_price, 2)
    if unit_price == 0 and quantity:
        unit_price = round(amount / quantity, 2) if amount else 0.0
    gst = num(payload.get("gst"), round(amount * 0.1, 2) if amount else 0.0)
    total_aud = num(payload.get("total_aud"), round(amount + gst, 2))
    amount_paid = num(payload.get("amount_paid"), 0.0)
    amount_due = num(payload.get("amount_due"), round(max(0, total_aud - amount_paid), 2))
    status = str(payload.get("status") or ("Paid" if amount_due == 0 and total_aud > 0 else "Unpaid")).strip()
    category = str(payload.get("category") or categorise_invoice_line(description)).strip() or "other"

    inv = {
        "invoice_number": invoice_number,
        "invoice_date": str(payload.get("invoice_date") or "").strip(),
        "due_date": str(payload.get("due_date") or "").strip(),
        "po_reference": str(payload.get("po_reference") or "").strip(),
        "project": str(payload.get("project") or "Ironbark").strip() or "Ironbark",
        "client": str(payload.get("client") or "Argo Coal Management Pty Ltd").strip(),
        "abn": str(payload.get("abn") or "").strip(),
        "subtotal": amount,
        "gst": gst,
        "total_aud": total_aud,
        "amount_paid": amount_paid,
        "amount_due": amount_due,
        "status": status,
    }
    line = {
        "description": description,
        "quantity": quantity,
        "unit_price": unit_price,
        "gst_rate": str(payload.get("gst_rate") or "10%").strip(),
        "amount": amount,
        "category": category,
        "line_date": str(payload.get("line_date") or inv["invoice_date"]).strip(),
        "site_name": str(payload.get("site_name") or "").strip(),
        "hole_num": str(payload.get("hole_num") or "").strip(),
        "activity_code": str(payload.get("activity_code") or category).strip(),
        "unit": str(payload.get("unit") or "each").strip(),
        "chargeable": payload.get("chargeable", True),
        "source_category": str(payload.get("source_category") or "manual").strip(),
    }

    source_file = f"manual:{contractor}:{invoice_number}"
    match_summary = {}
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO invoices
                    (source_file,contractor,invoice_number,invoice_date,due_date,po_reference,
                     project,client,abn,subtotal,gst,total_aud,amount_paid,amount_due,status,pdf_data)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NULL)
                    RETURNING id
                """, (source_file, contractor, inv["invoice_number"], inv["invoice_date"],
                      inv["due_date"], inv["po_reference"], inv["project"], inv["client"], inv["abn"],
                      inv["subtotal"], inv["gst"], inv["total_aud"], inv["amount_paid"], inv["amount_due"], inv["status"]))
                invoice_id = cur.fetchone()["id"]
                cur.execute("""
                    INSERT INTO invoice_lines
                    (invoice_id,contractor,invoice_number,description,quantity,unit_price,gst_rate,amount,category,
                     line_date,site_name,hole_num,activity_code,unit,chargeable,source_category)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (invoice_id, contractor, inv["invoice_number"], line["description"], line["quantity"],
                      line["unit_price"], line["gst_rate"], line["amount"], line["category"], line["line_date"],
                      line["site_name"], line["hole_num"], line["activity_code"], line["unit"], line["chargeable"],
                      line["source_category"]))
                try:
                    match_invoice_to_eos(cur, invoice_id, contractor, inv["po_reference"])
                except Exception:
                    pass
                cur.execute("""
                    SELECT
                        COUNT(*) AS total,
                        SUM(CASE WHEN match_status='exact_match' THEN 1 ELSE 0 END) AS exact,
                        SUM(CASE WHEN match_status='close_match' THEN 1 ELSE 0 END) AS close,
                        SUM(CASE WHEN match_status='no_eos_data' THEN 1 ELSE 0 END) AS no_eos,
                        SUM(CASE WHEN match_status='invoice_over_eos' THEN 1 ELSE 0 END) AS over,
                        SUM(CASE WHEN match_status='invoice_under_eos' THEN 1 ELSE 0 END) AS under,
                        SUM(COALESCE(amount,0)) AS invoiced,
                        SUM(COALESCE(matched_eos_cost,0)) AS matched_eos,
                        SUM(COALESCE(variance,0)) AS variance
                    FROM invoice_lines WHERE invoice_id=%s
                """, (invoice_id,))
                match_summary = dict(cur.fetchone() or {})
            conn.commit()
        return {
            "status": "created",
            "id": invoice_id,
            "invoice_number": inv["invoice_number"],
            "contractor": contractor,
            "total_aud": inv["total_aud"],
            "line_count": 1,
            "match_summary": match_summary,
        }
    except Exception as e:
        raise HTTPException(500, f"Database error: {str(e)}")


@app.get("/invoices")
def get_invoices(contractor: str = Query(...)):
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT id, source_file, contractor, invoice_number,
                           project, invoice_date, due_date, po_reference, client, abn,
                           subtotal, gst, total_aud, amount_paid, amount_due,
                           status, notes, billing_month, query_notes, version
                    FROM invoices
                    WHERE contractor=%s
                    ORDER BY invoice_date DESC
                """, (contractor,))
                invoices = [dict(r) for r in cur.fetchall()]

                for inv in invoices:
                    try:
                        cur.execute("""
                            SELECT COUNT(*) AS line_count,
                                   SUM(CASE WHEN match_status='exact_match' THEN 1 ELSE 0 END) AS exact_matches,
                                   SUM(CASE WHEN match_status='close_match' THEN 1 ELSE 0 END) AS close_matches,
                                   SUM(CASE WHEN match_status LIKE '%%over%%' THEN 1 ELSE 0 END) AS over_count,
                                   SUM(CASE WHEN match_status LIKE '%%under%%' THEN 1 ELSE 0 END) AS under_count,
                                   SUM(CASE WHEN match_status='no_eos_data' THEN 1 ELSE 0 END) AS unmatched_count
                            FROM invoice_lines WHERE invoice_id=%s
                        """, (inv["id"],))
                        row = cur.fetchone()
                        if row:
                            inv.update(dict(row))
                        else:
                            inv.update({"line_count":0,"exact_matches":0,"close_matches":0,"over_count":0,"under_count":0,"unmatched_count":0})
                    except Exception:
                        inv.update({"line_count":0,"exact_matches":0,"close_matches":0,"over_count":0,"under_count":0,"unmatched_count":0})

                return invoices
    except Exception as e:
        raise HTTPException(500, f"Invoices error: {str(e)}")


@app.get("/invoices/{invoice_id}/lines")
def get_invoice_lines(invoice_id: int):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM invoice_lines WHERE invoice_id=%s ORDER BY line_date NULLS LAST, id", (invoice_id,))
            return [dict(r) for r in cur.fetchall()]



@app.patch("/invoices/{invoice_id}")
async def update_invoice(invoice_id: int, request: Request):
    payload = await request.json()
    safe = {"billing_month","status","notes","amount_paid","amount_due","po_reference","query_notes","version","contractor","project","invoice_date","subtotal","gst","total_aud","invoice_number"}
    u = {k:v for k,v in payload.items() if k in safe and k != "version"}
    if not u: raise HTTPException(400, "No valid fields")
    # Auto-increment version on any edit
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT version FROM invoices WHERE id=%s", (invoice_id,))
            row = cur.fetchone()
            current_version = int(row["version"] or 1) if row else 1
            u["version"] = current_version + 1
            set_clause = ", ".join(f"{k}=%s" for k in u)
            vals = list(u.values()) + [invoice_id]
            cur.execute(f"UPDATE invoices SET {set_clause} WHERE id=%s", vals)
        conn.commit()
    return {"status": "updated", "version": u["version"]}


@app.get("/invoices/{invoice_id}/pdf")
def get_invoice_pdf(invoice_id: int):
    """Return the stored PDF file for viewing/downloading."""
    from fastapi.responses import Response
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT source_file, pdf_data FROM invoices WHERE id=%s", (invoice_id,))
            row = cur.fetchone()
            if not row or not row["pdf_data"]:
                raise HTTPException(404, "PDF not found for this invoice")
            filename = row["source_file"] or f"invoice_{invoice_id}.pdf"
            return Response(
                content=bytes(row["pdf_data"]),
                media_type="application/pdf",
                headers={"Content-Disposition": f'inline; filename="{filename}"'}
            )


@app.delete("/invoices/{invoice_id}")
def delete_invoice(invoice_id: int):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM invoices WHERE id=%s", (invoice_id,))
        conn.commit()
    return {"status":"deleted"}


@app.get("/reconciliation")
def get_reconciliation(contractor: str = Query(...)):
    """
    Full reconciliation: invoice totals vs EOS activity costs by category.
    """
    with get_conn() as conn:
        with conn.cursor() as cur:

            # Invoice totals by category
            cur.execute("""
                SELECT l.category,
                    SUM(l.amount) AS invoice_total,
                    COUNT(l.id)   AS line_count,
                    COUNT(DISTINCT l.invoice_number) AS invoice_count
                FROM invoice_lines l
                WHERE l.contractor=%s
                GROUP BY l.category
                ORDER BY invoice_total DESC
            """, (contractor,))
            inv_cats = {r["category"]: dict(r) for r in cur.fetchall()}

            # EOS costs by category
            cur.execute("""
                SELECT
                    CASE
                        WHEN code IN ('Drill_Core','Drill_Chip_or_Open_hole') THEN 'drilling_metres'
                        WHEN code LIKE '%%Standby%%' THEN 'standby'
                        WHEN code LIKE '%%Travel%%' THEN 'travel'
                        WHEN code LIKE '%%Safety%%' OR code LIKE '%%Training%%'
                          OR code LIKE '%%Prestart%%' THEN 'safety'
                        WHEN code LIKE 'D_Backhoe%%' OR code LIKE 'D_Water%%' THEN 'equipment'
                        WHEN code LIKE '%%Setup%%' OR code LIKE '%%Surface%%' THEN 'setup'
                        ELSE 'active'
                    END AS category,
                    SUM(line_cost) AS eos_total,
                    COUNT(*) AS activity_count
                FROM activities
                WHERE contractor=%s AND line_cost IS NOT NULL
                GROUP BY 1
            """, (contractor,))
            eos_cats = {r["category"]: dict(r) for r in cur.fetchall()}

            # Invoice header summary
            cur.execute("""
                SELECT invoice_number, invoice_date, due_date, po_reference,
                       total_aud, amount_paid, amount_due, status
                FROM invoices WHERE contractor=%s ORDER BY invoice_date
            """, (contractor,))
            invoices = [dict(r) for r in cur.fetchall()]

            # Grand totals
            cur.execute("SELECT SUM(total_aud) AS t, SUM(amount_paid) AS p, SUM(amount_due) AS d FROM invoices WHERE contractor=%s", (contractor,))
            inv_totals = dict(cur.fetchone())

            cur.execute("SELECT SUM(line_cost) AS t FROM activities WHERE contractor=%s AND line_cost IS NOT NULL", (contractor,))
            eos_grand = float(cur.fetchone()["t"] or 0)

            # Line-level discrepancies
            cur.execute("""
                SELECT l.invoice_number, l.description, l.category,
                       l.quantity, l.unit_price, l.amount,
                       l.matched_eos_cost, l.variance, l.match_status
                FROM invoice_lines l
                WHERE l.contractor=%s
                  AND l.match_status NOT IN ('exact_match','no_eos_data')
                ORDER BY ABS(COALESCE(l.variance,0)) DESC
                LIMIT 50
            """, (contractor,))
            discrepancies = [dict(r) for r in cur.fetchall()]

    # Build comparison rows
    all_cats = sorted(set(list(inv_cats.keys()) + list(eos_cats.keys())))
    comparison = []
    for cat in all_cats:
        inv = float(inv_cats.get(cat, {}).get("invoice_total", 0) or 0)
        eos = float(eos_cats.get(cat, {}).get("eos_total", 0) or 0)
        var = inv - eos
        comparison.append({
            "category":      cat,
            "invoice_total": round(inv, 2),
            "eos_total":     round(eos, 2),
            "variance":      round(var, 2),
            "variance_pct":  round(var/eos*100, 1) if eos else None,
            "status":        "match" if abs(var) < 1 else "over" if var > 0 else "under",
        })

    return {
        "invoices":      invoices,
        "comparison":    comparison,
        "discrepancies": discrepancies,
        "totals": {
            "invoice_total":  round(float(inv_totals.get("t") or 0), 2),
            "invoice_paid":   round(float(inv_totals.get("p") or 0), 2),
            "invoice_due":    round(float(inv_totals.get("d") or 0), 2),
            "eos_total":      round(eos_grand, 2),
            "grand_variance": round(float(inv_totals.get("t") or 0) - eos_grand, 2),
        },
    }


@app.post("/reconciliation/rematch")
def rematch_all(contractor: str = Query(...)):
    """Re-run EOS matching for all invoices of a contractor."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, po_reference FROM invoices WHERE contractor=%s", (contractor,))
            invoices = cur.fetchall()
            for inv in invoices:
                match_invoice_to_eos(cur, inv["id"], contractor, inv["po_reference"])
        conn.commit()
    return {"status": "rematched", "count": len(invoices)}


@app.post("/reconciliation/ai-audit")
async def ai_audit_reconciliation(request: Request):
    """Use Gemini AI to intelligently audit invoices against EOS field data."""
    payload = await request.json()
    contractor = payload.get("contractor", "Allianz Drilling")
    month = payload.get("month", "")

    if not GEMINI_API_KEY:
        raise HTTPException(500, "GEMINI_API_KEY not configured on server. Add it in Render environment variables.")

    with get_conn() as conn:
        with conn.cursor() as cur:
            # Get invoices
            if month:
                cur.execute("""SELECT invoice_number, invoice_date, billing_month, total_aud,
                       amount_paid, amount_due, status, po_reference
                    FROM invoices WHERE contractor=%s
                    AND (billing_month=%s OR invoice_date LIKE %s)
                    ORDER BY invoice_date""", (contractor, month, f"%{month}%"))
            else:
                cur.execute("""SELECT invoice_number, invoice_date, billing_month, total_aud,
                       amount_paid, amount_due, status, po_reference
                    FROM invoices WHERE contractor=%s ORDER BY invoice_date""", (contractor,))
            invoices = [dict(r) for r in cur.fetchall()]

            inv_lines = []
            for inv in invoices:
                cur.execute("""SELECT il.* FROM invoice_lines il
                    JOIN invoices i ON i.id=il.invoice_id
                    WHERE i.invoice_number=%s AND i.contractor=%s""",
                    (inv["invoice_number"], contractor))
                inv_lines.extend([dict(r) for r in cur.fetchall()])

            # EOS summary
            cur.execute("""SELECT date, hole_num, site_name, code, notes, total_time,
                       total_metres, bit_type, unit_rate, quantity, line_cost, rate_basis
                FROM activities WHERE contractor=%s AND line_cost IS NOT NULL
                ORDER BY date""", (contractor,))
            activities = [dict(r) for r in cur.fetchall()]

            cur.execute("SELECT code, description, rate, unit FROM hourly_rates WHERE contractor=%s", (contractor,))
            rates = [dict(r) for r in cur.fetchall()]

            cur.execute("""SELECT consumable, SUM(COALESCE(NULLIF(quantity,'')::FLOAT,1)) AS total_qty,
                       MAX(unit_price) AS unit_price, SUM(line_cost) AS total_cost
                FROM consumables WHERE contractor=%s AND line_cost > 0
                GROUP BY consumable""", (contractor,))
            consumables = [dict(r) for r in cur.fetchall()]

    inv_total = sum(float(i.get("total_aud") or 0) for i in invoices)
    eos_total = sum(float(a.get("line_cost") or 0) for a in activities)
    cons_total = sum(float(c.get("total_cost") or 0) for c in consumables)

    eos_by_code = {}
    for a in activities:
        code = a.get("code") or "unknown"
        eos_by_code[code] = eos_by_code.get(code, 0) + float(a.get("line_cost") or 0)

    eos_by_date = {}
    for a in activities:
        d = a.get("date") or "unknown"
        eos_by_date[d] = eos_by_date.get(d, 0) + float(a.get("line_cost") or 0)

    prompt = f"""You are a mining industry financial auditor specialising in drilling contractor invoices for Australian coal exploration. Analyse this data and identify discrepancies, overcharges, and anomalies.

CONTRACTOR: {contractor}
PERIOD: {month or 'All dates'}

INVOICES ({len(invoices)} invoices, total ${inv_total:,.2f}):
{json.dumps(invoices, indent=2, default=str)[:3000]}

INVOICE LINE ITEMS ({len(inv_lines)} lines):
{json.dumps(inv_lines[:50], indent=2, default=str)[:3000]}

EOS FIELD DATA SUMMARY:
Total EOS calculated cost: ${eos_total:,.2f}
Total consumables cost: ${cons_total:,.2f}
Grand variance (Invoice - EOS - Consumables): ${inv_total - eos_total - cons_total:,.2f}

Cost by activity code:
{json.dumps(eos_by_code, indent=2, default=str)[:2000]}

Daily totals (EOS):
{json.dumps(dict(list(eos_by_date.items())[:30]), indent=2, default=str)[:1500]}

SCHEDULE OF RATES:
{json.dumps(rates[:30], indent=2, default=str)[:1500]}

CONSUMABLES:
{json.dumps(consumables, indent=2, default=str)[:1000]}

Provide your audit as JSON with this exact structure:
{{
  "summary": "2-3 sentence executive summary",
  "grand_variance": {{"invoiced": 0, "eos_calculated": 0, "consumables": 0, "difference": 0, "percentage": 0}},
  "findings": [
    {{
      "severity": "critical or warning or info",
      "category": "overcharge or undercharge or rate_mismatch or missing_data or duplicate or suspicious",
      "title": "short title",
      "detail": "detailed explanation",
      "amount": 0,
      "recommendation": "action to take"
    }}
  ],
  "rate_check": [
    {{
      "code": "H_Active",
      "schedule_rate": 745,
      "invoiced_rate": 0,
      "status": "match or mismatch or not_found",
      "note": "explanation"
    }}
  ],
  "recommendations": ["list of action items"]
}}

Return ONLY valid JSON. No markdown fences. No text outside the JSON object."""

    url = gemini_generate_content_url(GEMINI_API_KEY)
    gemini_payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.2, "maxOutputTokens": 8192}
    }

    try:
        async with httpx.AsyncClient(timeout=90.0) as client:
            resp = await client.post(url, json=gemini_payload)
    except Exception as e:
        raise HTTPException(502, f"Gemini request failed: {str(e)}")

    if resp.status_code != 200:
        raise HTTPException(502, f"Gemini API error: {resp.status_code} - {resp.text[:200]}")

    result = resp.json()
    try:
        text = result["candidates"][0]["content"]["parts"][0]["text"]
        text = text.strip()
        if text.startswith("```"): text = text.split("\n", 1)[1]
        if text.endswith("```"): text = text[:-3]
        text = text.strip()
        if text.lower().startswith("json"): text = text[4:].strip()
        audit = json.loads(text)
    except (KeyError, IndexError, json.JSONDecodeError) as e:
        raw = result.get("candidates", [{}])[0].get("content", {}).get("parts", [{}])[0].get("text", "")
        return {"status": "partial", "raw_response": raw[:3000], "error": str(e)}

    return {"status": "ok", "audit": audit}


@app.post("/reprice")
def reprice_activities(contractor: str = Query(...)):
    """Fast rationalisation: fix dates, holes, sites, locations, then batch reprice."""

    SITE_TO_HOLE = {
        "CD-26-001":"CD1817C","CD-26-002":"CD1818C","CD-26-004":"CD1819C",
        "CD-26-005":"CD1820C","CD-26-006":"CD1821C","CD-26-007":"CD1822C",
        "CD-26-008":"CD1823C","CD-26-009":"CD1824C","CD-26-010":"CD1825C",
        "CD-26-011":"CD1826C","CD-26-012":"CD1827C","CD-26-013":"CD1828C",
        "CD-26-014":"CD1829C","CD-26-016":"CD1830C","CD-26-017":"CD1831C",
        "CD-26-018":"CD1832C","CD-26-019":"CD1833C","CD-26-020":"CD1834C",
        "CD-26-021":"CD1833C","CD-26-022":"CD1834C","CD-26-023":"CD1835C",
    }
    REVERSE_HOLE = {}
    for k, v in SITE_TO_HOLE.items():
        REVERSE_HOLE[v] = k

    months_map = {"january":"01","february":"02","march":"03","april":"04",
                  "may":"05","june":"06","july":"07","august":"08",
                  "september":"09","october":"10","november":"11","december":"12"}

    def normalise_date(d):
        if not d: return d
        d = d.strip()
        if re.match(r"^\d{2}/\d{2}/\d{4}$", d): return d
        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", d)
        if m: return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"
        m = re.match(r"^(\d{1,2})-(\w+)-(\d{4})$", d)
        if m:
            mon = months_map.get(m.group(2).lower(), "01")
            return f"{int(m.group(1)):02d}/{mon}/{m.group(3)}"
        return d

    stats = {"total":0,"dates_fixed":0,"holes_fixed":0,"sites_fixed":0,"priced":0,"locked_skipped":0}
    skipped_codes = {}

    with get_conn() as conn:
        with conn.cursor() as cur:
            # ── Load ALL data in 4 queries total ──────────────────────
            cur.execute("SELECT * FROM activities WHERE contractor=%s", (contractor,))
            rows = [dict(r) for r in cur.fetchall()]
            stats["total"] = len(rows)
            locked_keys = locked_activity_sheet_keys(cur, contractor)

            cur.execute("SELECT * FROM drilling_rates WHERE contractor=%s", (contractor,))
            all_dr = [dict(r) for r in cur.fetchall()]

            cur.execute("SELECT * FROM hourly_rates WHERE contractor=%s", (contractor,))
            all_hr = [dict(r) for r in cur.fetchall()]

            dr_years = sorted(set(r["year"] for r in all_dr))
            hr_years = sorted(set(r["year"] for r in all_hr))

        # ── Build rate lookup dicts in memory ─────────────────────────
        # hourly: {(year, code): rate}
        hr_lookup = {}
        for r in all_hr:
            hr_lookup[(r["year"], r["code"])] = float(r["rate"])

        # drilling: {(year, bit_type, depth): rate}  — depth is the matching band
        dr_lookup = {}
        for r in all_dr:
            key = (r["year"], normalise_drilling_bit_key(r["bit_type"]))
            if key not in dr_lookup: dr_lookup[key] = []
            dr_lookup[key].append((float(r["depth_from"]), float(r["depth_to"]), float(r["rate"])))

        rate_context = build_rate_context(all_hr, all_dr)

        def get_hr_mem(code, year):
            for try_year in [year, str(int(year)-1) if year.isdigit() else year, str(int(year)+1) if year.isdigit() else year, "2025"]:
                r = hr_lookup.get((try_year, code))
                if r is not None: return r
            return None

        def get_dr_mem(bit_key, depth, year):
            for try_year in [year, str(int(year)-1) if year.isdigit() else year, str(int(year)+1) if year.isdigit() else year, "2025"]:
                bands = dr_lookup.get((try_year, bit_key), [])
                for frm, to, rate in bands:
                    if frm <= depth < to:
                        return rate
            return None

        def extract_year(date_str):
            if not date_str: return "2026"
            parts = date_str.replace("-","/").split("/")
            for p in parts:
                if len(p) == 4 and p.isdigit(): return p
            return "2026"

        def parse_hours(t):
            if not t: return 0
            t = str(t).strip()
            m = re.match(r"(\d+):(\d+)", t)
            if m: return int(m.group(1)) + int(m.group(2))/60.0
            try: return float(t)
            except: return 0

        # ── Process all rows in memory ────────────────────────────────
        batch_updates = []
        for row in rows:
            if row_is_in_locked_sheet(row, locked_keys):
                stats["locked_skipped"] += 1
                continue
            updates = {}
            rid = row["id"]

            # 1. Date
            old_date = row.get("date","") or ""
            new_date = normalise_date(old_date)
            if new_date != old_date:
                updates["date"] = new_date
                row["date"] = new_date
                stats["dates_fixed"] += 1

            # 2. Hole ID from site
            hole = row.get("hole_num","") or ""
            site = row.get("site_name","") or ""
            if (not hole or hole == "0") and site in SITE_TO_HOLE:
                updates["hole_num"] = SITE_TO_HOLE[site]; row["hole_num"] = SITE_TO_HOLE[site]; stats["holes_fixed"] += 1
            elif hole in SITE_TO_HOLE:
                updates["hole_num"] = SITE_TO_HOLE[hole]; row["hole_num"] = SITE_TO_HOLE[hole]; stats["holes_fixed"] += 1
            elif site in SITE_TO_HOLE and hole != SITE_TO_HOLE[site]:
                updates["hole_num"] = SITE_TO_HOLE[site]; row["hole_num"] = SITE_TO_HOLE[site]; stats["holes_fixed"] += 1

            # 3. Site from hole
            if not site and row.get("hole_num","") in REVERSE_HOLE:
                updates["site_name"] = REVERSE_HOLE[row["hole_num"]]; stats["sites_fixed"] += 1

            # 4. Location
            loc = (row.get("location","") or "").lower()
            if loc in ("cdm sth","cdm south","cdm","carborough downs mine"):
                updates["location"] = "Carborough Downs"

            # 5. Reprice (in memory — no DB queries)
            year = extract_year(row.get("date",""))
            code = row.get("code","") or ""
            hours = parse_hours(row.get("total_time",""))
            bit_type = row.get("bit_type","") or ""
            depth = None
            mf = row.get("metres_from")
            mt = row.get("metres_to")
            if mf is not None and mt is not None:
                try: depth = (float(mf) + float(mt)) / 2
                except: pass
            metres = 0
            try: metres = float(row.get("total_metres",0) or 0)
            except: pass

            line_cost = None; unit_rate = None; quantity = None; rate_basis = None
            mcc_fix = mcc_reprice_from_row(row)

            # Drilling metres
            if metres > 0 and bit_type and mf is not None and mt is not None:
                priced = drilling_schedule_cost(row, rate_context)
                if priced is not None:
                    unit_rate = priced["unit_rate"]
                    quantity = priced["quantity"]
                    line_cost = priced["line_cost"]
                    rate_basis = priced["rate_basis"]
            elif code in DRILLING_METRE_CODES:
                unit_rate = 0; quantity = 0; line_cost = 0
                rate_basis = "drilling time covered by metreage; no metres recorded"

            # Day rates
            if line_cost is None:
                for dk, (dk_code, dk_unit) in DAY_RATE_CODES.items():
                    if dk in code or code in dk:
                        r = get_hr_mem(dk_code, year)
                        if r is not None:
                            unit_rate = r; quantity = 1; line_cost = r; rate_basis = f"${r:,.2f}/{dk_unit}"
                        break

            # Not chargeable
            if line_cost is None and is_not_chargeable_code(code, contractor):
                unit_rate = 0; quantity = round(hours,2) if hours > 0 else 1
                line_cost = 0; rate_basis = "not chargeable"

            # MCC weekly sheets often import as H_Active with equipment/role in notes.
            if line_cost is None and mcc_fix:
                code = mcc_fix["code"]
                row["code"] = code
                year = mcc_fix["rate_year"]
                unit_rate = mcc_fix["unit_rate"]
                quantity = mcc_fix["quantity"]
                line_cost = mcc_fix["line_cost"]
                rate_basis = mcc_fix["rate_basis"]
                updates["code"] = code

            # Equipment packages
            if line_cost is None and code.startswith("E_"):
                r = get_hr_mem(code, year)
                qty = round(hours,2) if hours > 0 else (row_num(row.get("quantity")) or 1)
                if r is not None:
                    unit_rate = r; quantity = qty
                    line_cost = round(r * qty, 2); rate_basis = f"equipment ${r:,.2f}/unit x {qty}"
                elif row.get("line_cost") is not None:
                    unit_rate = row.get("unit_rate"); quantity = row.get("quantity") or qty
                    line_cost = row.get("line_cost"); rate_basis = row.get("rate_basis") or "equipment charge"

            # Standby / inactive
            if line_cost is None and (code in STANDBY_CODES or "Standby" in code or code in INACTIVE_CODES):
                r = get_hr_mem(code, year)
                if r is None:
                    r = get_hr_mem("H_Inactive", year)
                if r is not None:
                    qty = round(hours,2) if hours > 0 else 1
                    unit_rate = r; quantity = qty; line_cost = round(r * qty, 2)
                    rate_basis = f"inactive ${r:,.2f}/hr x {qty}"

            # Active codes
            if line_cost is None and hours > 0 and (code in ACTIVE_CODES or code.startswith("H_")):
                r = get_hr_mem(code, year)
                if r is None:
                    r = get_hr_mem("H_Active", year)
                if r is not None:
                    unit_rate = r; quantity = round(hours,2)
                    line_cost = round(r * hours, 2); rate_basis = f"active ${r:,.2f}/hr x {hours:.2f}h"

            # Fallback: hours but no match
            if line_cost is None and hours > 0 and code:
                r = get_hr_mem(code, year)
                if r is None:
                    r = get_hr_mem("H_Active", year)
                if r is not None:
                    unit_rate = r; quantity = round(hours,2)
                    line_cost = round(r * hours, 2); rate_basis = f"fallback ${r:,.2f}/hr x {hours:.2f}h"

            if line_cost is not None:
                updates["rate_year"] = year
                updates["unit_rate"] = unit_rate
                updates["quantity"] = quantity
                updates["line_cost"] = line_cost
                updates["rate_basis"] = rate_basis
                stats["priced"] += 1
            else:
                skipped_codes[code or "empty"] = skipped_codes.get(code or "empty", 0) + 1

            if updates:
                batch_updates.append((updates, rid))

        # ── Batch update activities in one transaction ────────────────
        with conn.cursor() as cur:
            for updates, rid in batch_updates:
                set_clause = ", ".join(f"{k}=%s" for k in updates)
                vals = list(updates.values()) + [rid]
                cur.execute(f"UPDATE activities SET {set_clause} WHERE id=%s", vals)

            cur.execute("SELECT * FROM activities WHERE contractor=%s", (contractor,))
            repriced_rows = [dict(r) for r in cur.fetchall()]
            locked_target_totals = minimum_shift_group_totals(repriced_rows, locked_keys)
            repriced_rows, restored_coreplan_rows = update_restored_coreplan_activity_line_costs(cur, repriced_rows)
            min_shift_updates = []
            original_by_id = {r["id"]: r for r in repriced_rows}
            excluded_min_shift_keys = minimum_shift_excluded_keys(cur, contractor) if minimum_shift_rule(contractor) else set()
            for row in adjust_imported_minimum_shift_rows(repriced_rows, contractor, excluded_min_shift_keys, locked_target_totals):
                if not is_imported_minimum_shift_row(row):
                    continue
                original = original_by_id.get(row["id"])
                if not original:
                    continue
                if (
                    round(float(original.get("line_cost") or 0), 2) != round(float(row.get("line_cost") or 0), 2)
                    or round(float(original.get("quantity") or 0), 2) != round(float(row.get("quantity") or 0), 2)
                    or (original.get("rate_basis") or "") != (row.get("rate_basis") or "")
                ):
                    min_shift_updates.append(row)
            for row in min_shift_updates:
                cur.execute(
                    """
                    UPDATE activities
                    SET quantity=%s, line_cost=%s, rate_basis=%s
                    WHERE id=%s
                    """,
                    (row.get("quantity"), row.get("line_cost"), row.get("rate_basis"), row["id"]),
                )
            stats["minimum_shift_rows_adjusted"] = len(min_shift_updates)
            stats["coreplan_line_costs_restored"] = restored_coreplan_rows

            # ── Reprice consumables ───────────────────────────────────────
            try:
                cur.execute("SELECT * FROM consumables WHERE contractor=%s", (contractor,))
                cons_rows = [dict(r) for r in cur.fetchall()]

                cur.execute("SELECT * FROM consumable_rates WHERE contractor=%s", (contractor,))
                all_cr = [dict(r) for r in cur.fetchall()]

                # Build consumable rate lookup: {normalised_product: unit_price}
                cr_lookup = {}
                for r in all_cr:
                    key = r["product"].strip().upper()
                    cr_lookup[key] = float(r["unit_price"] or 0)
                    cr_lookup[key.replace(" ", "")] = float(r["unit_price"] or 0)

                cons_priced = 0
                cons_locked_skipped = 0
                for crow in cons_rows:
                    if row_is_in_locked_sheet(crow, locked_keys):
                        cons_locked_skipped += 1
                        continue
                    if contractor == "Mitchells Drilling" and crow.get("line_cost") is not None:
                        continue
                    product = (crow.get("consumable") or crow.get("type") or "").strip()
                    product_upper = product.upper()
                    product_nospace = product_upper.replace(" ", "")

                    price = cr_lookup.get(product_upper)
                    if price is None:
                        price = cr_lookup.get(product_nospace)
                    if price is None:
                        for rk, rv in cr_lookup.items():
                            if rk in product_upper or product_upper in rk:
                                price = rv
                                break

                    if price is not None and price > 0:
                        qty = 1
                        try:
                            qty = float(crow.get("quantity") or 1)
                        except (ValueError, TypeError):
                            qty = 1
                        lc = round(price * qty, 2)
                        cur.execute("UPDATE consumables SET unit_price=%s, line_cost=%s WHERE id=%s",
                                    (price, lc, crow["id"]))
                        cons_priced += 1

                stats["consumables_priced"] = cons_priced
                stats["consumables_total"] = len(cons_rows)
                stats["consumables_locked_skipped"] = cons_locked_skipped
            except Exception as e:
                stats["consumables_error"] = str(e)
                stats["consumables_priced"] = 0
                stats["consumables_total"] = 0
                stats["consumables_locked_skipped"] = 0

        conn.commit()

    minimum_shift_topups = sync_allianz_minimum_shift_topups(contractor)

    return {
        "status": "rationalised",
        "total": stats["total"],
        "dates_fixed": stats["dates_fixed"],
        "holes_fixed": stats["holes_fixed"],
        "sites_fixed": stats["sites_fixed"],
        "priced": stats["priced"],
        "locked_skipped": stats["locked_skipped"],
        "consumables_priced": stats.get("consumables_priced", 0),
        "consumables_total": stats.get("consumables_total", 0),
        "consumables_locked_skipped": stats.get("consumables_locked_skipped", 0),
        "minimum_shift_rows_adjusted": stats.get("minimum_shift_rows_adjusted", 0),
        "drilling_rate_years": dr_years,
        "hourly_rate_years": hr_years,
        "minimum_shift_topups": minimum_shift_topups,
        "skipped_codes": skipped_codes,
    }


@app.post("/rates/recalculate-database")
async def recalculate_database_from_rates(request: Request):
    payload = await request.json()
    contractor = payload.get("contractor", "Allianz Drilling")

    cleanup = cleanup_coreplan_doubleups_for_contractor(contractor)
    minimum_shift_topups = sync_allianz_minimum_shift_topups(contractor)
    reprice = reprice_activities(contractor)
    reprice["minimum_shift_topups"] = reprice.get("minimum_shift_topups") or minimum_shift_topups

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, po_reference FROM invoices WHERE contractor=%s", (contractor,))
            invoices = [dict(r) for r in cur.fetchall()]
            for inv in invoices:
                match_invoice_to_eos(cur, inv["id"], contractor, inv.get("po_reference") or "")

            cur.execute("""
                SELECT
                    COUNT(*) AS line_count,
                    SUM(CASE WHEN match_status IN ('exact_match','close_match') THEN 1 ELSE 0 END) AS accounted_lines,
                    SUM(CASE WHEN match_status='no_eos_data' THEN 1 ELSE 0 END) AS no_eos_lines,
                    SUM(COALESCE(amount,0)) AS invoice_total,
                    SUM(COALESCE(matched_eos_cost,0)) AS matched_eos_total,
                    SUM(COALESCE(variance,0)) AS variance
                FROM invoice_lines
                WHERE contractor=%s
            """, (contractor,))
            invoice_match_summary = dict(cur.fetchone() or {})
        conn.commit()

    return {
        "status": "recalculated",
        "contractor": contractor,
        "cleanup": cleanup,
        "reprice": reprice,
        "invoices_rematched": len(invoices),
        "invoice_match_summary": invoice_match_summary,
    }



# ── Borehole Planning ─────────────────────────────────────────────────────────

@lru_cache(maxsize=1)
def agd84_amg55_transformer():
    from pyproj import Transformer
    return Transformer.from_crs("EPSG:20355", "EPSG:4326", always_xy=True)


def agd84_amg55_to_wgs84(easting, northing):
    if easting in (None, "", 0) or northing in (None, "", 0):
        return None, None
    try:
        lng, lat = agd84_amg55_transformer().transform(float(easting), float(northing))
        if 110 <= lng <= 155 and -45 <= lat <= -10:
            return lat, lng
    except Exception:
        pass
    return None, None


def apply_borehole_wgs84(row: dict):
    lat, lng = agd84_amg55_to_wgs84(row.get("easting"), row.get("northing"))
    if lat is not None and lng is not None:
        row["lat"] = lat
        row["lng"] = lng
    return row


@app.get("/boreholes")
def get_boreholes(contractor: Optional[str] = Query(None)):
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                if contractor:
                    cur.execute("""
                        SELECT b.*,
                            COALESCE(SUM(a.line_cost),0) AS eos_cost,
                            COALESCE(SUM(CASE WHEN a.code LIKE 'Drill_%%'
                                OR a.code IN ('H_Min_Shift','H_Casing_Install','H_Circulation_Flush','H_Circulation_Lost','H_Reaming','H_Rig_Cementing','H_Surface_Setup','H_Tripping_Rods')
                                THEN a.line_cost ELSE 0 END),0) AS drilling_cost,
                            SUM(CASE WHEN a.code LIKE 'Drill_%%' THEN a.total_metres ELSE 0 END) AS drilling_metres,
                            SUM(a.total_metres) AS eos_metres
                        FROM boreholes b
                        LEFT JOIN activities a ON a.contractor=%s
                            AND (
                                a.hole_num=b.hole_id
                                OR (COALESCE(b.site_id,'')<>'' AND a.site_name=b.site_id)
                            )
                        WHERE b.contractor IN (%s, 'Company')
                        GROUP BY b.id ORDER BY b.drill_order
                    """, (contractor, contractor))
                else:
                    cur.execute("""
                        SELECT b.*,
                            COALESCE(SUM(a.line_cost),0) AS eos_cost,
                            COALESCE(SUM(CASE WHEN a.code LIKE 'Drill_%%'
                                OR a.code IN ('H_Min_Shift','H_Casing_Install','H_Circulation_Flush','H_Circulation_Lost','H_Reaming','H_Rig_Cementing','H_Surface_Setup','H_Tripping_Rods')
                                THEN a.line_cost ELSE 0 END),0) AS drilling_cost,
                            SUM(CASE WHEN a.code LIKE 'Drill_%%' THEN a.total_metres ELSE 0 END) AS drilling_metres,
                            SUM(a.total_metres) AS eos_metres
                        FROM boreholes b
                        LEFT JOIN activities a ON a.contractor=b.contractor
                            AND (
                                a.hole_num=b.hole_id
                                OR (COALESCE(b.site_id,'')<>'' AND a.site_name=b.site_id)
                            )
                        GROUP BY b.id ORDER BY b.drill_order
                    """)
                rows = [dict(r) for r in cur.fetchall()]
                if contractor:
                    deduped = {}
                    fallback_fields = {
                        "project","planned_year","site_id","drill_order","days_budgeted",
                        "bh_type","bit_type","purpose","easting","northing","rl",
                        "chip_depth","eoh_depth","total_core","seam_tk","lat","lng",
                        "drilling_budget_total","earthworks_budget_total","geophysical_budget_total",
                        "geological_support_budget_total","misc_budget_total","budget_total","actual_total","assigned_rig","scheduled_start",
                        "scheduled_end"
                    }
                    for row in rows:
                        key = row.get("hole_id")
                        if not key:
                            continue
                        current = deduped.get(key)
                        if not current:
                            deduped[key] = row
                        elif current.get("contractor") == "Company" and row.get("contractor") == contractor:
                            merged = {**current, **row}
                            for field in fallback_fields:
                                if merged.get(field) in (None, "", 0) and current.get(field) not in (None, "", 0):
                                    merged[field] = current.get(field)
                            deduped[key] = merged
                        elif row.get("contractor") == "Company":
                            for field in fallback_fields:
                                if current.get(field) in (None, "", 0) and row.get(field) not in (None, "", 0):
                                    current[field] = row.get(field)
                    rows = sorted(deduped.values(), key=lambda r: (r.get("drill_order") is None, r.get("drill_order") or 999999))
                return [apply_borehole_wgs84(row) for row in rows]
    except Exception as e:
        raise HTTPException(500, f"Boreholes error: {str(e)}")


@app.post("/boreholes/import_budget")
async def import_budget(request: Request):
    """Import boreholes from the Excel budget file."""
    try:
        payload = await request.json()
    except:
        raise HTTPException(400, "Invalid JSON")
    contractor = payload.get("contractor", "Allianz Drilling")
    boreholes = payload.get("boreholes", [])
    if not boreholes:
        raise HTTPException(400, "No boreholes provided")
    for b in boreholes:
        lat, lng = agd84_amg55_to_wgs84(b.get("easting"), b.get("northing"))
        b["_lat"] = lat if lat is not None else b.get("lat")
        b["_lng"] = lng if lng is not None else b.get("lng")
    with get_conn() as conn:
        with conn.cursor() as cur:
            imported = 0
            merged_by_site = 0
            removed_placeholders = 0
            for b in boreholes:
                incoming_hole_id = str(b.get("hole_id") or "").strip()
                site_id = str(b.get("site_id") or "").strip()
                if not incoming_hole_id:
                    continue
                merge_hole_id = incoming_hole_id
                if site_id:
                    cur.execute("""
                        SELECT hole_id
                        FROM boreholes
                        WHERE contractor=%s AND site_id=%s
                        ORDER BY CASE WHEN hole_id<>%s THEN 0 ELSE 1 END, id
                        LIMIT 1
                    """, (contractor, site_id, incoming_hole_id))
                    existing = cur.fetchone()
                    if existing and existing.get("hole_id"):
                        merge_hole_id = existing["hole_id"]
                if merge_hole_id != incoming_hole_id:
                    merged_by_site += 1
                cur.execute("""
                INSERT INTO boreholes
                (contractor,project,planned_year,site_id,hole_id,drill_order,days_budgeted,
                 bh_type,bit_type,purpose,easting,northing,rl,chip_depth,eoh_depth,total_core,
                 seam_tk,lat,lng,status,drilling_budget_total,earthworks_budget_total,
                 geophysical_budget_total,geological_support_budget_total,misc_budget_total,budget_total)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (contractor,hole_id) DO UPDATE SET
                    project=EXCLUDED.project, planned_year=EXCLUDED.planned_year,
                    site_id=EXCLUDED.site_id, drill_order=EXCLUDED.drill_order,
                    days_budgeted=EXCLUDED.days_budgeted,
                    bh_type=EXCLUDED.bh_type, bit_type=EXCLUDED.bit_type,
                    purpose=EXCLUDED.purpose, easting=EXCLUDED.easting,
                    northing=EXCLUDED.northing, rl=EXCLUDED.rl,
                    chip_depth=EXCLUDED.chip_depth, eoh_depth=EXCLUDED.eoh_depth,
                    total_core=EXCLUDED.total_core, seam_tk=EXCLUDED.seam_tk,
                    lat=EXCLUDED.lat, lng=EXCLUDED.lng,
                    drilling_budget_total=EXCLUDED.drilling_budget_total,
                    earthworks_budget_total=EXCLUDED.earthworks_budget_total,
                    geophysical_budget_total=EXCLUDED.geophysical_budget_total,
                    geological_support_budget_total=EXCLUDED.geological_support_budget_total,
                    misc_budget_total=EXCLUDED.misc_budget_total,
                    budget_total=EXCLUDED.budget_total
                """, (contractor, b.get("project",""), b.get("planned_year",""),
                   site_id, merge_hole_id, b.get("drill_order"),
                   b.get("days") or b.get("days_budgeted"),
                   b.get("type") or b.get("bh_type"), b.get("bit_type"), b.get("purpose"),
                   b.get("easting"), b.get("northing"), b.get("rl"),
                   b.get("chip_depth"), b.get("eoh_depth"), b.get("total_core"),
                   b.get("seam_tk"), b.get("_lat"), b.get("_lng"),
                   b.get("status", "Planned"),
                   b.get("drilling_budget_total"),
                   b.get("earthworks_budget_total"),
                   b.get("geophysical_budget_total"),
                   b.get("geological_support_budget_total"),
                   b.get("misc_budget_total"),
                   b.get("budget_total")))
                imported += 1
                if merge_hole_id != incoming_hole_id and site_id:
                    cur.execute("""
                        DELETE FROM boreholes
                        WHERE contractor=%s AND site_id=%s AND hole_id=%s
                    """, (contractor, site_id, incoming_hole_id))
                    removed_placeholders += cur.rowcount
        conn.commit()
    return {
        "status": "imported",
        "count": imported,
        "merged_by_site": merged_by_site,
        "removed_placeholders": removed_placeholders,
    }


@app.patch("/boreholes/{hole_id}")
async def update_borehole(hole_id: str, request: Request):
    payload = await request.json()
    contractor = payload.pop("contractor", "Allianz Drilling")
    safe = {"status","notes","days_budgeted","drilling_budget_total","earthworks_budget_total","geophysical_budget_total","geological_support_budget_total","misc_budget_total","budget_total","actual_total","drill_order","project","planned_year","site_id","bh_type","bit_type","purpose","easting","northing","rl","chip_depth","eoh_depth","total_core","seam_tk","lat","lng","assigned_rig","scheduled_start","scheduled_end","hole_id"}
    u = {k:v for k,v in payload.items() if k in safe}
    if not u: raise HTTPException(400, "No valid fields")
    params = {**u, "old_hole_id": hole_id, "contractor": contractor}
    set_clause = ",".join(f"{k}=%({k})s" for k in u)
    with get_conn() as conn:
        with conn.cursor() as cur:
            if "easting" in u or "northing" in u:
                cur.execute("SELECT easting,northing FROM boreholes WHERE hole_id=%s AND contractor IN (%s, 'Company') ORDER BY CASE WHEN contractor=%s THEN 0 ELSE 1 END LIMIT 1", (hole_id, contractor, contractor))
                existing = cur.fetchone() or {}
                easting = u.get("easting", existing.get("easting"))
                northing = u.get("northing", existing.get("northing"))
                lat, lng = agd84_amg55_to_wgs84(easting, northing)
                if lat is not None and lng is not None:
                    u["lat"] = lat
                    u["lng"] = lng
                    params["lat"] = lat
                    params["lng"] = lng
                    set_clause = ",".join(f"{k}=%({k})s" for k in u)
            cur.execute(
                f"UPDATE boreholes SET {set_clause} WHERE hole_id=%(old_hole_id)s AND contractor=%(contractor)s",
                params
            )
            if cur.rowcount == 0 and contractor != "Company":
                legacy = {**params, "contractor": "Company"}
                cur.execute(
                    f"UPDATE boreholes SET {set_clause} WHERE hole_id=%(old_hole_id)s AND contractor=%(contractor)s",
                    legacy
                )
            if cur.rowcount == 0:
                raise HTTPException(404, f"Borehole {hole_id} not found for contractor {contractor}")
        conn.commit()
    return {"status": "updated"}


@app.delete("/boreholes/{hole_id}")
def delete_borehole(hole_id: str, contractor: Optional[str] = Query(None)):
    with get_conn() as conn:
        with conn.cursor() as cur:
            if contractor:
                cur.execute("DELETE FROM boreholes WHERE hole_id=%s AND contractor=%s", (hole_id, contractor))
                if cur.rowcount == 0 and contractor != "Company":
                    cur.execute("DELETE FROM boreholes WHERE hole_id=%s AND contractor='Company'", (hole_id,))
            else:
                cur.execute("DELETE FROM boreholes WHERE hole_id=%s", (hole_id,))
        conn.commit()
    return {"status": "deleted"}


@app.get("/boreholes/summary")
def get_borehole_summary(contractor: Optional[str] = Query(None)):
    """Budget vs actual summary for reconciliation."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                where = "WHERE contractor=%s" if contractor else ""
                params = (contractor,) if contractor else ()
                cur.execute(f"""
                    SELECT
                        COUNT(*) AS total_holes,
                        SUM(days_budgeted) AS total_days_budgeted,
                        SUM(budget_total) AS total_budget,
                        COUNT(CASE WHEN status='Complete' THEN 1 END) AS completed,
                        COUNT(CASE WHEN status='In Progress' THEN 1 END) AS in_progress,
                        COUNT(CASE WHEN status='Planned' THEN 1 END) AS planned
                    FROM boreholes {where}
                """, params)
                summary = dict(cur.fetchone())
                cur.execute(f"""
                    SELECT COALESCE(SUM(a.line_cost),0) AS actual_total
                    FROM activities a
                    JOIN boreholes b ON b.hole_id=a.hole_num
                    {("WHERE a.contractor=%s" if contractor else "")}
                """, params)
                summary["actual_total"] = float(cur.fetchone()["actual_total"] or 0)
                return summary
    except Exception as e:
        raise HTTPException(500, f"Summary error: {str(e)}")


# ── Projects ──────────────────────────────────────────────────────────────────

@app.get("/projects")
def get_projects(contractor: str = Query(...)):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM projects WHERE contractor=%s ORDER BY program, name, year", (contractor,))
            return [dict(r) for r in cur.fetchall()]


@app.post("/projects")
async def add_project(request: Request):
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON")
    name = payload.get("name", "").strip()
    if not name:
        raise HTTPException(400, "name is required")
    contractor = payload.get("contractor", "Allianz Drilling")
    program = str(payload.get("program", payload.get("category", "Exploration")) or "Exploration").strip() or "Exploration"
    year = payload.get("year", "")
    notes = payload.get("notes", "")
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO projects (contractor, program, name, year, notes)
                    VALUES (%s, %s, %s, %s, %s)
                    ON CONFLICT (contractor, name) DO UPDATE SET program=EXCLUDED.program, year=EXCLUDED.year, notes=EXCLUDED.notes
                    RETURNING id
                """, (contractor, program, name, year, notes))
                pid = cur.fetchone()["id"]
            conn.commit()
        return {"status": "created", "id": pid, "program": program, "name": name}
    except Exception as e:
        raise HTTPException(500, f"Failed: {str(e)}")


@app.delete("/projects/{project_id}")
def delete_project(project_id: int):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM projects WHERE id=%s", (project_id,))
        conn.commit()
    return {"status": "deleted"}


# ── Gemini Vision OCR for handwritten drill logs ──────────────────────────────

# Project budgets
def clean_project_budget_row(row: dict, contractor_default: str = "Company") -> dict:
    project = str(row.get("project") or "").strip()
    section = str(row.get("section") or row.get("category") or "").strip()
    if not project:
        raise HTTPException(400, "project is required")
    if not section:
        raise HTTPException(400, "section is required")
    try:
        amount = float(str(row.get("budget_amount", row.get("budget", 0)) or 0).replace("$", "").replace(",", ""))
    except Exception:
        amount = 0.0
    return {
        "contractor": str(row.get("contractor") or contractor_default or "Company").strip() or "Company",
        "program": str(row.get("program") or "Exploration").strip() or "Exploration",
        "project": project,
        "section": section,
        "vendor": str(row.get("vendor") or row.get("contractor_vendor") or "").strip(),
        "budget_amount": amount,
        "allocation": str(row.get("allocation") or "Project level").strip() or "Project level",
        "notes": str(row.get("notes") or "").strip(),
    }


@app.get("/project-budgets")
def get_project_budgets(
    contractor: str = Query("Company"),
    project: Optional[str] = Query(None),
    program: Optional[str] = Query(None),
):
    where = ["contractor=%s"]
    params = [contractor]
    if project:
        where.append("project=%s")
        params.append(project)
    if program:
        where.append("program=%s")
        params.append(program)
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT *
                FROM project_budgets
                WHERE {' AND '.join(where)}
                ORDER BY program, project, section, vendor
            """, params)
            return [dict(r) for r in cur.fetchall()]


@app.post("/project-budgets")
async def upsert_project_budgets(request: Request):
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON")
    contractor = str(payload.get("contractor") or "Company").strip() or "Company"
    rows = payload.get("budgets", payload.get("rows", None))
    if rows is None:
        rows = [payload]
    if not isinstance(rows, list) or not rows:
        raise HTTPException(400, "No budget rows provided")
    cleaned = [clean_project_budget_row(dict(r), contractor) for r in rows]
    with get_conn() as conn:
        with conn.cursor() as cur:
            for row in cleaned:
                cur.execute("""
                    INSERT INTO project_budgets
                    (contractor, program, project, section, vendor, budget_amount, allocation, notes, updated_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,NOW())
                    ON CONFLICT (contractor, project, section, vendor)
                    DO UPDATE SET
                        program=EXCLUDED.program,
                        budget_amount=EXCLUDED.budget_amount,
                        allocation=EXCLUDED.allocation,
                        notes=EXCLUDED.notes,
                        updated_at=NOW()
                """, (
                    row["contractor"], row["program"], row["project"], row["section"], row["vendor"],
                    row["budget_amount"], row["allocation"], row["notes"],
                ))
        conn.commit()
    return {"status": "saved", "count": len(cleaned)}


@app.patch("/project-budgets/{budget_id}")
async def update_project_budget(budget_id: int, request: Request):
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON")
    safe = {"contractor", "program", "project", "section", "vendor", "budget_amount", "allocation", "notes"}
    update = {k: payload[k] for k in safe if k in payload}
    if not update:
        return {"status": "unchanged"}
    if "budget_amount" in update:
        try:
            update["budget_amount"] = float(str(update["budget_amount"] or 0).replace("$", "").replace(",", ""))
        except Exception:
            update["budget_amount"] = 0.0
    set_clause = ", ".join([f"{k}=%s" for k in update.keys()] + ["updated_at=NOW()"])
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(f"UPDATE project_budgets SET {set_clause} WHERE id=%s", list(update.values()) + [budget_id])
        conn.commit()
    return {"status": "updated"}


@app.delete("/project-budgets/{budget_id}")
def delete_project_budget(budget_id: int):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM project_budgets WHERE id=%s", (budget_id,))
        conn.commit()
    return {"status": "deleted"}


@app.get("/cost-centre-forecast")
def get_cost_centre_forecast(
    site: str = Query("IB"),
    cost_centre: str = Query("TECEXP"),
    year: int = Query(2026),
):
    """Return the Ironbark three-GL Finance baseline, actuals and remaining plan."""
    from datetime import datetime, timedelta, timezone

    brisbane_today = datetime.now(timezone(timedelta(hours=10))).date()
    forecast_project = "Ironbark" if site.upper() == "IB" else site
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT *
                FROM cost_centre_forecasts
                WHERE site=%s AND cost_centre=%s AND year=%s
                ORDER BY month, category
            """, (site, cost_centre, year))
            rows = [dict(r) for r in cur.fetchall()]

            live_actuals = {}
            live_sources = {}
            contractor_actuals = {}
            contractor_actual_sources = {}

            def add_actual(month, category, amount, source):
                key = f"{int(month)}:{category}"
                live_actuals[key] = round(live_actuals.get(key, 0) + float(amount or 0), 2)
                live_sources.setdefault(key, []).append({"source": source, "amount": round(float(amount or 0), 2)})

            def add_contractor_actual(month, contractor, expense_gl, amount, source):
                contractor_name = str(contractor or "Unassigned").strip() or "Unassigned"
                category = contractor_gl_category(expense_gl)
                key = f"{int(month)}:{contractor_name}"
                contractor_actuals[key] = round(contractor_actuals.get(key, 0) + float(amount or 0), 2)
                contractor_actual_sources.setdefault(key, []).append({
                    "source": source,
                    "expense_gl": str(expense_gl or "").strip(),
                    "amount": round(float(amount or 0), 2),
                })
                add_actual(month, category, amount, source)

            cur.execute("""
                SELECT name, category, program, sites,
                       COALESCE(NULLIF(BTRIM(expense_gl), ''),
                         CASE WHEN category='Drilling' THEN '4350'
                              WHEN category IN ('Earthworks','Labour') THEN '4250'
                              ELSE '4200' END
                       ) AS expense_gl
                FROM contractors
                WHERE COALESCE(active, TRUE)=TRUE
                  AND EXISTS (
                    SELECT 1
                    FROM unnest(string_to_array(COALESCE(program,''), ',')) AS assigned_program
                    WHERE BTRIM(assigned_program)='Exploration'
                  )
                  AND EXISTS (
                    SELECT 1
                    FROM unnest(string_to_array(COALESCE(sites,'Ironbark'), ',')) AS assigned_site
                    WHERE BTRIM(assigned_site)=%s
                  )
                ORDER BY name
            """, (forecast_project,))
            exploration_contractors = [dict(row) for row in cur.fetchall()]

            cur.execute("""
                WITH received_invoices AS (
                  SELECT
                    CASE
                      WHEN i.invoice_date ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$' THEN CAST(SUBSTRING(i.invoice_date,6,2) AS INTEGER)
                      WHEN i.invoice_date ~ '^[0-9]{1,2}/[0-9]{1,2}/[0-9]{4}$' THEN CAST(SPLIT_PART(i.invoice_date,'/',2) AS INTEGER)
                      WHEN i.billing_month ~ '^[0-9]{1,2}/[0-9]{4}$' THEN CAST(SPLIT_PART(i.billing_month,'/',1) AS INTEGER)
                    END AS month,
                    CASE
                      WHEN i.invoice_date ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$' THEN CAST(SUBSTRING(i.invoice_date,1,4) AS INTEGER)
                      WHEN i.invoice_date ~ '^[0-9]{1,2}/[0-9]{1,2}/[0-9]{4}$' THEN CAST(SPLIT_PART(i.invoice_date,'/',3) AS INTEGER)
                      WHEN i.billing_month ~ '^[0-9]{1,2}/[0-9]{4}$' THEN CAST(SPLIT_PART(i.billing_month,'/',2) AS INTEGER)
                    END AS invoice_year,
                    COALESCE(NULLIF(BTRIM(i.contractor), ''), 'Unassigned') AS contractor,
                    COALESCE(
                      NULLIF(BTRIM(c.expense_gl), ''),
                      CASE
                        WHEN COALESCE(c.category,'')='Drilling' THEN '4350'
                        WHEN COALESCE(c.category,'') IN ('Earthworks','Labour') THEN '4250'
                        ELSE '4200'
                      END
                    ) AS expense_gl,
                    COALESCE(i.total_aud,0) AS amount
                  FROM invoices i
                  LEFT JOIN contractors c ON c.name=i.contractor
                  WHERE COALESCE(i.project,'') ILIKE '%%Ironbark%%'
                    AND COALESCE(i.status,'') <> 'Rejected'
                    AND EXISTS (
                      SELECT 1
                      FROM unnest(string_to_array(COALESCE(c.sites,'Ironbark'), ',')) AS assigned_site
                      WHERE BTRIM(assigned_site)=%s
                    )
                )
                SELECT month, contractor, expense_gl, SUM(amount) AS amount
                FROM received_invoices
                WHERE invoice_year=%s
                GROUP BY month, contractor, expense_gl
                ORDER BY month, contractor
            """, (forecast_project, year))
            for row in cur.fetchall():
                if row["month"]:
                    add_contractor_actual(
                        row["month"],
                        row["contractor"],
                        row["expense_gl"],
                        row["amount"],
                        "Invoices Received",
                    )

            # Drilling Activity Reports provide the current incurred cost before
            # an invoice is received. Add only the unbilled balance for each
            # contractor/month so a later invoice replaces, rather than doubles,
            # the operational accrual.
            cur.execute("""
                SELECT CAST(SUBSTRING(a.date,6,2) AS INTEGER) AS month,
                       COALESCE(NULLIF(BTRIM(a.contractor), ''), 'Unassigned') AS contractor,
                       COALESCE(NULLIF(BTRIM(c.expense_gl), ''), '4350') AS expense_gl,
                       SUM(COALESCE(a.line_cost,0)) AS amount
                FROM activities a
                LEFT JOIN contractors c ON c.name=a.contractor
                WHERE a.date ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                  AND SUBSTRING(a.date,1,4)=%s
                  AND COALESCE(a.program,'')='Exploration'
                  AND COALESCE(c.category,'')='Drilling'
                  AND EXISTS (
                    SELECT 1
                    FROM unnest(string_to_array(COALESCE(c.sites,'Ironbark'), ',')) AS assigned_site
                    WHERE BTRIM(assigned_site)=%s
                  )
                  AND (
                    COALESCE(a.project,'') ILIKE '%%Ironbark%%'
                    OR COALESCE(a.hole_num,'') ILIKE 'IB%%'
                    OR COALESCE(a.site_name,'') ILIKE 'IB%%'
                  )
                GROUP BY CAST(SUBSTRING(a.date,6,2) AS INTEGER),
                         COALESCE(NULLIF(BTRIM(a.contractor), ''), 'Unassigned'),
                         COALESCE(NULLIF(BTRIM(c.expense_gl), ''), '4350')
                ORDER BY month, contractor
            """, (str(year), forecast_project))
            for row in cur.fetchall():
                contractor_key = f"{int(row['month'])}:{row['contractor']}"
                received_amount = float(contractor_actuals.get(contractor_key, 0) or 0)
                unbilled_amount = max(float(row["amount"] or 0) - received_amount, 0)
                if unbilled_amount > 0:
                    add_contractor_actual(
                        row["month"],
                        row["contractor"],
                        row["expense_gl"],
                        unbilled_amount,
                        "Activity Reports (unbilled)",
                    )

            # Remaining 4350 is driven by the unfinished Ironbark borehole plan.
            # Actual drilling already recorded against Planned/In Progress holes is
            # deducted so that it is not counted once in July and again in the plan.
            cur.execute("""
                SELECT b.hole_id, b.site_id, b.status, b.scheduled_start,
                       b.drill_order, b.days_budgeted, b.assigned_rig,
                       b.drilling_budget_total,
                       COALESCE(SUM(CASE WHEN COALESCE(ac.category,'')='Drilling'
                           THEN COALESCE(a.line_cost,0) ELSE 0 END),0) AS drilling_actual
                FROM boreholes b
                LEFT JOIN activities a ON
                    a.date ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                    AND SUBSTRING(a.date,1,4)=%s
                    AND (
                        a.hole_num=b.hole_id
                        OR (COALESCE(b.site_id,'')<>'' AND a.site_name=b.site_id)
                    )
                LEFT JOIN contractors ac ON ac.name=a.contractor
                WHERE b.contractor='Company'
                  AND COALESCE(b.project,'')='Ironbark'
                  AND COALESCE(b.planned_year,'')=%s
                  AND LOWER(COALESCE(b.status,'Planned')) NOT IN ('complete','cancelled')
                GROUP BY b.id
                ORDER BY b.drill_order NULLS LAST, b.hole_id
            """, (str(year), str(year)))
            plan_holes = [dict(row) for row in cur.fetchall()]

    plan_rows = []
    for hole in plan_holes:
        budget = float(hole.get("drilling_budget_total") or 0)
        actual = float(hole.get("drilling_actual") or 0)
        remaining = max(budget - actual, 0) if budget > 0 else 0
        deducted_actual = min(actual, budget) if budget > 0 else 0
        plan_rows.append({
            "hole_id": hole.get("hole_id"),
            "site_id": hole.get("site_id"),
            "status": hole.get("status") or "Planned",
            "scheduled_start": hole.get("scheduled_start"),
            "drill_order": hole.get("drill_order"),
            "days_budgeted": float(hole.get("days_budgeted") or 0),
            "assigned_rig": hole.get("assigned_rig") or "Unassigned",
            "drilling_budget": round(budget, 2),
            "drilling_actual": round(actual, 2),
            "deducted_actual": round(deducted_actual, 2),
            "remaining_drilling": round(remaining, 2),
        })

    # Schedule open holes in Borehole Planning order. Each assigned rig has a
    # separate cursor, allowing rigs to operate in parallel. A hole's remaining
    # budget is spread evenly over its remaining drill days and calendar months.
    schedule_base = brisbane_today + timedelta(days=1)
    rig_cursors = {}
    plan_forecast_cents = {}

    def parse_plan_date(raw_value):
        raw = str(raw_value or "").strip()
        for date_format in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(raw[:10], date_format).date()
            except (TypeError, ValueError):
                continue
        return None

    for hole in plan_rows:
        remaining = float(hole["remaining_drilling"] or 0)
        rig = str(hole.get("assigned_rig") or "Unassigned").strip() or "Unassigned"
        stored_start = parse_plan_date(hole.get("scheduled_start"))
        start_date = max(schedule_base, rig_cursors.get(rig, schedule_base), stored_start or schedule_base)
        budget_days = max(1, int(ceil(float(hole.get("days_budgeted") or 1))))
        budget = float(hole.get("drilling_budget") or 0)
        if budget > 0 and remaining < budget:
            remaining_days = max(1, int(ceil(budget_days * remaining / budget)))
        else:
            remaining_days = budget_days
        end_date = start_date + timedelta(days=remaining_days - 1)
        rig_cursors[rig] = end_date + timedelta(days=1)
        hole["forecast_start"] = start_date.isoformat()
        hole["forecast_end"] = end_date.isoformat()
        hole["forecast_drill_days"] = remaining_days

        total_cents = int(round(remaining * 100))
        cents_per_day, extra_cents = divmod(total_cents, remaining_days)
        for day_index in range(remaining_days):
            spend_date = start_date + timedelta(days=day_index)
            period = spend_date.strftime("%Y-%m")
            day_cents = cents_per_day + (1 if day_index < extra_cents else 0)
            plan_forecast_cents[period] = plan_forecast_cents.get(period, 0) + day_cents

    plan_forecast_periods = {
        f"{period}:4350 - Drilling Services": round(cents / 100, 2)
        for period, cents in sorted(plan_forecast_cents.items())
    }
    plan_forecast = {
        f"{month}:4350 - Drilling Services": round(plan_forecast_cents.get(f"{year}-{month:02d}", 0) / 100, 2)
        for month in range(1, 13)
    }

    remaining_total = round(sum(row["remaining_drilling"] for row in plan_rows), 2)
    plan_summary = {
        "project": "Ironbark",
        "hole_count": len(plan_rows),
        "planned_holes": sum(1 for row in plan_rows if str(row["status"]).lower() == "planned"),
        "in_progress_holes": sum(1 for row in plan_rows if str(row["status"]).lower() == "in progress"),
        "budgeted_holes": sum(1 for row in plan_rows if row["drilling_budget"] > 0),
        "missing_budget_holes": sum(1 for row in plan_rows if row["drilling_budget"] <= 0),
        "gross_drilling_budget": round(sum(row["drilling_budget"] for row in plan_rows), 2),
        "drilling_actual_on_open_holes": round(sum(row["drilling_actual"] for row in plan_rows), 2),
        "drilling_actual_deducted": round(sum(row["deducted_actual"] for row in plan_rows), 2),
        "remaining_drilling_budget": remaining_total,
        "forecast_in_year": round(sum(plan_forecast.values()), 2),
        "forecast_after_year": round(remaining_total - sum(plan_forecast.values()), 2),
        "scheduled_holes": len(plan_rows),
        "unphased_holes": 0,
        "missing_days_holes": sum(1 for row in plan_rows if float(row.get("days_budgeted") or 0) <= 0),
        "first_forecast_start": min((row["forecast_start"] for row in plan_rows), default=None),
        "latest_forecast_end": max((row["forecast_end"] for row in plan_rows), default=None),
        "allocation_method": "Remaining drilling budgets are scheduled by Borehole Planning drill order and assigned rig from the day after the as-of date. Each hole occupies its remaining budgeted drill days, and spend is spread evenly across those calendar days and months.",
    }

    return {
        "site": site,
        "division": rows[0]["division"] if rows else "TEC",
        "cost_centre": cost_centre,
        "program": rows[0]["program"] if rows else "Exploration",
        "year": year,
        "as_of": brisbane_today.isoformat(),
        "rows": rows,
        "live_actuals": live_actuals,
        "live_sources": live_sources,
        "exploration_contractors": exploration_contractors,
        "contractor_actuals": contractor_actuals,
        "contractor_actual_sources": contractor_actual_sources,
        "plan_forecast": plan_forecast,
        "plan_forecast_periods": plan_forecast_periods,
        "plan_summary": plan_summary,
        "plan_holes": plan_rows,
    }


@app.patch("/cost-centre-forecast/{row_id}")
async def update_cost_centre_forecast(row_id: int, request: Request):
    payload = await request.json()
    safe = {"manual_accrual", "forecast_override", "notes"}
    update = {key: payload[key] for key in safe if key in payload}
    if not update:
        raise HTTPException(400, "No valid fields")
    for field in ("manual_accrual", "forecast_override"):
        if field in update and update[field] not in (None, ""):
            try:
                update[field] = float(str(update[field]).replace("$", "").replace(",", ""))
            except Exception:
                raise HTTPException(400, f"{field} must be numeric")
        elif field in update:
            update[field] = None if field == "forecast_override" else 0
    set_clause = ", ".join([f"{key}=%s" for key in update] + ["updated_at=NOW()"])
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE cost_centre_forecasts SET {set_clause} WHERE id=%s RETURNING *",
                list(update.values()) + [row_id],
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Forecast row not found")
        conn.commit()
    return dict(row)


# Gemini Vision OCR
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")

GEMINI_PROMPT = """You are reading a scanned handwritten drilling End-of-Shift report (DEPCO Drill Log format).

Extract ALL data from this image into a JSON object with these exact fields:

{
  "log_number": "the No. at top right e.g. 099839",
  "client": "client name from Client field",
  "location": "Location/Area field",
  "hole_num": "Hole # field e.g. CR57",
  "date": "Date field in d/m/yyyy format",
  "day": "Day of week",
  "shift": "Day or Night",
  "start_time": "Start Time e.g. 0600",
  "finish_time": "Finish Time e.g. 1730",
  "travel_hours": "Travel Hours if filled",
  "driller": "Driller name",
  "offsider1": "Offsider 1 name",
  "offsider2": "Offsider 2 name",
  "rig_no": "Rig No.",
  "drill_rig": "Rod Trailer No. next to Rig No.",
  "activities": [
    {
      "comments": "handwritten description of the activity",
      "time_from": "HH:MM format e.g. 05:30",
      "time_to": "HH:MM format e.g. 06:00",
      "total_time": "duration in H:MM or fraction e.g. 0:30 or 0.5",
      "metres_from": null or number,
      "metres_to": null or number,
      "total_metres": null or number
    }
  ]
}

IMPORTANT RULES:
- Read the DRILLING section table carefully - each row has: COMMENTS, TIME FROM, TIME TO, TOTAL TIME, METERS FROM, METERS TO, TOTAL METERS
- Times are in 24hr format (e.g. 0530 means 05:30, 1415 means 14:15)
- Total Time is usually in fractions like 2, 1/2, 3/4 etc. Convert: 1/2=0:30, 1/4=0:15, 3/4=0:45
- The handwriting may be difficult - do your best to read it
- If a field is empty or unreadable, use null
- Return ONLY valid JSON, no markdown, no explanation
"""

GEMINI_OCR_RESPONSE_SCHEMA = {
    "type": "object",
    "properties": {
        "log_number": {"type": ["string", "null"]},
        "client": {"type": ["string", "null"]},
        "location": {"type": ["string", "null"]},
        "hole_num": {"type": ["string", "null"]},
        "date": {"type": ["string", "null"]},
        "day": {"type": ["string", "null"]},
        "shift": {"type": ["string", "null"]},
        "start_time": {"type": ["string", "null"]},
        "finish_time": {"type": ["string", "null"]},
        "travel_hours": {"type": ["string", "number", "null"]},
        "driller": {"type": ["string", "null"]},
        "offsider1": {"type": ["string", "null"]},
        "offsider2": {"type": ["string", "null"]},
        "rig_no": {"type": ["string", "null"]},
        "drill_rig": {"type": ["string", "null"]},
        "activities": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "comments": {"type": ["string", "null"]},
                    "time_from": {"type": ["string", "null"]},
                    "time_to": {"type": ["string", "null"]},
                    "total_time": {"type": ["string", "number", "null"]},
                    "metres_from": {"type": ["number", "null"]},
                    "metres_to": {"type": ["number", "null"]},
                    "total_metres": {"type": ["number", "null"]},
                },
                "required": [
                    "comments", "time_from", "time_to", "total_time",
                    "metres_from", "metres_to", "total_metres",
                ],
                "additionalProperties": False,
            },
        },
    },
    "required": [
        "log_number", "client", "location", "hole_num", "date", "day",
        "shift", "start_time", "finish_time", "travel_hours", "driller",
        "offsider1", "offsider2", "rig_no", "drill_rig", "activities",
    ],
    "additionalProperties": False,
}


async def ocr_with_gemini(pdf_bytes: bytes) -> dict:
    """Send a PDF page image to Gemini Vision and extract structured data."""
    if not GEMINI_API_KEY:
        raise HTTPException(500, "GEMINI_API_KEY not configured on server")

    # Convert PDF to image using pdfplumber
    from PIL import Image
    import io

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        if not pdf.pages:
            raise HTTPException(400, "PDF has no pages")
        page = pdf.pages[0]
        # Render page to image
        img = page.to_image(resolution=300)
        img_buffer = io.BytesIO()
        img.original.save(img_buffer, format='PNG')
        img_bytes = img_buffer.getvalue()

    b64_image = base64.b64encode(img_bytes).decode('utf-8')

    # Call Gemini API
    url = gemini_generate_content_url(GEMINI_API_KEY)

    payload = {
        "contents": [{
            "parts": [
                {"text": GEMINI_PROMPT},
                {
                    "inline_data": {
                        "mime_type": "image/png",
                        "data": b64_image
                    }
                }
            ]
        }],
        "generationConfig": {
            "temperature": 0.1,
            # Gemini 3.5 Flash can spend much of this allowance on thinking.
            # OCR is structured extraction, so keep thinking low and allow the
            # model's full output capacity for the JSON response.
            "maxOutputTokens": 65536,
            "thinkingConfig": {
                "thinkingLevel": "low",
            },
            "responseFormat": {
                "text": {
                    "mimeType": "APPLICATION_JSON",
                    "schema": GEMINI_OCR_RESPONSE_SCHEMA,
                }
            },
        }
    }

    async with httpx.AsyncClient(timeout=90.0) as client:
        for attempt in range(2):
            resp = await client.post(url, json=payload)
            if resp.status_code != 200:
                raise HTTPException(502, f"Gemini API error: {resp.status_code} - {resp.text[:200]}")

            result = resp.json()
            try:
                text = gemini_response_text(result)
                # Clean up - remove markdown fences if present
                text = text.strip()
                if text.startswith("```"):
                    text = text.split("\n", 1)[1] if "\n" in text else text[3:]
                if text.endswith("```"):
                    text = text[:-3]
                text = text.strip()
                if text.startswith("json"):
                    text = text[4:].strip()
                return json.loads(text)
            except (KeyError, IndexError, json.JSONDecodeError) as e:
                finish_reason = result.get("candidates", [{}])[0].get("finishReason", "unknown")
                if finish_reason == "MAX_TOKENS" and attempt == 0:
                    continue
                raise HTTPException(
                    422,
                    f"Could not parse Gemini response ({finish_reason}): {str(e)}",
                )

    raise HTTPException(422, "Gemini OCR did not return a complete response")


@app.post("/import/ocr")
async def import_ocr_pdf(
    file: UploadFile = File(...),
    contractor: str = Form(default="DEPCO Drilling"),
    program: str = Form(default=""),
    ocr_data: Optional[str] = Form(default=None),
):
    """Import a handwritten drill log PDF using reviewed or fresh Gemini OCR data."""
    filename = file.filename
    content = await file.read()

    # Check if already imported
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT 1 FROM imported_files WHERE filename=%s AND contractor=%s",
                        (filename, contractor))
            if cur.fetchone():
                return {"status": "skipped", "filename": filename}

    if ocr_data is not None:
        if len(ocr_data) > 1_000_000:
            raise HTTPException(400, "Reviewed OCR data is too large")
        try:
            data = json.loads(ocr_data)
        except json.JSONDecodeError as e:
            raise HTTPException(400, f"Reviewed OCR data is invalid: {str(e)}")
        if not isinstance(data, dict) or not isinstance(data.get("activities"), list):
            raise HTTPException(400, "Reviewed OCR data must contain an activities list")
        if len(data["activities"]) > 500:
            raise HTTPException(400, "Reviewed OCR data contains too many activities")
        if any(not isinstance(activity, dict) for activity in data["activities"]):
            raise HTTPException(400, "Each reviewed OCR activity must be an object")
    else:
        # Keep direct API compatibility, but the UI normally previews and reviews first.
        try:
            data = await ocr_with_gemini(content)
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(422, f"OCR failed: {str(e)}")

    # Convert to activity rows
    activities = data.get("activities", [])
    hole_num = data.get("hole_num", "")
    date_str = data.get("date", "")
    location = data.get("location", "")
    driller = data.get("driller", "")
    shift = data.get("shift", "Day")
    site_name = location

    rows = []
    for act in activities:
        time_from = act.get("time_from", "")
        time_to = act.get("time_to", "")
        total_time = act.get("total_time", "")

        # Normalise total_time to H:MM
        if total_time and isinstance(total_time, (int, float)):
            h = int(total_time)
            m = int((total_time - h) * 60)
            total_time = f"{h}:{m:02d}"

        metres_from = act.get("metres_from")
        metres_to = act.get("metres_to")
        total_metres = act.get("total_metres")
        comments = act.get("comments", "")

        rows.append({
            "source_file": filename, "contractor": contractor,
            "date": date_str, "hole_num": hole_num,
            "site_name": site_name, "location": location,
            "drill_rig": data.get("rig_no", ""), "client": data.get("client", ""),
            "contract": "", "shift": shift,
            "time_from": time_from, "time_to": time_to,
            "total_time": total_time,
            "bit_type": "", "diameter": "",
            "metres_from": float(metres_from) if metres_from else None,
            "metres_to": float(metres_to) if metres_to else None,
            "total_metres": float(total_metres) if total_metres else None,
            "code": "", "notes": comments,
            "rate_year": None, "unit_rate": None, "quantity": None,
            "line_cost": None, "rate_basis": None, "po_id": None,
        })

    rows = apply_import_activity_scope(rows, contractor, program)

    # Save to database
    with get_conn() as conn:
        with conn.cursor() as cur:
            if rows:
                psycopg2.extras.execute_batch(cur, """
                    INSERT INTO activities
                    (source_file,contractor,date,hole_num,site_name,program,project,location,drill_rig,
                     client,contract,shift,time_from,time_to,total_time,bit_type,diameter,
                     metres_from,metres_to,total_metres,code,notes,
                     rate_year,unit_rate,quantity,line_cost,rate_basis,po_id)
                    VALUES
                    (%(source_file)s,%(contractor)s,%(date)s,%(hole_num)s,%(site_name)s,%(program)s,%(project)s,
                     %(location)s,%(drill_rig)s,%(client)s,%(contract)s,%(shift)s,
                     %(time_from)s,%(time_to)s,%(total_time)s,%(bit_type)s,%(diameter)s,
                     %(metres_from)s,%(metres_to)s,%(total_metres)s,%(code)s,%(notes)s,
                     %(rate_year)s,%(unit_rate)s,%(quantity)s,%(line_cost)s,%(rate_basis)s,%(po_id)s)
                """, rows)
            cur.execute("INSERT INTO imported_files (filename,contractor) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                        (filename, contractor))
            cur.execute("""
                INSERT INTO source_files (filename, contractor, file_type, pdf_data)
                VALUES (%s, %s, 'ocr', %s) ON CONFLICT (filename, contractor) DO NOTHING
            """, (filename, contractor, psycopg2.Binary(content)))
        conn.commit()

    return {
        "status": "imported",
        "filename": filename,
        "ocr_data": data,
        "rows": len(rows),
        "hole_num": hole_num,
        "date": date_str,
        "contractor": contractor,
    }


@app.post("/import/ocr/preview")
async def preview_ocr_pdf(
    file: UploadFile = File(...),
):
    """Preview OCR results without saving — for testing."""
    content = await file.read()
    data = await ocr_with_gemini(content)
    return {"ocr_data": data, "activity_count": len(data.get("activities", []))}


@app.get("/source_files/{filename}")
def get_source_file(filename: str, contractor: Optional[str] = Query(None)):
    """Return a stored source PDF for viewing."""
    from fastapi.responses import Response
    with get_conn() as conn:
        with conn.cursor() as cur:
            if contractor:
                cur.execute("SELECT pdf_data FROM source_files WHERE filename=%s AND contractor=%s", (filename, contractor))
            else:
                cur.execute("SELECT pdf_data FROM source_files WHERE filename=%s LIMIT 1", (filename,))
            row = cur.fetchone()
            if not row or not row["pdf_data"]:
                raise HTTPException(404, "Source file not found")
            return Response(
                content=bytes(row["pdf_data"]),
                media_type="application/pdf",
                headers={"Content-Disposition": f'inline; filename="{filename}"'}
            )

@app.delete("/reset")
def reset_db(contractor: Optional[str]=Query(None)):
    with get_conn() as conn:
        with conn.cursor() as cur:
            if contractor:
                for tbl in ("activities","consumables","crew","imported_files"):
                    cur.execute(f"DELETE FROM {tbl} WHERE contractor=%s",(contractor,))
            else:
                for tbl in ("activities","consumables","crew","imported_files","purchase_orders"):
                    cur.execute(f"DELETE FROM {tbl}")
        conn.commit()
    return {"status":"cleared","contractor":contractor or "all"}
