"""MCC weekly timesheet parsing and daily-to-weekly reconciliation."""

import re
from collections import defaultdict
from io import BytesIO

from mcc_rates import MCC_SCHEDULE_DATE, mcc_schedule_match


MCC_WEEKLY_REQUIRED_HEADERS = {
    "Created by",
    "Shift Time Start",
    "Site Location",
    "Job Description - Role",
    "Job Description - Location",
    "Job Description - Hours",
    "Job Description - Description of Work Performed",
}

MCC_WORKSTREAM_LABELS = {
    "ARG-002": "ARG-002 - Gas Riser Civil Works",
    "ARG-003": "ARG-003 - SIS Drill Civil Works",
    "ARG-004": "ARG-004 - SIS Drill Watercart Works",
    "ARG-005": "ARG-005 - Exploration Civils & Support Works",
}


def _excel_dt(value):
    if value is None:
        return "", ""
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y"), value.strftime("%H:%M")
    return str(value), ""


def _decimal_hours_to_time(hours):
    try:
        total = int(round(float(hours) * 60))
    except Exception:
        return ""
    return f"{total // 60}:{total % 60:02d}"


def mcc_program_from_location(text):
    value = (text or "").lower()
    if "arg-002" in value or "gas riser" in value:
        return "Gas Riser"
    if "arg-003" in value or "arg-004" in value or "sis" in value:
        return "SIS"
    if "arg-005" in value or "exploration" in value:
        return "Exploration"
    return ""


def mcc_workstream_label(text):
    """Return one stable workstream label despite source whitespace variants."""
    normalised = re.sub(r"\s+", " ", str(text or "")).strip()
    match = re.search(r"\bARG\s*-\s*(00[2-5])\b", normalised, re.I)
    if not match:
        return normalised
    return MCC_WORKSTREAM_LABELS.get(f"ARG-{match.group(1)}", normalised)


def mcc_hole_from_text(text):
    value = text or ""
    patterns = [
        r"\bIB[-\s]?(\d{2})[-\s]?(\d{3})\b",
        r"\bIB\s?(\d{2})[-\s]?(\d{2})\b",
        r"\bGR[-\s]?(\d{1,2})\b",
        r"\bSISMG\d{2}[-\s]?\d{2}[A-Z0-9]*\b",
        r"\bMG\d{2}[-\s]?\d{2}[A-Z0-9]*\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, value, re.I)
        if not match:
            continue
        raw = match.group(0).upper().replace(" ", "-")
        if raw.startswith("IB") and len(match.groups()) >= 2:
            second = match.group(2)
            if len(second) == 2:
                second = second.zfill(3)
            return f"IB-{match.group(1)}-{second}"
        return raw.replace("--", "-")
    return ""


def parse_mcc_weekly_xlsx(content, filename, contractor="MCC Group"):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ValueError(f"openpyxl is not available: {exc}") from exc

    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    activities, crew, seen, source_lines = [], [], set(), []
    header = {"date": "", "hole_num": "", "site_name": "", "contractor": contractor}

    eligible_sheets = []
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(value or "").strip() for value in rows[0]]
        if not MCC_WEEKLY_REQUIRED_HEADERS.issubset(set(headers)):
            continue
        eligible_sheets.append((worksheet.title, rows, headers))

    # Older workbooks contain an all-work summary plus invoice workstream tabs.
    # The workstream tabs are the reconciled ARG detail and can contain corrections
    # that did not flow back to the summary. Prefer them to avoid duplicates and
    # exclude unrelated FITZ work from the Argo invoice basis.
    workstream_sheets = [
        item for item in eligible_sheets
        if re.match(r"^ARG[-\s]", item[0], re.I)
    ]
    sheets_to_parse = workstream_sheets or eligible_sheets

    for _sheet_name, rows, headers in sheets_to_parse:
        index = {name: position for position, name in enumerate(headers)}

        for raw in rows[1:]:
            def cell(name):
                position = index.get(name, -1)
                return raw[position] if 0 <= position < len(raw) else None

            created_by = cell("Created by")
            start = cell("Shift Time Start")
            end = cell("Shift Time End")
            site = cell("Site Location")
            role = cell("Job Description - Role")
            location = mcc_workstream_label(cell("Job Description - Location"))
            hours = cell("Job Description - Hours")
            description = cell("Job Description - Description of Work Performed")
            equipment = cell("Job Description - Equipment")
            smu_start = cell("Job Description - SMU Start")
            smu_finish = cell("Job Description - SMU Finish")
            smu_total = cell("Job Description - SMU Total")
            if not any([created_by, start, end, site, role, location, hours, description, equipment]):
                continue

            report_date, time_from = _excel_dt(start)
            _, time_to = _excel_dt(end)
            program = mcc_program_from_location(location)
            hole = mcc_hole_from_text(" ".join([str(description or ""), str(location or "")]))
            source_key = tuple(str(value or "") for value in (
                report_date, time_from, time_to, created_by, role, location,
                hours, description, equipment, smu_start, smu_finish, smu_total,
            ))
            if source_key in seen:
                continue
            seen.add(source_key)

            notes_parts = [
                str(description or "").strip(),
                f"Role: {role}" if role else "",
                f"Program: {program}" if program else "",
                f"Workstream: {location}" if location else "",
                f"Equipment: {equipment}" if equipment else "",
                f"SMU: {smu_start} to {smu_finish} ({smu_total})"
                if equipment and (smu_start is not None or smu_finish is not None or smu_total is not None) else "",
                f"Created by: {created_by}" if created_by else "",
            ]
            try:
                raw_hours = hours
                if raw_hours in (None, "") and equipment and smu_total not in (None, ""):
                    raw_hours = smu_total
                hours_value = float(raw_hours) if raw_hours not in (None, "") else None
            except Exception:
                hours_value = None

            base_row = {
                "source_file": filename,
                "contractor": contractor,
                "date": report_date,
                "hole_num": hole,
                "site_name": str(site or "").strip() or hole,
                "program": program,
                "project": str(location or "").strip() or program,
                "location": str(location or "").strip(),
                "drill_rig": str(equipment or "").strip(),
                "client": "ARGO",
                "contract": str(location or "").strip(),
                "shift": "",
                "time_from": time_from,
                "time_to": time_to,
                "total_time": _decimal_hours_to_time(hours),
                "bit_type": "",
                "diameter": "",
                "metres_from": None,
                "metres_to": None,
                "total_metres": None,
                "rate_year": report_date[-4:] if len(report_date) >= 4 else "",
                "po_id": None,
            }
            priced_lines = []
            labour_rate = mcc_schedule_match(role, "labour")
            equipment_rate = mcc_schedule_match(equipment, "equipment")
            if labour_rate:
                priced_lines.append(("Labour", labour_rate))
            if equipment_rate:
                priced_lines.append(("Equipment", equipment_rate))

            if not priced_lines:
                activities.append({
                    **base_row,
                    "code": "H_Active",
                    "notes": " | ".join(part for part in notes_parts if part),
                    "unit_rate": None,
                    "quantity": hours_value,
                    "line_cost": None,
                    "rate_basis": "MCC weekly EOS import - unpriced",
                })
            else:
                for charge_type, rate in priced_lines:
                    quantity = 1 if rate["unit"] == "day" else hours_value
                    line_cost = round(float(quantity or 0) * float(rate["rate"]), 2) if quantity is not None else None
                    activities.append({
                        **base_row,
                        "code": rate["code"],
                        "notes": " | ".join(
                            part for part in notes_parts + [
                                f"Charge type: {charge_type}",
                                f"Schedule item: {rate['description']}",
                            ] if part
                        ),
                        "unit_rate": rate["rate"],
                        "quantity": quantity,
                        "line_cost": line_cost,
                        "rate_basis": (
                            f"MCC schedule {MCC_SCHEDULE_DATE} - {rate['description']} "
                            f"({rate['unit']}); excludes accommodation and diesel"
                        ),
                    })

            if created_by:
                crew.append({
                    "source_file": filename,
                    "contractor": contractor,
                    "date": report_date,
                    "hole_num": hole,
                    "site_name": base_row["site_name"],
                    "role": str(role or ""),
                    "name": str(created_by or ""),
                    "hours": "" if hours_value is None else str(hours_value),
                })
            if not header["date"] and report_date:
                header.update({
                    "date": report_date,
                    "site_name": base_row["site_name"],
                    "hole_num": hole,
                    "contractor": contractor,
                })
            source_lines.append(f"{report_date} {created_by or ''} {location or ''} {description or ''}")

    if not activities:
        raise ValueError("No MCC weekly timesheet rows found in workbook")
    return header, activities, [], crew, "\n".join(source_lines[:500])


def build_mcc_daily_weekly_audit(rows):
    """Compare operational daily quantities with authoritative weekly quantities."""
    grouped = defaultdict(lambda: {
        "daily_quantity": 0.0,
        "weekly_quantity": 0.0,
        "daily_cost": 0.0,
        "weekly_cost": 0.0,
        "daily_rows": 0,
        "weekly_rows": 0,
        "description": "",
        "unit": "",
    })
    weekly_files = set()
    daily_files = set()

    for row in rows or []:
        source_type = str(row.get("source_file_type") or "")
        if source_type not in {"mcc_site_services_pdf", "mcc_weekly_xlsx"}:
            continue
        code = str(row.get("code") or "Unpriced").strip() or "Unpriced"
        key = (str(row.get("date") or ""), code)
        item = grouped[key]
        item["description"] = str(row.get("rate_description") or item["description"] or code)
        item["unit"] = str(row.get("rate_unit") or item["unit"] or "")
        quantity = float(row.get("quantity") or 0)
        cost = float(row.get("line_cost") or 0)
        filename = str(row.get("source_file") or "")
        if source_type == "mcc_weekly_xlsx":
            item["weekly_quantity"] += quantity
            item["weekly_cost"] += cost
            item["weekly_rows"] += 1
            if filename:
                weekly_files.add(filename)
        else:
            item["daily_quantity"] += quantity
            item["daily_cost"] += cost
            item["daily_rows"] += 1
            if filename:
                daily_files.add(filename)

    comparison = []
    for (report_date, code), item in sorted(grouped.items()):
        quantity_variance = round(item["weekly_quantity"] - item["daily_quantity"], 2)
        tolerance = 0.01 if item["unit"] == "day" else 0.25
        if not item["weekly_rows"]:
            status = "missing_weekly"
        elif not item["daily_rows"]:
            status = "missing_daily"
        elif abs(quantity_variance) <= tolerance:
            status = "match"
        else:
            status = "variance"
        comparison.append({
            "date": report_date,
            "code": code,
            "description": item["description"],
            "unit": item["unit"],
            "daily_quantity": round(item["daily_quantity"], 2),
            "weekly_quantity": round(item["weekly_quantity"], 2),
            "quantity_variance": quantity_variance,
            "daily_estimate": round(item["daily_cost"], 2),
            "weekly_cost": round(item["weekly_cost"], 2),
            "cost_variance": round(item["weekly_cost"] - item["daily_cost"], 2),
            "status": status,
            "daily_rows": item["daily_rows"],
            "weekly_rows": item["weekly_rows"],
        })

    totals = {
        "daily_estimate": round(sum(item["daily_cost"] for item in grouped.values()), 2),
        "weekly_cost": round(sum(item["weekly_cost"] for item in grouped.values()), 2),
        "cost_variance": round(sum(item["weekly_cost"] - item["daily_cost"] for item in grouped.values()), 2),
        "matched": sum(item["status"] == "match" for item in comparison),
        "exceptions": sum(item["status"] != "match" for item in comparison),
        "daily_files": len(daily_files),
        "weekly_files": len(weekly_files),
    }
    return {"totals": totals, "comparison": comparison}
