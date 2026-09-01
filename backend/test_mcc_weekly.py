import unittest
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook

from mcc_weekly import build_mcc_daily_weekly_audit, parse_mcc_weekly_xlsx


HEADERS_WITHOUT_SHIFT_END = [
    "Created by",
    "Shift Time Start",
    "Site Location",
    "Job Description - Role",
    "Job Description - Location",
    "Job Description - Hours",
    "Job Description - Description of Work Performed",
    "Job Description - Equipment",
    "Job Description - SMU Start",
    "Job Description - SMU Finish",
    "Job Description - SMU Total",
]


def workbook_bytes(rows, duplicate_sheet=False):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Weekly"
    sheet.append(HEADERS_WITHOUT_SHIFT_END)
    for row in rows:
        sheet.append(row)
    if duplicate_sheet:
        copy = workbook.create_sheet("Workstream")
        copy.append(HEADERS_WITHOUT_SHIFT_END)
        for row in rows:
            copy.append(row)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


class MCCWeeklyTests(unittest.TestCase):
    def test_imports_new_format_without_shift_end(self):
        content = workbook_bytes([[
            "Kristy Faram",
            datetime(2026, 7, 20, 5, 30),
            "Ironbark 1",
            "Multi Skilled Operator",
            "ARG-004 - SIS Drill Watercart Works",
            2.5,
            "Water MG08-02L sumps",
            "WT01 - Hino FM500 13kL Watercart",
            100,
            102,
            2,
        ]])

        header, activities, _, crew, _ = parse_mcc_weekly_xlsx(
            content, "20260726-MCC-ARG-TimeSheetReport.xlsx"
        )

        self.assertEqual(header["date"], "20/07/2026")
        self.assertEqual(len(activities), 3)
        self.assertEqual({row["program"] for row in activities}, {"SIS"})
        self.assertEqual({row["code"] for row in activities}, {
            "MCC_MULTI_SKILLED_OPERATOR",
            "MCC_BODY_WATER_TRUCK",
            "MCC_LIGHT_VEHICLE",
        })
        self.assertEqual(sum(row["line_cost"] for row in activities), 515.0)
        self.assertTrue(all(row["time_to"] == "" for row in activities))
        self.assertEqual(len(crew), 1)

    def test_canonicalises_workstream_whitespace_by_arg_code(self):
        content = workbook_bytes([[
            "Kristy Faram",
            datetime(2026, 7, 20, 5, 30),
            "Ironbark 1",
            "Multi Skilled Operator",
            "ARG-005 -   Exploration Civils &   Support Works",
            2,
            "Rehabilitate IB-26-018",
            None,
            None,
            None,
            None,
        ]])

        _, activities, _, _, _ = parse_mcc_weekly_xlsx(content, "weekly.xlsx")

        self.assertEqual(activities[0]["program"], "Exploration")
        self.assertEqual(activities[0]["location"], "ARG-005 - Exploration Civils & Support Works")
        self.assertEqual(activities[0]["project"], "ARG-005 - Exploration Civils & Support Works")
        self.assertIn("Workstream: ARG-005 - Exploration Civils & Support Works", activities[0]["notes"])

    def test_deduplicates_summary_and_workstream_sheets(self):
        row = [
            "Michael Chilby",
            datetime(2026, 7, 21, 5, 30),
            "Ironbark 1",
            "Supervisor",
            "ARG-002 - Gas Riser Civil Works",
            4,
            "Empty GR21 sumps",
            "MCC39 - Vac Truck - PER DAY",
            0,
            1,
            1,
        ]
        _, activities, _, crew, _ = parse_mcc_weekly_xlsx(
            workbook_bytes([row], duplicate_sheet=True), "weekly.xlsx"
        )

        self.assertEqual(len(activities), 3)
        self.assertEqual(len(crew), 1)
        self.assertEqual(sum(item["line_cost"] for item in activities), 1395.0)

    def test_prefers_arg_workstream_tabs_over_stale_summary(self):
        summary_row = [
            "Adam Vine", datetime(2026, 6, 22, 6), "Ironbark 1", "Supervisor",
            "ARG-003 - SIS Drill Civil Works", 8, "Provide supervision", None,
            None, None, None,
        ]
        corrected_row = list(summary_row)
        corrected_row[5] = 12.5
        corrected_row[6] = "Supervise spotter catcher works"
        workbook = Workbook()
        summary = workbook.active
        summary.title = "20260621-ARG-WeeklySheets"
        summary.append(HEADERS_WITHOUT_SHIFT_END)
        summary.append(summary_row)
        workstream = workbook.create_sheet("ARG-003 SIS Civils")
        workstream.append(HEADERS_WITHOUT_SHIFT_END)
        workstream.append(corrected_row)
        stream = BytesIO()
        workbook.save(stream)

        _, activities, _, _, _ = parse_mcc_weekly_xlsx(stream.getvalue(), "weekly.xlsx")

        self.assertEqual(len(activities), 2)
        self.assertEqual(activities[0]["quantity"], 12.5)
        self.assertEqual(activities[0]["line_cost"], 1500.0)

    def test_uses_smu_total_when_nightshift_hours_are_blank(self):
        content = workbook_bytes([[
            "Kristy Faram", datetime(2026, 8, 23, 5, 30), "Ironbark 1",
            "Multi Skilled Operator", "ARG-004 - SIS Drill Watercart Works", None,
            "Nightshift watercart hours", "WT01 - Hino FM500 13kL Watercart",
            1752.1, 1753.8, 1.7,
        ]])

        _, activities, _, crew, _ = parse_mcc_weekly_xlsx(content, "weekly.xlsx")

        self.assertEqual({row["quantity"] for row in activities}, {1.0, 1.7})
        self.assertEqual(sum(row["line_cost"] for row in activities), 411.0)
        self.assertEqual(crew[0]["hours"], "1.7")

    def test_equipment_uses_smu_and_day_hire_is_charged_once_per_day(self):
        rows = [
            [
                "Michael Chilby", datetime(2026, 8, 22, 5, 30), "Ironbark 1",
                "Supervisor", "ARG-005 - Exploration Civils & Support Works", 3,
                "Rehab pads", "LD04 - Caterpillar 432 Backhoe", 100, 102.4, 2.4,
            ],
            [
                "Michael Chilby", datetime(2026, 8, 22, 5, 30), "Ironbark 1",
                "Supervisor", "ARG-005 - Exploration Civils & Support Works", 2,
                "Empty first sump", "MCC39 - Vac Truck - PER DAY", 0, 1, 1,
            ],
            [
                "Michael Chilby", datetime(2026, 8, 22, 5, 30), "Ironbark 1",
                "Supervisor", "ARG-005 - Exploration Civils & Support Works", 1,
                "Empty second sump", "MCC39 - Vac Truck - PER DAY", 1, 2, 1,
            ],
        ]

        _, activities, _, _, _ = parse_mcc_weekly_xlsx(workbook_bytes(rows), "weekly.xlsx")

        equipment = [row for row in activities if "Charge type: Equipment" in row["notes"]]
        by_code = {row["code"]: row for row in equipment}
        self.assertEqual(by_code["MCC_BACKHOE"]["quantity"], 2.4)
        self.assertEqual(by_code["MCC_BACKHOE"]["line_cost"], 156.0)
        self.assertEqual(by_code["MCC_VAC_TRUCK"]["quantity"], 1)
        self.assertEqual(by_code["MCC_VAC_TRUCK"]["line_cost"], 810.0)
        self.assertEqual(sum(row["code"] == "MCC_VAC_TRUCK" for row in equipment), 1)
        self.assertEqual(by_code["MCC_LIGHT_VEHICLE"]["quantity"], 1)
        self.assertEqual(by_code["MCC_LIGHT_VEHICLE"]["line_cost"], 105.0)

    def test_adds_one_light_vehicle_for_each_program_workday(self):
        rows = [
            ["A", datetime(2026, 8, 17, 5, 30), "Ironbark 1", "Supervisor", "ARG-005 - Exploration Civils & Support Works", 3, "Work A", None, None, None, None],
            ["B", datetime(2026, 8, 17, 6, 0), "Ironbark 1", "Multi Skilled Operator", "ARG-005 - Exploration Civils & Support Works", 2, "Work B", None, None, None, None],
            ["C", datetime(2026, 8, 18, 5, 30), "Ironbark 1", "Supervisor", "ARG-005 - Exploration Civils & Support Works", 4, "Work C", None, None, None, None],
        ]

        _, activities, _, _, _ = parse_mcc_weekly_xlsx(workbook_bytes(rows), "weekly.xlsx")
        vehicles = [row for row in activities if row["code"] == "MCC_LIGHT_VEHICLE"]

        self.assertEqual(len(vehicles), 2)
        self.assertEqual({row["date"] for row in vehicles}, {"17/08/2026", "18/08/2026"})
        self.assertEqual(sum(row["line_cost"] for row in vehicles), 210.0)

    def test_audit_keeps_weekly_cost_as_authoritative(self):
        rows = [
            {
                "date": "20/08/2026",
                "code": "MCC_BODY_WATER_TRUCK",
                "quantity": 8,
                "line_cost": 640,
                "source_file": "daily.pdf",
                "source_file_type": "mcc_site_services_pdf",
                "rate_description": "Body Water Truck",
                "rate_unit": "hour",
            },
            {
                "date": "20/08/2026",
                "code": "MCC_BODY_WATER_TRUCK",
                "quantity": 7.5,
                "line_cost": 600,
                "source_file": "weekly.xlsx",
                "source_file_type": "mcc_weekly_xlsx",
                "rate_description": "Body Water Truck",
                "rate_unit": "hour",
            },
        ]

        audit = build_mcc_daily_weekly_audit(rows)

        self.assertEqual(audit["totals"]["weekly_cost"], 600.0)
        self.assertEqual(audit["totals"]["daily_estimate"], 640.0)
        self.assertEqual(audit["totals"]["cost_variance"], -40.0)
        self.assertEqual(audit["comparison"][0]["status"], "variance")

    def test_audit_applies_workgroup_allocation_to_quantity_and_cost(self):
        rows = [
            {
                "date": "20/08/2026",
                "code": "MCC_BODY_WATER_TRUCK",
                "quantity": 8,
                "line_cost": 640,
                "allocation_percent": 50,
                "source_file": "daily.pdf",
                "source_file_type": "mcc_site_services_pdf",
                "rate_description": "Body Water Truck",
                "rate_unit": "hour",
            },
            {
                "date": "20/08/2026",
                "code": "MCC_BODY_WATER_TRUCK",
                "quantity": 4,
                "line_cost": 320,
                "source_file": "weekly.xlsx",
                "source_file_type": "mcc_weekly_xlsx",
                "rate_description": "Body Water Truck",
                "rate_unit": "hour",
            },
        ]

        audit = build_mcc_daily_weekly_audit(rows)

        self.assertEqual(audit["comparison"][0]["daily_quantity"], 4.0)
        self.assertEqual(audit["comparison"][0]["daily_estimate"], 320.0)
        self.assertEqual(audit["comparison"][0]["status"], "match")
        self.assertEqual(audit["totals"]["cost_variance"], 0.0)

    def test_audit_accepts_light_vehicle_per_represented_day(self):
        rows = [{
            "date": "20/08/2026",
            "code": "MCC_LIGHT_VEHICLE",
            "quantity": 1,
            "line_cost": 105,
            "source_file": "weekly.xlsx",
            "source_file_type": "mcc_weekly_xlsx",
            "rate_description": "Light Vehicle",
            "rate_unit": "day",
        }, {
            "date": "20/08/2026",
            "code": "MCC_LIGHT_VEHICLE",
            "quantity": 1,
            "line_cost": 105,
            "source_file": "daily.pdf",
            "source_file_type": "mcc_site_services_pdf",
            "rate_description": "Light Vehicle",
            "rate_unit": "day",
        }]

        audit = build_mcc_daily_weekly_audit(rows)

        self.assertEqual(audit["comparison"][0]["status"], "accepted_rule")
        self.assertEqual(audit["totals"]["accepted"], 1)
        self.assertEqual(audit["totals"]["exceptions"], 0)
        self.assertEqual(audit["totals"]["daily_estimate"], 105.0)
        self.assertEqual(audit["totals"]["cost_variance"], 0.0)

    def test_audit_does_not_require_daily_reports_before_10_august(self):
        rows = [
            {
                "date": "09/08/2026",
                "code": "MCC_SUPERVISOR",
                "quantity": 5,
                "line_cost": 600,
                "source_file": "weekly-before.xlsx",
                "source_file_type": "mcc_weekly_xlsx",
                "rate_description": "Supervisor",
                "rate_unit": "hour",
            },
            {
                "date": "10/08/2026",
                "code": "MCC_SUPERVISOR",
                "quantity": 5,
                "line_cost": 600,
                "source_file": "weekly-after.xlsx",
                "source_file_type": "mcc_weekly_xlsx",
                "rate_description": "Supervisor",
                "rate_unit": "hour",
            },
        ]

        audit = build_mcc_daily_weekly_audit(rows)
        by_date = {row["date"]: row for row in audit["comparison"]}

        self.assertEqual(by_date["09/08/2026"]["status"], "not_required")
        self.assertEqual(by_date["10/08/2026"]["status"], "missing_daily")
        self.assertEqual(audit["totals"]["not_required"], 1)
        self.assertEqual(audit["totals"]["exceptions"], 1)
        self.assertEqual(audit["totals"]["daily_estimate"], 600.0)
        self.assertEqual(audit["totals"]["cost_variance"], 600.0)

    def test_audit_ignores_construction_trade_from_daily_comparison(self):
        rows = [
            {
                "date": "20/08/2026",
                "code": "MCC_CONSTRUCTION_TRADE",
                "quantity": 8,
                "line_cost": 800,
                "source_file": "daily.pdf",
                "source_file_type": "mcc_site_services_pdf",
                "rate_description": "Construction Trade",
                "rate_unit": "hour",
            },
            {
                "date": "20/08/2026",
                "code": "MCC_CONSTRUCTION_TRADE",
                "quantity": 10,
                "line_cost": 1000,
                "source_file": "weekly.xlsx",
                "source_file_type": "mcc_weekly_xlsx",
                "rate_description": "Construction Trade",
                "rate_unit": "hour",
            },
        ]

        audit = build_mcc_daily_weekly_audit(rows)

        self.assertEqual(audit["comparison"][0]["status"], "ignored_daily")
        self.assertEqual(audit["totals"]["ignored_daily"], 1)
        self.assertEqual(audit["totals"]["exceptions"], 0)
        self.assertEqual(audit["totals"]["daily_estimate"], 1000.0)
        self.assertEqual(audit["totals"]["cost_variance"], 0.0)


if __name__ == "__main__":
    unittest.main()
